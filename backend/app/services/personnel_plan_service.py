"""
Personnel Plan ("Personal Plan Data" tab) Service
Handles CRUD for personnel plan inputs and Excel import, mirrors
sales_plan_service.py / purchase_plan_service.py's shape.
"""
import io
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.personnel_plan import PersonnelPlan
import structlog

logger = structlog.get_logger()

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _num(v):
    """Blank template cells sometimes hold stray non-numeric placeholders —
    treat anything that isn't a real number as "no value yet"."""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _default_content():
    return {
        "meta": {"type": "", "department": ""},
        "headcount": {"year_prev": None, "year_curr": None, "rows": [], "total": None},
        "recruitment_permanent": {"year": None, "rows": [], "total": None},
        "recruitment_temporary": {"year": None, "rows": [], "total": None},
    }


class PersonnelPlanService:

    async def list_personnel_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
    ) -> dict:
        q = select(PersonnelPlan).order_by(PersonnelPlan.plan_year.desc(), PersonnelPlan.department)
        if plan_year:
            q = q.where(PersonnelPlan.plan_year == plan_year)
        if department:
            q = q.where(PersonnelPlan.department == department)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_personnel_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PersonnelPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_personnel_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(PersonnelPlan, plan_id)
        if not row:
            q = select(PersonnelPlan).where(
                PersonnelPlan.plan_year  == payload.get("plan_year"),
                PersonnelPlan.department == payload.get("department", ""),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()
        if row:
            row.content    = payload.get("content", row.content)
            row.status     = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = PersonnelPlan(
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department", ""),
                content    = payload.get("content", _default_content()),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def delete_personnel_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PersonnelPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(PersonnelPlan).where(PersonnelPlan.id == plan_id))
        return {"success": True, "message": f"Deleted personnel plan #{plan_id}"}

    def _find_row(self, ws, text, col=1, start=1):
        """First row (>= start) whose given column contains text (substring,
        case-insensitive) — used to locate sections by title rather than a
        fixed row number, so the parser tolerates the level list growing or
        shrinking, or blocks shifting up/down a row or two."""
        text = text.lower()
        for r in range(start, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v and text in str(v).lower():
                return r
        return None

    def _headcount_row(self, ws, r):
        return {
            "prev_permanent": _num(ws.cell(row=r, column=3).value),
            "prev_temporary": _num(ws.cell(row=r, column=4).value),
            "prev_total":     _num(ws.cell(row=r, column=5).value),
            "curr_permanent": _num(ws.cell(row=r, column=6).value),
            "curr_temporary": _num(ws.cell(row=r, column=7).value),
            "curr_total":     _num(ws.cell(row=r, column=8).value),
            "inc_permanent":  _num(ws.cell(row=r, column=9).value),
            "inc_temporary":  _num(ws.cell(row=r, column=10).value),
            "inc_total":      _num(ws.cell(row=r, column=11).value),
        }

    def _parse_headcount(self, ws) -> Optional[dict]:
        header_row = self._find_row(ws, "level", col=1)
        if header_row is None:
            return None
        year_prev = ws.cell(row=header_row, column=3).value  # C
        year_curr = ws.cell(row=header_row, column=6).value  # F
        # Two sub-header rows below (period label, then Permanent/Temporary/Total
        # labels) — data starts 3 rows below the "Level" header.
        rows = []
        total = None
        r = header_row + 3
        while r <= ws.max_row:
            level = ws.cell(row=r, column=1).value
            if level is None:
                break
            values = self._headcount_row(ws, r)
            if str(level).strip().lower() == "total":
                total = values
                break
            rows.append({"level": str(level).strip(), **values, "notes": ws.cell(row=r, column=12).value or ""})
            r += 1
        return {"year_prev": year_prev, "year_curr": year_curr, "rows": rows, "total": total}

    def _parse_recruitment(self, ws, section_title, after_row=1) -> Optional[dict]:
        section_row = self._find_row(ws, section_title, col=1, start=after_row)
        if section_row is None:
            return None
        header_row = self._find_row(ws, "level", col=1, start=section_row)
        if header_row is None:
            return None
        year = ws.cell(row=header_row, column=3).value  # C
        rows = []
        total = None
        r = header_row + 2  # month sub-header is +1, data starts +2
        while r <= ws.max_row:
            level = ws.cell(row=r, column=1).value
            if level is None:
                break
            months = [_num(ws.cell(row=r, column=c).value) for c in range(3, 15)]  # C-N
            total_val = _num(ws.cell(row=r, column=15).value)  # O
            if str(level).strip().lower() == "total":
                total = {"months": months, "total": total_val}
                break
            rows.append({"level": str(level).strip(), "months": months, "total": total_val})
            r += 1
        return {"year": year, "rows": rows, "total": total}, section_row

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "Personal plan template.xlsx"
        layout — meta at C6 (Type) / C9 (Department); headcount-by-level
        block, then "Recruitment Schedule - Permanent" and "- Temporary"
        blocks, each located by section title text rather than a fixed row
        number (tolerates the level list growing/shrinking). Sheet-name
        agnostic — matched by "[ H1 ]" in cell A1, same convention as the
        Purchase Plan template's "[ P1 ]"."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        target_ws = None
        for ws in wb.worksheets:
            if str(ws.cell(row=1, column=1).value or "").strip() == "[ H1 ]":
                target_ws = ws
                break
        if target_ws is None:
            return {"success": False, "error": "No recognizable data sheet found — expected a sheet with "
                                                 "'[ H1 ]' in cell A1, matching the Personal Plan template layout."}
        ws = target_ws

        meta = {
            "type":       ws.cell(row=6, column=3).value or "",
            "department": ws.cell(row=9, column=3).value or "",
        }

        headcount = self._parse_headcount(ws)
        rp_result = self._parse_recruitment(ws, "recruitment schedule - permanent")
        recruitment_permanent, rp_row = rp_result if rp_result else (None, 1)
        rt_result = self._parse_recruitment(ws, "recruitment schedule - temporary", after_row=(rp_row or 1) + 1)
        recruitment_temporary, _ = rt_result if rt_result else (None, None)

        if not headcount and not recruitment_permanent and not recruitment_temporary:
            return {"success": False, "error": "Sheet matched the template signature but no recognizable "
                                                 "data blocks (headcount / recruitment schedule) were found."}

        content = {
            "meta": meta,
            "headcount": headcount or _default_content()["headcount"],
            "recruitment_permanent": recruitment_permanent or _default_content()["recruitment_permanent"],
            "recruitment_temporary": recruitment_temporary or _default_content()["recruitment_temporary"],
        }
        payload = {
            "plan_year":  plan_year,
            "department": meta["department"],
            "content":    content,
            "status":     "draft",
        }
        result = await self.upsert_personnel_plan(db, payload, username)
        if result["success"]:
            result["rows_imported"] = {
                "headcount": len(content["headcount"]["rows"]),
                "recruitment_permanent": len(content["recruitment_permanent"]["rows"]),
                "recruitment_temporary": len(content["recruitment_temporary"]["rows"]),
            }
        return result

    def _to_dict(self, row: PersonnelPlan) -> dict:
        return {
            "id":          row.id,
            "plan_year":   row.plan_year,
            "department":  row.department,
            "content":     row.content,
            "status":      row.status,
            "created_by":  row.created_by,
            "created_at":  row.created_at.isoformat() if row.created_at else None,
            "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
        }
