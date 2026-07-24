"""
Investment Plan ("Investment Plan Data" tab) Service
Handles CRUD for investment plan inputs and Excel import, mirrors
sales_plan_service.py / manufacture_plan_service.py's shape (flat headers+rows).
"""
import io
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.investment_plan import InvestmentPlan
import structlog

logger = structlog.get_logger()

HEADERS = [
    "No", "Clarification", "Priority", "Item", "Purpose", "Picture", "QTY", "Lifetime (Year)",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Total", "Notes (Replacement/Additional)",
]


def _num(v):
    """Blank template cells sometimes hold stray non-numeric placeholders —
    treat anything that isn't a real number as "no value yet"."""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _default_content():
    return {"meta": {"type": "", "department": "", "team_code": "", "team_name": "", "year": None},
            "headers": HEADERS, "rows": []}


class InvestmentPlanService:

    async def list_investment_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
        team_code: Optional[str] = None,
    ) -> dict:
        q = select(InvestmentPlan).order_by(
            InvestmentPlan.plan_year.desc(), InvestmentPlan.department, InvestmentPlan.team_code
        )
        if plan_year:
            q = q.where(InvestmentPlan.plan_year == plan_year)
        if department:
            q = q.where(InvestmentPlan.department == department)
        if team_code:
            q = q.where(InvestmentPlan.team_code == team_code)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_investment_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(InvestmentPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_investment_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(InvestmentPlan, plan_id)
        if not row:
            q = select(InvestmentPlan).where(
                InvestmentPlan.plan_year  == payload.get("plan_year"),
                InvestmentPlan.department == payload.get("department", ""),
                InvestmentPlan.team_code  == str(payload.get("team_code", "")),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()
        if row:
            row.content    = payload.get("content", row.content)
            row.status     = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = InvestmentPlan(
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department", ""),
                team_code  = str(payload.get("team_code", "")),
                team_name  = payload.get("team_name", ""),
                content    = payload.get("content", _default_content()),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def delete_investment_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(InvestmentPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(InvestmentPlan).where(InvestmentPlan.id == plan_id))
        return {"success": True, "message": f"Deleted investment plan #{plan_id}"}

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "investment plan template.xlsx"
        layout — meta at C6 (Type) / C9 (Department) / C11 (Team Code/Name,
        a single combined field in this template — e.g. "All"); two-row
        header at 13-14; data from row 15, identified by a non-empty
        Clarification (column B) rather than a fixed row range, since blank
        template rows in between (no Clarification, only a leftover total
        formula) are skipped rather than treated as the end of the list —
        stops only at a "Total" row in column A. Sheet-name agnostic —
        matched by "[ I1 ]" in cell A1, same convention as the other plan
        templates."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        ws = None
        for sheet in wb.worksheets:
            if str(sheet.cell(row=1, column=1).value or "").strip() == "[ I1 ]":
                ws = sheet
                break
        if ws is None:
            return {"success": False, "error": "No recognizable data sheet found — expected a sheet with "
                                                 "'[ I1 ]' in cell A1, matching the Investment Plan template layout."}

        meta = {
            "type":       ws.cell(row=6,  column=3).value or "",
            "department": ws.cell(row=9,  column=3).value or "",
            "team_code":  ws.cell(row=11, column=3).value or "",
            "team_name":  "",
            "year":       ws.cell(row=13, column=9).value,
        }

        rows = []
        r = 15
        while r <= ws.max_row:
            no = ws.cell(row=r, column=1).value
            if no is not None and str(no).strip().lower() == "total":
                break
            clarification = ws.cell(row=r, column=2).value
            if clarification is None:
                # Blank template gap row (no category, just a leftover total
                # formula) — skip, don't stop; real rows may resume after it.
                r += 1
                continue
            rows.append([
                no,
                str(clarification),
                ws.cell(row=r, column=3).value or "",   # Priority
                ws.cell(row=r, column=4).value or "",   # Item
                ws.cell(row=r, column=5).value or "",   # Purpose
                ws.cell(row=r, column=6).value or "",   # Picture
                _num(ws.cell(row=r, column=7).value) or None,   # QTY
                _num(ws.cell(row=r, column=8).value) or None,   # Lifetime (Year)
                *[_num(ws.cell(row=r, column=c).value) for c in range(9, 21)],  # I-T Jan-Dec
                _num(ws.cell(row=r, column=21).value),   # U Total
                ws.cell(row=r, column=22).value or "",   # V Notes
            ])
            r += 1

        if not rows:
            return {"success": False, "error": "No data rows found starting at row 15 — file doesn't match the expected template"}

        content = {"meta": meta, "headers": HEADERS, "rows": rows}
        payload = {
            "plan_year":  plan_year,
            "department": meta["department"],
            "team_code":  str(meta["team_code"]),
            "team_name":  meta["team_name"],
            "content":    content,
            "status":     "draft",
        }
        result = await self.upsert_investment_plan(db, payload, username)
        if result["success"]:
            result["rows_imported"] = len(rows)
        return result

    def _to_dict(self, row: InvestmentPlan) -> dict:
        return {
            "id":          row.id,
            "plan_year":   row.plan_year,
            "department":  row.department,
            "team_code":   row.team_code,
            "team_name":   row.team_name,
            "content":     row.content,
            "status":      row.status,
            "created_by":  row.created_by,
            "created_at":  row.created_at.isoformat() if row.created_at else None,
            "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
        }
