"""
Sales Plan Service
Handles CRUD for sales plan inputs and Excel export (S1 Value / S2 Unit)
"""
import os
import json
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete, desc
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.sales_plan import SalesPlan, SalesPlanHistory
import structlog

logger = structlog.get_logger()

EXPORT_DIR = "/tmp/sales_plan_exports"
os.makedirs(EXPORT_DIR, exist_ok=True)


class SalesPlanService:

    async def list_sales_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
        team_code: Optional[str] = None,
        plan_type: Optional[str] = None,
    ) -> dict:
        q = select(SalesPlan).order_by(SalesPlan.plan_year.desc(), SalesPlan.department, SalesPlan.team_code)
        if plan_year:
            q = q.where(SalesPlan.plan_year == plan_year)
        if department:
            q = q.where(SalesPlan.department == department)
        if team_code:
            q = q.where(SalesPlan.team_code == team_code)
        if plan_type:
            q = q.where(SalesPlan.plan_type == plan_type)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_sales_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(SalesPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_sales_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(SalesPlan, plan_id)
        if not row:
            q = select(SalesPlan).where(
                SalesPlan.plan_year  == payload.get("plan_year"),
                SalesPlan.department == payload.get("department", ""),
                SalesPlan.team_code  == payload.get("team_code", ""),
                SalesPlan.plan_type  == payload.get("plan_type", "value"),
            )
            # A team can submit more than one plan under the same
            # department/team_code (e.g. Business Development's CMO and
            # Service Agreement plans both use team_code 62) — distinguished
            # by [ Type ]/[ Area ] meta, not by team_code. Only apply this
            # extra match when the incoming content actually carries a meta
            # dict (Excel import always does; the manual "New Plan" form
            # doesn't set one at all) so manual entries keep matching purely
            # on year/department/team_code/plan_type as before.
            meta = (payload.get("content") or {}).get("meta")
            if meta and "type" in meta:
                q = q.where(
                    SalesPlan.content["meta"]["type"].astext == str(meta.get("type") or ""),
                    SalesPlan.content["meta"]["area"].astext == str(meta.get("area") or ""),
                )
            result = await db.execute(q)
            row = result.scalars().first()
        if row:
            row.content   = payload.get("content", row.content)
            row.status    = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = SalesPlan(
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department", ""),
                team_code  = payload.get("team_code", ""),
                team_name  = payload.get("team_name", ""),
                plan_type  = payload.get("plan_type", "value"),
                content    = payload.get("content", {}),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str,
                            sheets: Optional[str] = None) -> dict:
        """Parse an uploaded Excel matching the "(S1) Sales plan_Value.xlsx"
        template family — sheet-agnostic, like every other PAC Simulation
        upload: any worksheet whose A1 is "[ S1 ]" is read as one plan (one
        sheet = one plan), everything else (e.g. an "Index_Team Code"
        reference sheet some exports carry as sheet 1) is skipped. Blindly
        reading wb.worksheets[0] used to be this file's behavior and broke
        on any workbook where the first sheet isn't the data sheet — it
        would silently parse a lookup table as sales figures and 500 trying
        to save a numeric Team Code as the text `department` column.

        `sheets` optionally narrows which tabs even get looked at — a
        1-based, in-tab-order spec like "1", "1-2", or "1,3,5-7" — for
        workbooks where only some sheets should be (re-)imported this run.
        None/blank means every sheet is considered, same as before this
        parameter existed.

        Two header layouts exist in the wild, detected per-sheet from
        row 14 col B:
          - "Local" layout (col B is anything other than "Country"):
            No, Product, Jan-Dec (D-O), Total Value (P), Total Unit (Q),
            Price (R) — one currency.
          - "Export/CMO" layout (col B == "Country"):
            No, Country, Customer, Product, Jan-Dec (E-P), Total Value (Q),
            Total Unit (R), Price USD (S), Price IDR (T) — used when each
            product line ships to a different country/customer and is
            priced in USD converted to IDR.
        Both normalize to the same stored row shape: [no, country, customer,
        product, jan..dec, total_value, total_unit, price_usd, price_idr]
        (20 items) — Local rows carry "" for country/customer/price_usd."""
        import io
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        try:
            allowed_sheet_nums = self._parse_sheet_range(sheets, len(wb.worksheets))
        except ValueError as e:
            return {"success": False, "error": str(e)}

        def _num(v):
            # Blank template cells sometimes hold a stray placeholder like
            # "\" instead of being truly empty — anything non-numeric just
            # means "no value yet", not a parse error.
            try:
                return int(v or 0)
            except (TypeError, ValueError):
                return 0

        headers = ["No", "Country", "Customer", "Product", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total Value", "Total Unit", "Price (USD)", "Price (IDR)"]

        imported = []
        for sheet_num, ws in enumerate(wb.worksheets, start=1):
            if allowed_sheet_nums is not None and sheet_num not in allowed_sheet_nums:
                continue
            if str(ws.cell(row=1, column=1).value or "").strip() != "[ S1 ]":
                continue

            meta = {
                "type":       ws.cell(row=6,  column=3).value or "",
                "area":       ws.cell(row=8,  column=3).value or "",
                "department": ws.cell(row=10, column=3).value or "",
                "team_code":  ws.cell(row=12, column=3).value or "",
                "team_name":  str(ws.cell(row=12, column=4).value or "").lstrip("/ ").strip(),
                # The Excel *tab* name, not any in-sheet cell — lets the
                # plan picker show which sheet a plan came from (e.g.
                # "Public"/"Private") when several sheets share the same
                # department/team_code and only differ by tab name.
                "sheet_name": ws.title,
            }

            is_export_layout = str(ws.cell(row=14, column=2).value or "").strip() == "Country"

            rows = []
            for r in range(16, ws.max_row + 1):
                no = ws.cell(row=r, column=1).value
                if is_export_layout:
                    country  = ws.cell(row=r, column=2).value
                    customer = ws.cell(row=r, column=3).value
                    product  = ws.cell(row=r, column=4).value
                    if no is None or product is None:
                        continue
                    months = [_num(ws.cell(row=r, column=c).value) for c in range(5, 17)]  # E-P
                    total_value = _num(ws.cell(row=r, column=17).value)
                    total_unit  = _num(ws.cell(row=r, column=18).value)
                    price_usd   = ws.cell(row=r, column=19).value
                    price_usd   = price_usd if isinstance(price_usd, (int, float)) else ""
                    price_idr   = ws.cell(row=r, column=20).value
                    price_idr   = price_idr if isinstance(price_idr, (int, float)) else ""
                    rows.append([no, str(country or ""), str(customer or ""), str(product),
                                 *months, total_value, total_unit, price_usd, price_idr])
                else:
                    product = ws.cell(row=r, column=2).value
                    if no is None or product is None:
                        continue
                    months = [_num(ws.cell(row=r, column=c).value) for c in range(4, 16)]  # D-O
                    total_value = _num(ws.cell(row=r, column=16).value)
                    total_unit  = _num(ws.cell(row=r, column=17).value)
                    price       = ws.cell(row=r, column=18).value
                    price       = price if isinstance(price, (int, float)) else ""
                    rows.append([no, "", "", str(product), *months, total_value, total_unit, "", price])

            if not rows:
                continue

            content = {"headers": headers, "rows": rows, "meta": meta}
            payload = {
                "plan_year":  plan_year,
                "department": meta["department"],
                "team_code":  str(meta["team_code"]),
                "team_name":  meta["team_name"],
                "plan_type":  "value",
                "content":    content,
                "status":     "draft",
            }
            result = await self.upsert_sales_plan(db, payload, username)
            if result["success"]:
                imported.append({"sheet": ws.title, "rows": len(rows), "id": result["data"]["id"]})

        if not imported:
            scope = f" in the selected sheet(s) ({sheets})" if allowed_sheet_nums is not None else ""
            return {"success": False, "error": f"No recognizable data sheets found{scope} — none match the Sales "
                                                 "Plan template layout (expected '[ S1 ]' in cell A1 of each data "
                                                 "sheet)."}
        return {"success": True, "imported": imported, "rows_imported": sum(x["rows"] for x in imported)}

    @staticmethod
    def _parse_sheet_range(spec: Optional[str], sheet_count: int) -> Optional[set[int]]:
        """'1' / '1-2' / '1,3,5-7' (1-based, in workbook tab order) -> the
        set of sheet numbers to consider; None (spec blank) means every
        sheet, preserving the pre-existing behavior. Raises ValueError with
        a user-facing message on a malformed or out-of-range spec, rather
        than silently processing zero (or the wrong) sheets."""
        spec = (spec or "").strip()
        if not spec:
            return None

        result: set[int] = set()
        for token in spec.split(","):
            token = token.strip()
            if not token:
                continue
            if "-" in token:
                a, _, b = token.partition("-")
                if not a.strip().isdigit() or not b.strip().isdigit():
                    raise ValueError(f"Invalid sheet range '{token}' — expected e.g. '1-3'.")
                start, end = int(a), int(b)
                if start < 1 or end < start:
                    raise ValueError(f"Invalid sheet range '{token}'.")
                result.update(range(start, end + 1))
            elif token.isdigit():
                result.add(int(token))
            else:
                raise ValueError(f"Invalid sheet number '{token}' — expected a number, e.g. '1' or '1-3'.")

        out_of_range = {n for n in result if n < 1 or n > sheet_count}
        if out_of_range:
            raise ValueError(
                f"Sheet number(s) {sorted(out_of_range)} out of range — this file only has {sheet_count} sheet(s)."
            )
        return result

    async def get_gross_sales_report_data(self, db: AsyncSession, plan_year: int) -> dict:
        """Flatten every Sales Plan (Value) product row for the given year
        into report lines for the Gross Sales Report export — Market comes
        from each plan's [ Type ] meta, Customer from [ Area ], matching how
        output_grossales2026.xlsx's Market/Customer columns are populated.
        The S1 template's Jan-Dec cells hold monthly Sales VALUE (Rp), not
        quantity — confirmed against a real dev record where row[14] (Total
        Value) and row[15] (Total Unit) exactly match output_grossales2026's
        Sales Amount Total / Sales Quantity Total for the same product, and
        row[Jan]/price divides out to the reference file's Sales Quantity
        Jan exactly. So Amount is the real entered data; Quantity is derived
        (Amount / Price) — the reverse of a plain qty*price sheet."""
        q = select(SalesPlan).where(
            SalesPlan.plan_year == plan_year,
            SalesPlan.plan_type == "value",
        ).order_by(SalesPlan.department, SalesPlan.team_code)
        result = await db.execute(q)
        plans = result.scalars().all()

        lines = []
        for plan in plans:
            content = plan.content or {}
            meta = content.get("meta", {})
            market = meta.get("type", "") or ""
            customer = meta.get("area", "") or ""
            if not market:
                # A blank [ Type ] marks a pre-aggregated rollup plan (e.g.
                # area="Total") rather than a real market/customer segment —
                # its monthly cells hold a second-order sum across other
                # plans, so mixing it into the line-item report would
                # double-count every product.
                continue
            for row in content.get("rows", []):
                if len(row) < 20:
                    continue
                country = row[1]
                customer_name = row[2]
                product = row[3]
                amounts = [v if isinstance(v, (int, float)) else 0 for v in row[4:16]]
                # Price (IDR) is what the report deals in (Rp) — Price (USD)
                # is only meaningful alongside an exchange rate, out of scope
                # for this report.
                price = row[19] if isinstance(row[19], (int, float)) else 0
                lines.append({
                    "market": market,
                    "customer": customer,
                    "country": str(country or ""),
                    "customer_name": str(customer_name or ""),
                    "product": str(product or ""),
                    "amounts": amounts,
                    "price": price,
                })

        if not lines:
            return {"success": False, "error": f"Tidak ada Sales Plan Data (Value) untuk tahun {plan_year}."}
        return {"success": True, "data": lines, "plan_year": plan_year}

    async def delete_sales_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(SalesPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(SalesPlan).where(SalesPlan.id == plan_id))
        return {"success": True, "message": f"Deleted sales plan #{plan_id}"}

    async def export_excel(self, db: AsyncSession, plan_id: int, plan_type: str, username: str) -> dict:
        row = await db.get(SalesPlan, plan_id)
        if not row:
            return {"success": False, "error": "Sales plan not found"}
        if row.plan_type != plan_type:
            return {"success": False, "error": f"Plan type mismatch: {row.plan_type} vs {plan_type}"}
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return {"success": False, "error": "openpyxl not installed"}

        content = row.content or {}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"S{1 if plan_type == 'value' else 2} Sales Plan"

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        sub_fill    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sub_font    = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        ws.cell(row=1, column=1, value=f"PT CKD OTTO Pharmaceuticals — Sales Plan {plan_type.title()} {row.plan_year}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value=f"Department: {row.department}  |  Team: {row.team_code} / {row.team_name}")
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)

        rows_data = content.get("rows", [])
        headers = content.get("headers", [
            "No", "Product / Description", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"
        ])

        # Header row
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for r_idx, row_data in enumerate(rows_data, 5):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx > 2 else "left", vertical="center")
                if c_idx > 2 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)

        filename = f"(S{1 if plan_type == 'value' else 2}) Sales Plan_{plan_type.title()}_{row.plan_year}_{row.team_code or 'ALL'}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        wb.save(filepath)

        history = SalesPlanHistory(
            sales_plan_id = row.id,
            plan_type     = plan_type,
            filename      = filename,
            file_path     = filepath,
            generated_by  = username,
        )
        db.add(history)
        await db.commit()
        await db.refresh(history)

        return {
            "success": True,
            "filename": filename,
            "file_path": filepath,
            "history_id": history.id,
        }

    def _to_dict(self, row: SalesPlan) -> dict:
        return {
            "id":         row.id,
            "plan_year":  row.plan_year,
            "department": row.department,
            "team_code":  row.team_code,
            "team_name":  row.team_name,
            "plan_type":  row.plan_type,
            "content":    row.content,
            "status":     row.status,
            "created_by": row.created_by,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
