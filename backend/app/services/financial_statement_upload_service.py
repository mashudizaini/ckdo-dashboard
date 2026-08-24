"""
Financial Statement — Excel Upload Source
─────────────────────────────────────────
Alternative to the live Oracle EBS GL_BALANCES query
(financial_statement_service.py), for the transition period where some
reports are still produced manually in Excel. Parses the same 3 sheets the
whole Financial Statement module was originally modeled after —
sumber/FS_CKD OTTO 2015-2026_sent.xlsx — and stores the result as JSON,
one row per report_type, so the API layer can serve either source through
the same response shape the frontend already consumes.

Sheet -> report_type:
  "Balance sheet"    -> balance_sheet
  "Profit or loss"   -> profit_loss       (annual, FY columns)
  "PL_monthly"       -> profit_loss_monthly (single MTD/YTD this-vs-last-year snapshot)
  "Cashflow"         -> cash_flow         (annual, bare-year columns — see parse_cash_flow_excel)

Label-matching notes (verified against the actual reference file — NOT
assumed):
  - Balance Sheet's line-item labels match the Oracle-mode bucket labels
    (BS_ASSET_CURRENT etc.) exactly, so those fixed buckets are reused for
    ordering/validation.
  - Profit or loss's NET SALES/COGS breakdown is per-customer in the manual
    file vs. per-channel in Oracle-mode — and EXPENSES also isn't a clean
    1:1 match (e.g. the file has "STORAGE" where Oracle-mode's fixed list
    has "CONFERENCE & CONVENTION"). Rather than force either sheet's lines
    into Oracle's bucket labels (which would silently misattribute rows),
    every section's line items are kept exactly as found in the file, in
    file order — only the section header / TOTAL row labels are matched
    literally (verified 1:1 against the file), since those drive the
    growth-rate chart and the table's bold TOTAL rows.
  - PL_monthly uses its own distinct TOTAL-row wording (e.g. "TOTAL INCOME
    TAX", "NET PROFIT (LOSS)") — different from both Balance Sheet's and
    the annual Profit or loss sheet's TOTAL labels — matching the labels
    ProfitLossMonthlyPanel already renders in the frontend.
"""
import io
import re
from datetime import datetime

import openpyxl

from app.services.financial_statement_service import (
    BS_ASSET_CURRENT, BS_ASSET_NONCURRENT, BS_LIAB_CURRENT, BS_LIAB_NONCURRENT, BS_EQUITY_ORDER,
)

_MONTH_RE = re.compile(r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\b", re.I)


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


# Known typo/spacing variant -> canonical label, found in the actual
# reference file (not a hypothetical).
_ALIASES = {
    "DEFFERED TAX INCOME ( EXPENSE)": "DEFERRED TAX INCOME (EXPENSE)",
}


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _sheet(wb, name: str):
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return wb[s]
    raise ValueError(f"Sheet '{name}' not found in the file — available sheets: {', '.join(wb.sheetnames)}")


def _row_label(ws, row: int, max_col: int = 6):
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip():
            label = _norm(v)
            return _ALIASES.get(label, label)
    return None


def _find_row(ws, label: str, max_row: int, start_row: int = 1) -> int | None:
    target = _norm(label)
    for r in range(start_row, max_row + 1):
        if _row_label(ws, r) == target:
            return r
    return None


def _lines_between(ws, header_row: int, total_row: int, value_cols: list[int]) -> list[dict]:
    """Every non-empty labeled row strictly between a section header and
    its TOTAL row, kept in file order with the file's own label — see
    module docstring for why this isn't matched against Oracle's bucket
    labels."""
    lines = []
    for r in range(header_row + 1, total_row):
        label = _row_label(ws, r)
        if not label:
            continue
        values = [_num(ws.cell(row=r, column=c).value) for c in value_cols]
        if all(v == 0 for v in values):
            continue
        lines.append({"label": label, "values": values})
    return lines


def _row_values(ws, row: int, value_cols: list[int]) -> list[float]:
    return [_num(ws.cell(row=row, column=c).value) for c in value_cols]


# ── Balance Sheet ────────────────────────────────────────────────────────

_BS_TOTAL_LABELS = {
    "current_assets":       "TOTAL CURRENT ASSETS",
    "noncurrent_assets":    "TOTAL NON CURRENT ASSETS",
    "assets":               "TOTAL ASSETS",
    "current_liabilities":  "TOTAL CURRENT LIABILITIES",
    "noncurrent_liabilities": "TOTAL NONCURRENT LIABILITIES",
    "liabilities":          "TOTAL LIABILITIES",
    "equity":               "TOTAL EQUITY",
    "liabilities_and_equity": "TOTAL LIABILITIES AND EQUITY",
}


def parse_balance_sheet_excel(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = _sheet(wb, "Balance sheet")

    year_cols: dict[int, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=6, column=c).value
        if isinstance(v, str) and re.match(r"^FY\s*\d{4}$", v.strip(), re.I):
            year_cols[int(re.search(r"\d{4}", v).group())] = c
    if not year_cols:
        raise ValueError("Year header (\"FY YYYY\") not found in row 6 of the 'Balance sheet' sheet.")
    years = sorted(year_cols)
    value_cols = [year_cols[y] for y in years]

    as_of_label = None
    for r in range(1, 6):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _MONTH_RE.search(v):
            as_of_label = v.strip()
            break

    label_values: dict[str, list[float]] = {}
    for r in range(7, ws.max_row + 1):
        label = _row_label(ws, r)
        if not label:
            continue
        values = _row_values(ws, r, value_cols)
        if all(v == 0 for v in values):
            continue
        label_values[label] = values

    zeros = [0.0] * len(years)

    def section(labels):
        return [{"label": l, "values": label_values.get(_norm(l), zeros)} for l in labels]

    def total(key, computed):
        return label_values.get(_norm(_BS_TOTAL_LABELS[key]), computed)

    current_assets = section(BS_ASSET_CURRENT)
    noncurrent_assets = section(BS_ASSET_NONCURRENT)
    current_liab = section(BS_LIAB_CURRENT)
    noncurrent_liab = section(BS_LIAB_NONCURRENT)
    equity = section(BS_EQUITY_ORDER)

    def sum_rows(rows_):
        return [sum(row["values"][i] for row in rows_) for i in range(len(years))]

    total_current_assets = total("current_assets", sum_rows(current_assets))
    total_noncurrent_assets = total("noncurrent_assets", sum_rows(noncurrent_assets))
    total_assets = total("assets", [a + b for a, b in zip(total_current_assets, total_noncurrent_assets)])
    total_current_liab = total("current_liabilities", sum_rows(current_liab))
    total_noncurrent_liab = total("noncurrent_liabilities", sum_rows(noncurrent_liab))
    total_liabilities = total("liabilities", [a + b for a, b in zip(total_current_liab, total_noncurrent_liab)])
    total_equity = total("equity", sum_rows(equity))
    total_liab_equity = total("liabilities_and_equity", [a + b for a, b in zip(total_liabilities, total_equity)])

    known = {_norm(l) for l in (BS_ASSET_CURRENT + BS_ASSET_NONCURRENT + BS_LIAB_CURRENT
                                 + BS_LIAB_NONCURRENT + BS_EQUITY_ORDER)} | {_norm(v) for v in _BS_TOTAL_LABELS.values()}
    unmapped = sorted(l for l in label_values if l not in known)

    return {
        "years": years, "as_of_label": as_of_label,
        "current_assets": current_assets, "total_current_assets": total_current_assets,
        "noncurrent_assets": noncurrent_assets, "total_noncurrent_assets": total_noncurrent_assets,
        "total_assets": total_assets,
        "current_liabilities": current_liab, "total_current_liabilities": total_current_liab,
        "noncurrent_liabilities": noncurrent_liab, "total_noncurrent_liabilities": total_noncurrent_liab,
        "total_liabilities": total_liabilities,
        "equity": equity, "total_equity": total_equity,
        "total_liabilities_and_equity": total_liab_equity,
        "check_diff": [a - b for a, b in zip(total_liab_equity, total_assets)],
        "unmapped_accounts": unmapped,
    }


# ── Profit or Loss (annual) ──────────────────────────────────────────────

def _parse_pl_sheet(ws, value_cols: list[int], totals: dict) -> dict:
    """Shared section-scanner for the annual 'Profit or loss' sheet and the
    'PL_monthly' sheet — same section structure, different TOTAL-row
    wording (passed in via `totals`) and different value columns."""
    max_row = ws.max_row

    def find(label):
        r = _find_row(ws, label, max_row)
        if r is None:
            raise ValueError(f"Row '{label}' not found in the sheet.")
        return r

    r_net_sales = find("NET SALES")
    r_total_net_sales = find(totals["total_net_sales"])
    r_cogs = find("COGS")
    r_total_cogs = find(totals["total_cogs"])
    r_gross_profit = find(totals["gross_profit"])
    r_expenses = find("EXPENSES")
    r_total_expenses = find(totals["total_expenses"])
    r_other = find("OTHER INCOME / EXPENSES")
    r_total_other = find(totals["total_other"])
    r_pbt = find(totals["profit_before_tax"])
    r_tax = find("INCOME TAX")
    r_total_tax = find(totals["total_tax"])
    r_pat = find(totals["profit_after_tax"])
    r_oci = find(totals["oci"])
    r_total_comprehensive = find(totals["total_comprehensive"])

    sales_lines = _lines_between(ws, r_net_sales, r_total_net_sales, value_cols)
    cogs_lines = _lines_between(ws, r_cogs, r_total_cogs, value_cols)
    expense_lines = _lines_between(ws, r_expenses, r_total_expenses, value_cols)
    other_lines = _lines_between(ws, r_other, r_total_other, value_cols)
    tax_lines = _lines_between(ws, r_tax, r_total_tax, value_cols)

    return {
        "sales_lines": sales_lines, "contra_lines": [], "total_net_sales": _row_values(ws, r_total_net_sales, value_cols),
        "cogs_lines": cogs_lines, "total_cogs": _row_values(ws, r_total_cogs, value_cols),
        "gross_profit": _row_values(ws, r_gross_profit, value_cols),
        "expense_lines": expense_lines, "total_expenses": _row_values(ws, r_total_expenses, value_cols),
        "other_lines": other_lines, "total_other": _row_values(ws, r_total_other, value_cols),
        "profit_before_tax": _row_values(ws, r_pbt, value_cols),
        "tax_lines": tax_lines, "total_tax": _row_values(ws, r_total_tax, value_cols),
        "profit_after_tax": _row_values(ws, r_pat, value_cols),
        "oci": _row_values(ws, r_oci, value_cols),
        "total_comprehensive": _row_values(ws, r_total_comprehensive, value_cols),
        "unmapped_accounts": [],
    }


_PL_ANNUAL_TOTALS = {
    "total_net_sales": "TOTAL NET SALES", "total_cogs": "TOTAL COGS", "gross_profit": "GROSS PROFIT",
    "total_expenses": "TOTAL EXPENSES", "total_other": "TOTAL OTHER INCOME (EXPENSES)",
    "profit_before_tax": "PROFIT (LOSS) BEFORE TAX", "total_tax": "TOTAL INCOME TAX BENEFIT (EXPENSE)",
    "profit_after_tax": "PROFIT (LOSS) AFTER TAX", "oci": "OTHER COMPREHENSIVE INCOME",
    "total_comprehensive": "TOTAL COMPREHENSIVE INCOME (LOSS) FOR THE YEAR",
}

_PL_MONTHLY_TOTALS = {
    "total_net_sales": "TOTAL NET SALES", "total_cogs": "TOTAL COGS", "gross_profit": "GROSS PROFIT",
    "total_expenses": "TOTAL EXPENSES", "total_other": "TOTAL OTHER INCOME (EXPENSE)",
    "profit_before_tax": "PROFIT (LOSS) BEFORE INCOME TAX", "total_tax": "TOTAL INCOME TAX",
    "profit_after_tax": "NET PROFIT (LOSS)", "oci": "OTHER COMPREHENSIVE INCOME (LOSS)",
    "total_comprehensive": "TOTAL COMPREHENSIVE INCOME (LOSS) FOR THE YEAR",
}


def parse_profit_loss_excel(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = _sheet(wb, "Profit or loss")

    year_cols: dict[int, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=6, column=c).value
        if isinstance(v, str) and re.match(r"^FY\s*\d{4}$", v.strip(), re.I):
            year_cols[int(re.search(r"\d{4}", v).group())] = c
    if not year_cols:
        raise ValueError("Year header (\"FY YYYY\") not found in row 6 of the 'Profit or loss' sheet.")
    years = sorted(year_cols)
    value_cols = [year_cols[y] for y in years]

    data = _parse_pl_sheet(ws, value_cols, _PL_ANNUAL_TOTALS)
    data["years"] = years
    data["columns"] = [f"FY {y}" for y in years]
    return data


# ── Profit or Loss Monthly ───────────────────────────────────────────────

def parse_profit_loss_monthly_excel(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = _sheet(wb, "PL_monthly")

    # Row 7: date labels spanning each year's MTD/YTD pair (e.g. "June 30,
    # 2025" over cols G:H, "June 30, 2026" over cols I:J). Row 8 confirms
    # the MTD/YTD sub-labels at those same 4 columns.
    date_cells = [(c, ws.cell(row=7, column=c).value) for c in range(1, ws.max_column + 1)
                  if isinstance(ws.cell(row=7, column=c).value, str) and ws.cell(row=7, column=c).value.strip()]
    if len(date_cells) < 2:
        raise ValueError("Date header (row 7) not found in the 'PL_monthly' sheet.")
    (col_last, date_last), (col_this, date_this) = date_cells[0], date_cells[1]
    value_cols = [col_last, col_last + 1, col_this, col_this + 1]  # MTD Last, YTD Last, MTD This, YTD This

    data = _parse_pl_sheet(ws, value_cols, _PL_MONTHLY_TOTALS)
    data["date_last"] = date_last.strip()
    data["date_this"] = date_this.strip()
    data["columns"] = ["MTD Last Year", "YTD Last Year", "MTD This Year", "YTD This Year"]

    # Which calendar month/year this snapshot represents — lets
    # save_upload() store it as its own row instead of overwriting
    # whatever month was uploaded before, so a Month+Year picker in the
    # UI can select between multiple stored snapshots.
    try:
        dt_this = datetime.strptime(data["date_this"], "%B %d, %Y")
        data["period_month"] = dt_this.month
        data["period_year"] = dt_this.year
    except ValueError:
        data["period_month"] = None
        data["period_year"] = None
    return data


# ── Cash Flow ─────────────────────────────────────────────────────────────
# Unlike the other 2 sheets, "Cashflow" has NO fixed/known line-item list to
# validate against — every label (Operating/Investing/Financing subtotals,
# and the many free-form line items under them) is carried through exactly
# as found, in file order, with its indentation level preserved via which
# column (B/C/D) the label sits in. There's also no live Oracle equivalent
# for this report at all (a statutory cash flow isn't a direct GL_BALANCES
# query — it's manually prepared/derived each period), so this is the only
# source, permanently, not a transitional one.
#
# Row 6 headers are bare years (2015, not "FY 2015"). The current/in-
# progress year is NOT a single column like the other 2 sheets — it's
# broken into one column per posted month (row 8: Jan, Feb, ...), matching
# how far the period has actually been closed. A single "annual" figure for
# that year is synthesized per row: Beginning Balance takes the first
# month's value, Ending Balance takes the last month's value, everything
# else (Cash In/Out and every line item) sums across the posted months.

_MONTH_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def _is_month_label(v) -> bool:
    return isinstance(v, str) and v.strip().upper()[:3] in _MONTH_ABBR


def parse_cash_flow_excel(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = _sheet(wb, "Cashflow")

    year_col: dict[int, int] = {}
    month_cols: list[int] = []
    for c in range(1, ws.max_column + 1):
        yv = ws.cell(row=6, column=c).value
        if isinstance(yv, (int, float)) and 2000 <= int(yv) <= 2100:
            year_col[int(yv)] = c
        if _is_month_label(ws.cell(row=8, column=c).value):
            month_cols.append(c)
    if not year_col:
        raise ValueError("Year header not found in row 6 of the 'Cashflow' sheet.")

    # The in-progress year is whichever row-6 year sits at/before the first
    # month column (the sheet reuses that same column for both the year
    # label and the first month's data).
    partial_year = None
    if month_cols:
        candidates = [y for y, c in year_col.items() if c <= month_cols[0]]
        if candidates:
            partial_year = max(candidates)

    years = sorted(year_col)

    as_of_label = None
    for r in range(1, 6):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _MONTH_RE.search(v):
            as_of_label = v.strip()
            break

    def row_values(row: int, label_upper: str) -> list[float]:
        out = []
        for y in years:
            if y == partial_year and month_cols:
                if "BEGINNING BALANCE" in label_upper:
                    out.append(_num(ws.cell(row=row, column=month_cols[0]).value))
                elif "ENDING BALANCE" in label_upper:
                    out.append(_num(ws.cell(row=row, column=month_cols[-1]).value))
                else:
                    out.append(sum(_num(ws.cell(row=row, column=c).value) for c in month_cols))
            else:
                out.append(_num(ws.cell(row=row, column=year_col[y]).value))
        return out

    rows_out = []
    monthly_ending_balance = []
    for r in range(9, ws.max_row + 1):
        label, level = None, None
        for lvl, col in enumerate((2, 3, 4)):  # B=level0 (trunk), C=level1 (Operating/Investing/Financing), D=level2 (line items)
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.strip():
                label, level = re.sub(r"\s+", " ", v.strip()), lvl
                break
        if not label:
            continue
        values = row_values(r, label.upper())
        if level == 2 and all(v == 0 for v in values):
            continue  # thin/blank line item for this year range — same convention as _lines_between
        rows_out.append({"label": label, "level": level, "type": "total" if level < 2 else "line", "values": values})

        # The trunk "Ending Balance" row already holds a running balance in
        # every posted month column (row_values() above only ever reads the
        # LAST one, to synthesize a single year-end figure) — capture the
        # per-month series too, so a monthly Actual trend (e.g. EIS
        # Administration's Cashflow chart) can be built without re-deriving
        # it from Cash In/Out.
        if level == 0 and "ENDING BALANCE" in label.upper() and partial_year and month_cols:
            for c in month_cols:
                month_label = str(ws.cell(row=8, column=c).value or "").strip().upper()[:3]
                if month_label not in _MONTH_ABBR:
                    continue
                monthly_ending_balance.append({
                    "month": _MONTH_ABBR.index(month_label) + 1,
                    "value": _num(ws.cell(row=r, column=c).value),
                })

    return {
        "years": years,
        "as_of_label": as_of_label,
        "partial_year": partial_year,
        "rows": rows_out,
        "monthly_ending_balance": monthly_ending_balance,
    }


_PARSERS = {
    "balance_sheet": parse_balance_sheet_excel,
    "profit_loss": parse_profit_loss_excel,
    "profit_loss_monthly": parse_profit_loss_monthly_excel,
    "cash_flow": parse_cash_flow_excel,
}


class FinancialStatementUploadService:

    async def save_upload(self, db, report_type: str, content: bytes, filename: str, username: str) -> dict:
        from app.models.financial_statement_upload import FinancialStatementUpload
        from sqlalchemy import select

        parser = _PARSERS.get(report_type)
        if not parser:
            return {"success": False, "error": f"Unknown report type: {report_type}"}
        try:
            parsed = parser(content)
        except Exception as e:
            return {"success": False, "error": str(e)}

        # profit_loss_monthly is keyed by (report_type, period_month,
        # period_year) — a new month's upload adds a new row instead of
        # overwriting a prior month's. Every other report_type keeps the
        # original single-row-per-type behavior (period_month IS NULL).
        period_month = period_year = None
        if report_type == "profit_loss_monthly":
            period_month, period_year = parsed.get("period_month"), parsed.get("period_year")
            if period_month is None or period_year is None:
                return {"success": False, "error": (
                    "Could not determine the month/year this snapshot represents from row 7's "
                    "date header — expected a format like \"June 30, 2026\"."
                )}

        month_filter = FinancialStatementUpload.period_month == period_month if period_month is not None \
            else FinancialStatementUpload.period_month.is_(None)
        year_filter = FinancialStatementUpload.period_year == period_year if period_year is not None \
            else FinancialStatementUpload.period_year.is_(None)
        query = (
            select(FinancialStatementUpload)
            .where(FinancialStatementUpload.report_type == report_type, month_filter, year_filter)
        )
        result = await db.execute(query)
        row = result.scalar_one_or_none()
        if row:
            row.content = parsed
            row.original_filename = filename
            row.uploaded_by = username
            row.uploaded_at = datetime.utcnow()
        else:
            row = FinancialStatementUpload(
                report_type=report_type, content=parsed,
                period_month=period_month, period_year=period_year,
                original_filename=filename, uploaded_by=username,
            )
            db.add(row)
        await db.commit()
        return {"success": True, "data": parsed, "uploaded_at": row.uploaded_at.isoformat() if row.uploaded_at else None}

    async def get_upload(self, db, report_type: str, month: int | None = None, year: int | None = None) -> dict | None:
        from app.models.financial_statement_upload import FinancialStatementUpload
        from sqlalchemy import select

        query = select(FinancialStatementUpload).where(FinancialStatementUpload.report_type == report_type)
        if report_type == "profit_loss_monthly":
            if month is not None and year is not None:
                query = query.where(FinancialStatementUpload.period_month == month, FinancialStatementUpload.period_year == year)
            else:
                # No month/year given — fall back to the most recently
                # uploaded snapshot (e.g. for /upload-status's overview).
                # nullslast() matters here: Postgres's default for DESC is
                # NULLS FIRST, which would otherwise put a legacy
                # pre-migration row (period_month/year still NULL) ahead
                # of every properly-dated snapshot.
                query = query.order_by(
                    FinancialStatementUpload.period_year.desc().nullslast(),
                    FinancialStatementUpload.period_month.desc().nullslast(),
                )
        result = await db.execute(query.limit(1))
        row = result.scalar_one_or_none()
        if not row:
            return None
        return {
            "content": row.content,
            "original_filename": row.original_filename,
            "uploaded_by": row.uploaded_by,
            "uploaded_at": row.uploaded_at.isoformat() if row.uploaded_at else None,
            "period_month": row.period_month,
            "period_year": row.period_year,
        }

    async def list_snapshots(self, db, report_type: str) -> list[dict]:
        """profit_loss_monthly only — every stored (month, year) snapshot,
        newest first, for the frontend's Month+Year picker."""
        from app.models.financial_statement_upload import FinancialStatementUpload
        from sqlalchemy import select

        result = await db.execute(
            select(FinancialStatementUpload.period_month, FinancialStatementUpload.period_year, FinancialStatementUpload.uploaded_at)
            .where(FinancialStatementUpload.report_type == report_type)
            .where(FinancialStatementUpload.period_month.isnot(None))
            .order_by(FinancialStatementUpload.period_year.desc(), FinancialStatementUpload.period_month.desc())
        )
        return [
            {"month": m, "year": y, "uploaded_at": u.isoformat() if u else None}
            for m, y, u in result.fetchall()
        ]
