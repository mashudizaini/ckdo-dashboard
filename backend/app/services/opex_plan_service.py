"""
OPEX Plan ("OPEX Plan Data" tab) Service
Handles CRUD for OPEX plan inputs and Excel import, mirrors
purchase_plan_service.py's multi-sheet-per-upload shape (one plan per
Department/Team data sheet) combined with manufacture_plan_service.py /
investment_plan_service.py's flat headers+rows content shape.
"""
import io
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.opex_plan import OpexPlan
import structlog

logger = structlog.get_logger()

HEADERS = [
    "No", "Managerial Account No", "Managerial Account Name",
    "Chart of Account No", "Chart of Account Name", "Controll/Uncontrolled",
    "Sales & Mkt", "Strategy Development", "Plant", "Admin",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Total",
]


def _num(v):
    """Blank template cells sometimes hold stray non-numeric placeholders —
    treat anything that isn't a real number as "no value yet"."""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _default_content():
    return {"meta": {"type": "", "department": "", "team_code": "", "team_name": ""},
            "headers": HEADERS, "rows": []}


class OpexPlanService:

    async def list_opex_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
        team_code: Optional[str] = None,
    ) -> dict:
        q = select(OpexPlan).order_by(
            OpexPlan.plan_year.desc(), OpexPlan.department, OpexPlan.team_code
        )
        if plan_year:
            q = q.where(OpexPlan.plan_year == plan_year)
        if department:
            q = q.where(OpexPlan.department == department)
        if team_code:
            q = q.where(OpexPlan.team_code == team_code)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_opex_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(OpexPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_opex_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(OpexPlan, plan_id)
        if not row:
            q = select(OpexPlan).where(
                OpexPlan.plan_year  == payload.get("plan_year"),
                OpexPlan.department == payload.get("department", ""),
                OpexPlan.team_code  == str(payload.get("team_code", "")),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()
        if row:
            row.content    = payload.get("content", row.content)
            row.status     = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = OpexPlan(
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department", ""),
                team_code  = str(payload.get("team_code", "")),
                team_name  = payload.get("team_name", ""),
                content    = payload.get("content", _default_content()),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def delete_opex_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(OpexPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(OpexPlan).where(OpexPlan.id == plan_id))
        return {"success": True, "message": f"Deleted OPEX plan #{plan_id}"}

    def _read_sheet_rows(self, ws) -> list:
        """Each budget line is a Chart-of-Account row (col D/E) grouped under
        a Managerial Account (col B/C) whose No/Code/Name are only populated
        on the group's first row (merged-cell style in the template) — carry
        the last seen No/Managerial Account No/Name forward across the
        group's continuation rows. Header at 13-14, data from row 15, ends
        at a "Total" row in column A. Gap rows (no Chart of Account at all)
        are skipped rather than treated as the end of the list."""
        rows = []
        r = 15
        last_no = last_ma_no = last_ma_name = None
        while r <= ws.max_row:
            no = ws.cell(row=r, column=1).value
            if no is not None and str(no).strip().lower() == "total":
                break
            coa_no = ws.cell(row=r, column=4).value
            coa_name = ws.cell(row=r, column=5).value
            if coa_no is None and coa_name is None:
                r += 1
                continue
            if no is not None:
                last_no = no
            ma_no = ws.cell(row=r, column=2).value
            if ma_no is not None:
                last_ma_no = ma_no
            ma_name = ws.cell(row=r, column=3).value
            if ma_name is not None and str(ma_name).strip() != "":
                last_ma_name = ma_name
            rows.append([
                last_no, last_ma_no, last_ma_name,
                coa_no, str(coa_name or ""),
                ws.cell(row=r, column=6).value or "",    # Controll/Uncontrolled
                ws.cell(row=r, column=7).value or "",    # Sales & Mkt
                ws.cell(row=r, column=8).value or "",    # Strategy Development
                ws.cell(row=r, column=9).value or "",    # Plant
                ws.cell(row=r, column=10).value or "",   # Admin
                *[_num(ws.cell(row=r, column=c).value) for c in range(11, 23)],  # K-V Jan-Dec
                _num(ws.cell(row=r, column=23).value),   # W Total
            ])
            r += 1
        return rows

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "(O1) OPEX Plan_Summary_Department.xlsx"
        template — one plan per data sheet (Department / Department_Team...),
        each with its own Type/Department/Team meta and budget-line list.
        Sheet-name agnostic — matched by "[ O1 ]" in cell A1, same convention
        as the other plan templates. Reference/index sheets (e.g. "Index_Team
        Code") don't have this and are skipped automatically."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        imported = []
        for ws in wb.worksheets:
            if str(ws.cell(row=1, column=1).value or "").strip() != "[ O1 ]":
                continue

            meta = {
                "type":       ws.cell(row=6,  column=3).value or "",
                "department": ws.cell(row=9,  column=3).value or "",
                "team_code":  ws.cell(row=11, column=3).value or "",
                "team_name":  str(ws.cell(row=11, column=4).value or "").lstrip("/ ").strip(),
            }
            rows = self._read_sheet_rows(ws)
            if not rows:
                continue

            content = {"meta": meta, "headers": HEADERS, "rows": rows}
            payload = {
                "plan_year":  plan_year,
                "department": meta["department"],
                "team_code":  str(meta["team_code"]),
                "team_name":  meta["team_name"],
                "content":    content,
                "status":     "draft",
            }
            result = await self.upsert_opex_plan(db, payload, username)
            if result["success"]:
                imported.append({"sheet": ws.title, "rows": len(rows), "id": result["data"]["id"]})

        if not imported:
            return {"success": False, "error": "No recognizable data sheets found — none of the sheets in this "
                                                 "file match the OPEX Plan template layout (expected '[ O1 ]' "
                                                 "in cell A1 of each data sheet)."}
        return {"success": True, "imported": imported}

    def _to_dict(self, row: OpexPlan) -> dict:
        return {
            "id":          row.id,
            "plan_year":   row.plan_year,
            "department":  row.department,
            "team_code":   row.team_code,
            "team_name":   row.team_name,
            "content":     row.content,
            "status":      row.status,
            "created_by":  row.created_by,
            "created_at":  row.created_at.isoformat() if row.created_at else None,
            "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
        }
