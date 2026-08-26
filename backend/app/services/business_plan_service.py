"""
Business Plan Service — PAC module.
Stores and retrieves Business Plan documents in PostgreSQL.
"""
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.business_plan import PACBusinessPlan
import structlog

logger = structlog.get_logger()


class BusinessPlanService:

    # ── List ─────────────────────────────────────────────────────────────────

    async def list_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        doc_type: Optional[str] = None,
        department: Optional[str] = None,
    ) -> dict:
        q = select(PACBusinessPlan).order_by(
            PACBusinessPlan.plan_year.desc(),
            PACBusinessPlan.doc_type,
            PACBusinessPlan.department,
        )
        if plan_year:
            q = q.where(PACBusinessPlan.plan_year == plan_year)
        if doc_type:
            q = q.where(PACBusinessPlan.doc_type == doc_type)
        if department:
            q = q.where(PACBusinessPlan.department.ilike(f"%{department}%"))

        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    # ── Get single ────────────────────────────────────────────────────────────

    async def get_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PACBusinessPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    # ── Upsert ───────────────────────────────────────────────────────────────

    async def upsert_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        """
        Create or update a Business Plan document.
        Unique key: doc_type + plan_year + department + team_code.
        If id is provided and exists — update that record.
        Otherwise try to find by unique key; if not found — create.
        """
        doc_id = payload.get("id")
        row = None

        if doc_id:
            row = await db.get(PACBusinessPlan, doc_id)

        if not row:
            # Try unique-key lookup
            q = select(PACBusinessPlan).where(
                PACBusinessPlan.doc_type   == payload.get("doc_type"),
                PACBusinessPlan.plan_year  == payload.get("plan_year"),
                PACBusinessPlan.department == (payload.get("department") or "ALL"),
                PACBusinessPlan.team_code  == (payload.get("team_code") or ""),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()

        if row:
            row.content    = payload.get("content", row.content)
            row.team_name  = payload.get("team_name",  row.team_name)
            row.plan_role  = payload.get("plan_role",  row.plan_role)
            row.status     = payload.get("status",     row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = PACBusinessPlan(
                doc_type   = payload.get("doc_type", "strategy_plan"),
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department") or "ALL",
                team_code  = payload.get("team_code") or "",
                team_name  = payload.get("team_name") or "",
                plan_role  = payload.get("plan_role") or "",
                content    = payload.get("content", {}),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)

        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    # ── Import from Excel ───────────────────────────────────────────────────────

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "Strategy_Action Plan -
        Mashudi.xlsx" template — header at F5 (Department), F7 (Team
        Code/Name as "code / name"), H7 (Role); a 3-column body starting
        right after the "Managerial Objective / Strategy / Action Plan*"
        header row, with labels like "(1) ...", "(a) ...", "(i) ..." in
        columns C, K, P respectively. Blank C/K cells mean the row
        continues the previous objective/strategy."""
        import io
        import re
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        ws = wb.worksheets[0]

        def _clean(v) -> str:
            return str(v).strip() if v is not None else ""

        def _split_label(text: str):
            m = re.match(r"^\s*\(([^)]+)\)\s*(.*)$", text)
            return (m.group(1), m.group(2).strip()) if m else ("", text.strip())

        department = _clean(ws.cell(row=5, column=6).value)   # F5
        team_raw   = _clean(ws.cell(row=7, column=6).value)   # F7
        if "/" in team_raw:
            team_code, team_name = [p.strip() for p in team_raw.split("/", 1)]
        else:
            team_code, team_name = team_raw, ""
        plan_role  = _clean(ws.cell(row=7, column=8).value)   # H7

        header_row = None
        for r in range(1, ws.max_row + 1):
            if _clean(ws.cell(row=r, column=3).value) == "Managerial Objective":
                header_row = r
                break
        start_row = (header_row + 1) if header_row else 11

        items = []
        cur_obj = None
        cur_strat = None

        for r in range(start_row, ws.max_row + 1):
            obj_val   = _clean(ws.cell(row=r, column=3).value)    # C
            strat_val = _clean(ws.cell(row=r, column=11).value)   # K
            act_val   = _clean(ws.cell(row=r, column=16).value)   # P

            if obj_val.startswith("*"):
                break  # footnote row (e.g. "*) Action plan detail related to ... KPI plan")

            if obj_val:
                num, text = _split_label(obj_val)
                cur_obj = {"obj_num": num or str(len(items) + 1), "obj_text": text, "strategies": []}
                items.append(cur_obj)
                cur_strat = None

            if strat_val:
                if cur_obj is None:
                    cur_obj = {"obj_num": str(len(items) + 1), "obj_text": "", "strategies": []}
                    items.append(cur_obj)
                letter, text = _split_label(strat_val)
                cur_strat = {"letter": letter or chr(97 + len(cur_obj["strategies"])), "text": text, "actions": []}
                cur_obj["strategies"].append(cur_strat)

            if act_val:
                if cur_obj is None:
                    cur_obj = {"obj_num": str(len(items) + 1), "obj_text": "", "strategies": []}
                    items.append(cur_obj)
                if cur_strat is None:
                    cur_strat = {"letter": chr(97 + len(cur_obj["strategies"])), "text": "", "actions": []}
                    cur_obj["strategies"].append(cur_strat)
                num, text = _split_label(act_val)
                cur_strat["actions"].append({"num": num or "i", "text": text})

        if not items:
            return {"success": False, "error": "No data rows found — file doesn't match the expected Strategy & Action Plan template"}

        payload = {
            "doc_type":   "strategy_plan",
            "plan_year":  plan_year,
            "department": department or "ALL",
            "team_code":  team_code,
            "team_name":  team_name,
            "plan_role":  plan_role,
            "content":    {"items": items},
            "status":     "draft",
        }
        return await self.upsert_plan(db, payload, username)

    # ── Import Reporting > Manufacturing Plan from Excel ────────────────────────

    _MFG_REPORT_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    async def import_manufacture_plan_report_excel(self, db: AsyncSession, file_bytes: bytes,
                                                     plan_year: int, username: str) -> dict:
        """Parse the "4.Manufacture_Plan" sheet from the Business Plan
        Report workbook (e.g. "01.V3.2026 Business_Plan_Report_...xlsx") —
        a hierarchical Batch Size (vial) plan: Total -> Local/CMO/Export ->
        (CMO/Export only) per-customer sub-group -> Liquid/Freeze Dry ->
        individual products, each carrying a prior-year actual total,
        this-year plan total, and a Jan-Dec monthly breakdown of the plan
        total.

        Row shape in the source sheet (columns, 1-based):
          C = group/subgroup/numbered-subtotal label (e.g. "Total",
              "Local", "CMO, Etana", "1. Liquid")   -- OR --
          E = a leaf label: either a bare "Liquid"/"Freeze Dry" quick
              breakdown, or a real product name
          H = prior-year actual total, I = this-year plan total,
          J..U = this-year Jan..Dec monthly plan breakdown (12 columns)

        Row TYPE / indentation LEVEL is inferred from which column carries
        the label and the label text itself — there's no explicit
        level/indent stored in the source file:
          "Total"/"Local"/"CMO"/"Export" (col C, no comma)      -> group,    level 0
          "CMO, X"/"Export, X" (col C, contains a comma)        -> group,    level 1 (nested under the group above)
          "1. Liquid"/"2. Freeze Dry" (col C, "N. " prefix)     -> subtotal, one level under the current group
          bare "Liquid"/"Freeze Dry" (col E)                    -> subtotal, one level under the current group
          any other label (col E)                               -> line,     one level under the current subtotal
        (Classified by the col E label text alone, not by whether column G
        — a per-row reference batch size — happens to be filled in: a
        product with zero planned production can have a blank G cell in
        the source, which would otherwise be indistinguishable from a bare
        "Liquid"/"Freeze Dry" breakdown row.) This is a heuristic
        reconstruction of the sheet's visual nesting, not something the
        file states explicitly — worth a spot-check against the source
        after the first real upload.

        Repeated page headers (title row, "4. Manufacture Plan" section
        title, the "Title"/month header rows the source repeats on every
        printed page) are skipped by pattern, not by row number, so this
        survives a different row count in a future year's file."""
        import io
        import re
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        ws = next((s for s in wb.worksheets if "manufactur" in s.title.lower()), wb.worksheets[0])

        def cell(r, c):
            return ws.cell(row=r, column=c).value

        def clean(v) -> str:
            return str(v).strip() if v is not None else ""

        def num(v) -> float:
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        # Column labels for the 2 summary + 12 monthly value slots — read
        # from whichever header row appears first, instead of hardcoding
        # "2025(E)"/"2026(P)", so a future year's file just works.
        col_labels = None
        for r in range(1, ws.max_row + 1):
            if clean(cell(r, 3)) == "Title":
                month_row = r + 1
                col_labels = [clean(cell(r, 8)) or "Prior Year", clean(cell(r, 9)) or "Plan Year"]
                col_labels += [clean(cell(month_row, c)) or self._MFG_REPORT_MONTH_NAMES[c - 10] for c in range(10, 22)]
                break
        if col_labels is None:
            return {"success": False, "error": (
                "Could not find the header row (column C = 'Title') — this doesn't look like the "
                "'4.Manufacture_Plan' sheet layout."
            )}

        rows_out = []
        group_base = 0      # indentation level of the most recent group/subgroup row
        subtotal_base = 1   # indentation level of the most recent Liquid/Freeze Dry subtotal
        for r in range(1, ws.max_row + 1):
            c_val, e_val = clean(cell(r, 3)), clean(cell(r, 5))
            if not c_val and not e_val:
                continue
            if (c_val == "Title" or e_val == "Jan" or c_val == "Jan"
                    or clean(cell(r, 2)).startswith("4. Manufacture")
                    or clean(cell(r, 1)).startswith("PT CKD OTTO")):
                continue

            values = [num(cell(r, 8)), num(cell(r, 9))] + [num(cell(r, c)) for c in range(10, 22)]

            if c_val:
                if re.match(r"^\d+\.\s*(Liquid|Freeze Dry)", c_val, re.IGNORECASE):
                    level = group_base + 1
                    subtotal_base = level
                    rows_out.append({"label": c_val, "level": level, "type": "subtotal", "values": values})
                elif "," in c_val:
                    group_base = 1
                    rows_out.append({"label": c_val, "level": 1, "type": "group", "values": values})
                else:
                    group_base = 0
                    row_type = "total" if c_val.lower() == "total" else "group"
                    rows_out.append({"label": c_val, "level": 0, "type": row_type, "values": values})
            elif e_val.lower() not in ("liquid", "freeze dry"):
                # A real product — classified by its NAME, not by whether G
                # (the reference batch size) happens to be filled in: a
                # product with zero planned production can have a blank G
                # cell in the source, which would otherwise be
                # indistinguishable from a bare "Liquid"/"Freeze Dry"
                # breakdown row.
                rows_out.append({"label": e_val, "level": subtotal_base + 1, "type": "line", "values": values})
            else:
                level = group_base + 1
                subtotal_base = level
                rows_out.append({"label": e_val, "level": level, "type": "subtotal", "values": values})

        if not rows_out:
            return {"success": False, "error": (
                "No recognizable rows found — check the sheet matches the '4.Manufacture_Plan' layout."
            )}

        payload = {
            "doc_type": "mfg_plan_report", "plan_year": plan_year,
            "department": "ALL", "team_code": "", "team_name": "",
            "content": {"columns": col_labels, "rows": rows_out},
            "status": "final",
        }
        return await self.upsert_plan(db, payload, username)

    # ── Delete ────────────────────────────────────────────────────────────────

    async def delete_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PACBusinessPlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(PACBusinessPlan).where(PACBusinessPlan.id == plan_id))
        return {"success": True, "message": f"Deleted plan #{plan_id}"}

    # ── Helper ────────────────────────────────────────────────────────────────

    def _to_dict(self, row: PACBusinessPlan) -> dict:
        return {
            "id":          row.id,
            "doc_type":    row.doc_type,
            "plan_year":   row.plan_year,
            "department":  row.department,
            "team_code":   row.team_code,
            "team_name":   row.team_name,
            "plan_role":   row.plan_role,
            "content":     row.content,
            "status":      row.status,
            "created_at":  row.created_at.isoformat() if row.created_at else None,
            "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
            "created_by":  row.created_by,
        }
