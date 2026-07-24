"""
Purchase Plan (Material) Service
Handles CRUD for purchase plan inputs and Excel import, mirrors
sales_plan_service.py's shape.
"""
import io
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.purchase_plan import PurchasePlanMaterial
import structlog

logger = structlog.get_logger()

HEADERS = [
    "No", "Type", "Item Code No.", "Item Code Name", "UOM", "MOQ",
    "Stock", "QTY Needed", "Final QTY to Order",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Total", "Unit Price (Orig)", "Unit Price (IDR)", "Total Price (Rp)",
]

PP_CATEGORIES = ["Summary", "Local", "CMO", "Export"]


def _num(v):
    """Blank template cells sometimes hold stray non-numeric placeholders —
    treat anything that isn't a real number as "no value yet"."""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _default_content():
    return {
        "meta": {"type": "", "department": "", "team_code": "", "team_name": "", "exchange_rate": 0},
        "headers": HEADERS,
        "items": [],
    }


class PurchasePlanService:

    async def list_purchase_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
        team_code: Optional[str] = None,
        plan_category: Optional[str] = None,
    ) -> dict:
        q = select(PurchasePlanMaterial).order_by(
            PurchasePlanMaterial.plan_year.desc(), PurchasePlanMaterial.department, PurchasePlanMaterial.team_code
        )
        if plan_year:
            q = q.where(PurchasePlanMaterial.plan_year == plan_year)
        if department:
            q = q.where(PurchasePlanMaterial.department == department)
        if team_code:
            q = q.where(PurchasePlanMaterial.team_code == team_code)
        if plan_category:
            q = q.where(PurchasePlanMaterial.plan_category == plan_category)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_purchase_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PurchasePlanMaterial, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_purchase_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(PurchasePlanMaterial, plan_id)
        if not row:
            q = select(PurchasePlanMaterial).where(
                PurchasePlanMaterial.plan_year     == payload.get("plan_year"),
                PurchasePlanMaterial.plan_category == payload.get("plan_category", "Local"),
                PurchasePlanMaterial.department    == payload.get("department", ""),
                PurchasePlanMaterial.team_code     == str(payload.get("team_code", "")),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()
        if row:
            row.content    = payload.get("content", row.content)
            row.status     = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = PurchasePlanMaterial(
                plan_year     = payload.get("plan_year", datetime.now().year),
                plan_category = payload.get("plan_category", "Local"),
                department    = payload.get("department", ""),
                team_code     = str(payload.get("team_code", "")),
                team_name     = payload.get("team_name", ""),
                content       = payload.get("content", _default_content()),
                status        = payload.get("status", "draft"),
                created_by    = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def delete_purchase_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(PurchasePlanMaterial, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(PurchasePlanMaterial).where(PurchasePlanMaterial.id == plan_id))
        return {"success": True, "message": f"Deleted purchase plan #{plan_id}"}

    def _read_sheet_items(self, ws) -> list:
        """Each item spans two physical rows in the template: an "Order" row
        (columns K-V = Jan-Dec quantities to order, W=1H total, X=Total,
        Y/Z=Unit Price orig/IDR, AA=Total Price) directly followed by a
        "Received" row (same K-V month columns, X=Total). Row layout —
        header at 13-14, data from row 15."""
        items = []
        r = 15
        while r <= ws.max_row:
            no = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=4).value
            if no is None and name is None:
                r += 1
                continue
            order_months = [_num(ws.cell(row=r, column=c).value) for c in range(11, 23)]  # K-V
            item = {
                "no": no,
                "type": ws.cell(row=r, column=2).value or "",
                "item_code_no": ws.cell(row=r, column=3).value or "",
                "item_code_name": str(name or ""),
                "uom": ws.cell(row=r, column=5).value or "",
                "moq": _num(ws.cell(row=r, column=6).value) or None,
                "stock": _num(ws.cell(row=r, column=7).value) or None,
                "qty_needed": _num(ws.cell(row=r, column=8).value) or None,
                "final_qty_to_order": _num(ws.cell(row=r, column=9).value) or None,
                "order": order_months,
                "order_total": _num(ws.cell(row=r, column=24).value),        # X
                "unit_price_orig": _num(ws.cell(row=r, column=25).value),    # Y
                "unit_price_idr": _num(ws.cell(row=r, column=26).value),     # Z
                "total_price": _num(ws.cell(row=r, column=27).value),        # AA
                "received": [0] * 12,
                "received_total": 0,
            }
            # The next row, if it's the paired "Received" row (col J label,
            # or simply a row with no "no"/name of its own), holds received qty.
            if r + 1 <= ws.max_row:
                next_no = ws.cell(row=r + 1, column=1).value
                next_name = ws.cell(row=r + 1, column=4).value
                if next_no is None and next_name is None:
                    item["received"] = [_num(ws.cell(row=r + 1, column=c).value) for c in range(11, 23)]
                    item["received_total"] = _num(ws.cell(row=r + 1, column=24).value)
                    r += 1
            items.append(item)
            r += 1
        return items

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "(P1-M) Purchase plan_Material.xlsx"
        template — one plan per data sheet (Summary/Local/CMO/Export), each
        with its own Type/Department/Team meta and item list."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        imported = []
        for ws in wb.worksheets:
            # Sheet-name-agnostic — a sheet counts as a Purchase Plan data
            # sheet purely by matching the template's layout signature
            # ("[ P1 ]" in A1), not by its name. Reference/index sheets
            # (e.g. "Team and Department List") don't have this and are
            # skipped automatically.
            if str(ws.cell(row=1, column=1).value or "").strip() != "[ P1 ]":
                continue

            meta = {
                "type":          ws.cell(row=6,  column=4).value or "",
                "department":    ws.cell(row=9,  column=4).value or "",
                "team_code":     ws.cell(row=11, column=4).value or "",
                "team_name":     str(ws.cell(row=11, column=5).value or "").lstrip("/ ").strip(),
                "exchange_rate": _num(ws.cell(row=12, column=24).value),  # X12
            }
            items = self._read_sheet_items(ws)
            if not items:
                continue

            category = self._infer_category(meta["type"], ws.title)
            content = {"meta": meta, "headers": HEADERS, "items": items}
            payload = {
                "plan_year":     plan_year,
                "plan_category": category,
                "department":    meta["department"],
                "team_code":     str(meta["team_code"]),
                "team_name":     meta["team_name"],
                "content":       content,
                "status":        "draft",
            }
            result = await self.upsert_purchase_plan(db, payload, username)
            if result["success"]:
                imported.append({"category": category, "items": len(items), "id": result["data"]["id"]})

        if not imported:
            return {"success": False, "error": "No recognizable data sheets found — none of the sheets in this "
                                                 "file match the Purchase Plan template layout (expected '[ P1 ]' "
                                                 "in cell A1 of each data sheet)."}
        return {"success": True, "imported": imported}

    def _infer_category(self, type_text: str, sheet_name: str) -> str:
        """plan_category no longer depends on the sheet being named exactly
        'Purchase Plan Material_Local' etc. — inferred instead from the
        sheet's own "[ Type ]" meta text (e.g. "... - Local"), falling back
        to the sheet's name, then a generic default so any single-sheet
        upload (e.g. a plain "Sheet1") still imports successfully."""
        for text in (type_text, sheet_name):
            low = str(text or "").lower()
            for category in PP_CATEGORIES:
                if category.lower() in low:
                    return category
            if "total" in low:
                return "Summary"
        return "Local"

    def _to_dict(self, row: PurchasePlanMaterial) -> dict:
        return {
            "id":            row.id,
            "plan_year":     row.plan_year,
            "plan_category": row.plan_category,
            "department":    row.department,
            "team_code":     row.team_code,
            "team_name":     row.team_name,
            "content":       row.content,
            "status":        row.status,
            "created_by":    row.created_by,
            "created_at":    row.created_at.isoformat() if row.created_at else None,
            "updated_at":    row.updated_at.isoformat() if row.updated_at else None,
        }
