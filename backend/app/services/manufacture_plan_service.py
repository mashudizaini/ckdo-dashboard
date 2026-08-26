"""
Manufacture Plan ("Manufacture Plan Data" tab) Service
Handles CRUD for manufacture plan inputs and Excel import, mirrors
sales_plan_service.py's shape (flat headers+rows).
"""
import io
from datetime import datetime
from typing import Optional
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.manufacture_plan import ManufacturePlan
import structlog

logger = structlog.get_logger()

HEADERS = [
    "No", "Customer", "Item Code", "Name", "Batch Size (Vial)", "Yield (%)",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Total Batch", "Total Qty (Before Yield)", "Total Qty (After Yield)", "Sales Quantity", "Coverage",
]


def _num(v):
    """Blank template cells sometimes hold stray non-numeric placeholders —
    treat anything that isn't a real number as "no value yet"."""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _default_content():
    return {"meta": {"type": "", "department": "", "team_code": "", "team_name": "", "year": None},
            "headers": HEADERS, "rows": []}


class ManufacturePlanService:

    async def list_manufacture_plans(
        self,
        db: AsyncSession,
        plan_year: Optional[int] = None,
        department: Optional[str] = None,
        team_code: Optional[str] = None,
    ) -> dict:
        q = select(ManufacturePlan).order_by(
            ManufacturePlan.plan_year.desc(), ManufacturePlan.department, ManufacturePlan.team_code
        )
        if plan_year:
            q = q.where(ManufacturePlan.plan_year == plan_year)
        if department:
            q = q.where(ManufacturePlan.department == department)
        if team_code:
            q = q.where(ManufacturePlan.team_code == team_code)
        result = await db.execute(q)
        rows = result.scalars().all()
        return {"success": True, "count": len(rows), "data": [self._to_dict(r) for r in rows]}

    async def get_manufacture_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(ManufacturePlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        return {"success": True, "data": self._to_dict(row)}

    async def upsert_manufacture_plan(self, db: AsyncSession, payload: dict, username: str) -> dict:
        plan_id = payload.get("id")
        row = None
        if plan_id:
            row = await db.get(ManufacturePlan, plan_id)
        if not row:
            q = select(ManufacturePlan).where(
                ManufacturePlan.plan_year  == payload.get("plan_year"),
                ManufacturePlan.department == payload.get("department", ""),
                ManufacturePlan.team_code  == str(payload.get("team_code", "")),
            )
            result = await db.execute(q)
            row = result.scalar_one_or_none()
        if row:
            row.content    = payload.get("content", row.content)
            row.status     = payload.get("status",  row.status)
            row.updated_at = datetime.utcnow()
        else:
            row = ManufacturePlan(
                plan_year  = payload.get("plan_year", datetime.now().year),
                department = payload.get("department", ""),
                team_code  = str(payload.get("team_code", "")),
                team_name  = payload.get("team_name", ""),
                content    = payload.get("content", _default_content()),
                status     = payload.get("status", "draft"),
                created_by = username,
            )
            db.add(row)
        await db.flush()
        await db.refresh(row)
        return {"success": True, "data": self._to_dict(row)}

    async def delete_manufacture_plan(self, db: AsyncSession, plan_id: int) -> dict:
        row = await db.get(ManufacturePlan, plan_id)
        if not row:
            return {"success": False, "error": "Not found"}
        await db.execute(delete(ManufacturePlan).where(ManufacturePlan.id == plan_id))
        return {"success": True, "message": f"Deleted manufacture plan #{plan_id}"}

    async def import_excel(self, db: AsyncSession, file_bytes: bytes, plan_year: int, username: str) -> dict:
        """Parse an uploaded Excel matching the "manufacture plan template.xlsx"
        layout — meta at C6 (Type) / C9 (Department) / C11+D11 (Team Code/Name);
        two-row header at 13-14; data from row 15 until a "Total" row in
        column A. Sheet-name agnostic — matched by "[ M1 ]" in cell A1, same
        convention as the Purchase/Personnel Plan templates."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Could not read Excel file: {e}"}

        ws = None
        for sheet in wb.worksheets:
            if str(sheet.cell(row=1, column=1).value or "").strip() == "[ M1 ]":
                ws = sheet
                break
        if ws is None:
            return {"success": False, "error": "No recognizable data sheet found — expected a sheet with "
                                                 "'[ M1 ]' in cell A1, matching the Manufacture Plan template layout."}

        meta = {
            "type":       ws.cell(row=6,  column=3).value or "",
            "department": ws.cell(row=9,  column=3).value or "",
            "team_code":  ws.cell(row=11, column=3).value or "",
            "team_name":  str(ws.cell(row=11, column=4).value or "").lstrip("/ ").strip(),
            "year":       ws.cell(row=13, column=7).value,
        }

        rows = []
        r = 15
        while r <= ws.max_row:
            no = ws.cell(row=r, column=1).value
            if no is None:
                break
            if str(no).strip().lower() == "total":
                break
            name = ws.cell(row=r, column=4).value
            if name is None:
                # Blank template placeholder row (No pre-filled, nothing else) — skip.
                r += 1
                continue
            rows.append([
                no,
                ws.cell(row=r, column=2).value or "",
                ws.cell(row=r, column=3).value or "",
                str(name),
                _num(ws.cell(row=r, column=5).value) or None,
                _num(ws.cell(row=r, column=6).value) or None,
                *[_num(ws.cell(row=r, column=c).value) for c in range(7, 19)],  # G-R Jan-Dec
                _num(ws.cell(row=r, column=19).value),  # S  Total Batch
                _num(ws.cell(row=r, column=20).value),  # T  Total Qty Before Yield
                _num(ws.cell(row=r, column=21).value),  # U  Total Qty After Yield
                _num(ws.cell(row=r, column=22).value),  # V  Sales Quantity
                _num(ws.cell(row=r, column=23).value),  # W  Coverage
            ])
            r += 1

        if not rows:
            return {"success": False, "error": "No data rows found starting at row 15 — file doesn't match the expected template"}

        content = {"meta": meta, "headers": HEADERS, "rows": rows}
        payload = {
            "plan_year":  plan_year,
            "department": meta["department"],
            "team_code":  str(meta["team_code"]),
            "team_name":  meta["team_name"],
            "content":    content,
            "status":     "draft",
        }
        result = await self.upsert_manufacture_plan(db, payload, username)
        if result["success"]:
            result["rows_imported"] = len(rows)
        return result

    # ── Reporting > Manufacturing Plan report ────────────────────────────────
    # Computed live from Simulation Data > Manufacture Plan — no separate
    # upload for this report; it's a re-formatting of data already entered
    # there, matching the layout of "4.Manufacture_Plan" in the reference
    # Business Plan Report workbook (Total -> Local/CMO/Export -> (CMO/
    # Export only) per-customer sub-group -> Liquid/Freeze Dry -> product,
    # with a prior-year and this-year monthly Jan-Dec breakdown).
    #
    # Two things the reference format shows that Manufacture Plan
    # Simulation Data doesn't capture as an explicit field, so they're
    # inferred here:
    #   - Local vs CMO vs Export: each plan's own free-text [Type] meta
    #     (e.g. "Commercial Production - Export") is matched by substring,
    #     not a fixed enum -- a Type mentioning neither "export" nor "cmo"
    #     defaults to Local.
    #   - Liquid vs Freeze Dry: no such field anywhere in the input at all
    #     -- classified from the product's base name (strength/dosage
    #     suffix stripped) against a short hardcoded list of known
    #     freeze-dried products. Extend _MFG_REPORT_FREEZE_DRY_PRODUCTS
    #     below if a new freeze-dried product is added to the catalog.
    # Vial quantity per month = monthly batch count x Batch Size (Vial),
    # the same "before yield" formula the existing Detail Manufacturing
    # Plan export already uses (report figures are batch-plan output, not
    # yield-adjusted).

    _MFG_REPORT_FREEZE_DRY_PRODUCTS = {"bortezomib", "pemetrexed"}

    @classmethod
    def _report_form(cls, product_name: str) -> str:
        import re
        base = re.sub(r"\s+[\d.,]+\s*(mg|mcg|g|ml|iu)\b.*$", "", product_name or "", flags=re.IGNORECASE).strip().lower()
        return "Freeze Dry" if any(k in base for k in cls._MFG_REPORT_FREEZE_DRY_PRODUCTS) else "Liquid"

    @staticmethod
    def _report_business_type(meta_type: str) -> str:
        t = (meta_type or "").lower()
        if "export" in t:
            return "Export"
        if "cmo" in t:
            return "CMO"
        return "Local"

    async def _report_rows_for_year(self, db: AsyncSession, plan_year: int) -> list:
        result = await self.list_manufacture_plans(db, plan_year=plan_year)
        if not result.get("success"):
            return []
        out = []
        for plan in result["data"]:
            business_type = self._report_business_type((plan.get("content") or {}).get("meta", {}).get("type"))
            for row in (plan.get("content") or {}).get("rows") or []:
                if len(row) < 18:
                    continue
                name = str(row[3] or "").strip()
                if not name:
                    continue
                customer = str(row[1] or "").strip()
                batch_size = _num(row[4]) or 0
                months = [(_num(row[6 + m]) or 0) * batch_size for m in range(12)]
                if not any(months):
                    continue
                out.append({
                    "business_type": business_type, "customer": customer,
                    "form": self._report_form(name), "product": name, "months": months,
                })
        return out

    @staticmethod
    def _report_sum_months(node) -> list:
        """node is either a [12 floats] leaf or a dict of children (possibly
        nested several levels) -- recursively sums whichever it is."""
        if isinstance(node, list):
            return list(node)
        total = [0.0] * 12
        for child in node.values():
            child_sum = ManufacturePlanService._report_sum_months(child)
            for i in range(12):
                total[i] += child_sum[i]
        return total

    @staticmethod
    def _report_build_tree(rows: list) -> dict:
        """[{business_type, customer, form, product, months}] -> nested
        dict: tree[business_type][customer_key][form][product] = [12 floats].
        customer_key is None for Local (the reference format has no
        customer sub-grouping there), else the row's own Customer value."""
        tree: dict = {}
        for r in rows:
            bt = r["business_type"]
            cust = r["customer"] if bt != "Local" else None
            months = (tree.setdefault(bt, {}).setdefault(cust, {})
                          .setdefault(r["form"], {}).setdefault(r["product"], [0.0] * 12))
            for i in range(12):
                months[i] += r["months"][i]
        return tree

    async def get_report(self, db: AsyncSession, plan_year: int) -> dict:
        this_rows = await self._report_rows_for_year(db, plan_year)
        prior_rows = await self._report_rows_for_year(db, plan_year - 1)
        if not this_rows and not prior_rows:
            return {"success": False, "error": (
                f"No Manufacture Plan data found in Simulation Data for {plan_year} (or {plan_year - 1}) — "
                "upload it there first."
            )}

        cur = self._report_build_tree(this_rows)
        prior = self._report_build_tree(prior_rows)
        sum_months = self._report_sum_months

        def node(tree, *path):
            n = tree
            for p in path:
                if not isinstance(n, dict) or p not in n:
                    return {}
                n = n[p]
            return n

        def emit(label, level, row_type, months_cur, months_prior):
            rows_out.append({
                "label": label, "level": level, "type": row_type,
                "values": [sum(months_prior), sum(months_cur)] + months_cur,
            })

        def emit_form_breakdown(cur_node, prior_node, level, numbered: bool):
            """The Liquid/Freeze Dry rows under a group -- numbered=False for
            the bare quick-summary pair, numbered=True for the "1. Liquid" /
            "2. Freeze Dry" pair that's followed by individual products."""
            for i, form in enumerate(("Liquid", "Freeze Dry"), 1):
                fc, fp = cur_node.get(form, {}), prior_node.get(form, {})
                if not fc and not fp:
                    continue
                mc, mp = (sum_months(fc) if fc else [0.0] * 12), (sum_months(fp) if fp else [0.0] * 12)
                label = f"{i}. {form}" if numbered else form
                emit(label, level, "subtotal", mc, mp)
                if numbered:
                    for product in sorted(set(fc.keys()) | set(fp.keys())):
                        pc, pp = fc.get(product, [0.0] * 12), fp.get(product, [0.0] * 12)
                        emit(product, level + 1, "line", pc, pp)

        rows_out: list = []
        total_cur, total_prior = sum_months(cur) if cur else [0.0] * 12, sum_months(prior) if prior else [0.0] * 12
        emit("Total", 0, "total", total_cur, total_prior)
        # Total's own quick Liquid/Freeze Dry breakdown (every business type combined)
        total_forms_cur: dict = {}
        total_forms_prior: dict = {}
        for bt in ("Local", "CMO", "Export"):
            for src, dst in ((cur, total_forms_cur), (prior, total_forms_prior)):
                for cust_forms in node(src, bt).values():
                    for form, products in cust_forms.items():
                        dst.setdefault(form, {})
                        for prod, months in products.items():
                            existing = dst[form].setdefault(prod, [0.0] * 12)
                            for i in range(12):
                                existing[i] += months[i]
        emit_form_breakdown(total_forms_cur, total_forms_prior, 1, numbered=False)

        for bt in ("Local", "CMO", "Export"):
            bt_cur, bt_prior = node(cur, bt), node(prior, bt)
            if not bt_cur and not bt_prior:
                continue
            bt_months_cur = sum_months(bt_cur) if bt_cur else [0.0] * 12
            bt_months_prior = sum_months(bt_prior) if bt_prior else [0.0] * 12
            emit(bt, 0, "group", bt_months_cur, bt_months_prior)

            # bare quick Liquid/Freeze Dry breakdown for this business type (all its customers combined)
            bt_forms_cur: dict = {}
            bt_forms_prior: dict = {}
            for src, dst in ((bt_cur, bt_forms_cur), (bt_prior, bt_forms_prior)):
                for cust_forms in src.values():
                    for form, products in cust_forms.items():
                        dst.setdefault(form, {})
                        for prod, months in products.items():
                            existing = dst[form].setdefault(prod, [0.0] * 12)
                            for i in range(12):
                                existing[i] += months[i]
            emit_form_breakdown(bt_forms_cur, bt_forms_prior, 1, numbered=False)

            if bt == "Local":
                emit_form_breakdown(bt_cur.get(None, {}), bt_prior.get(None, {}), 1, numbered=True)
            else:
                customers = sorted((set(bt_cur.keys()) | set(bt_prior.keys())) - {None},
                                    key=lambda c: c or "")
                for cust in customers:
                    cust_cur, cust_prior = bt_cur.get(cust, {}), bt_prior.get(cust, {})
                    cust_months_cur = sum_months(cust_cur) if cust_cur else [0.0] * 12
                    cust_months_prior = sum_months(cust_prior) if cust_prior else [0.0] * 12
                    emit(f"{bt}, {cust or '(Unspecified)'}", 1, "group", cust_months_cur, cust_months_prior)
                    emit_form_breakdown(cust_cur, cust_prior, 2, numbered=True)

        columns = [f"{plan_year - 1} (Actual/Plan)", f"{plan_year} (Plan)"] + HEADERS[6:18]
        return {"success": True, "plan_year": plan_year, "columns": columns, "rows": rows_out}

    def _to_dict(self, row: ManufacturePlan) -> dict:
        return {
            "id":          row.id,
            "plan_year":   row.plan_year,
            "department":  row.department,
            "team_code":   row.team_code,
            "team_name":   row.team_name,
            "content":     row.content,
            "status":      row.status,
            "created_by":  row.created_by,
            "created_at":  row.created_at.isoformat() if row.created_at else None,
            "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
        }
