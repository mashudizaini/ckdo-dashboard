"""
HR Employee Router
Route prefix : /api/v1/dashboard/hr/employees
Upload file Excel karyawan (export standar Talenta) → REPLACE seluruh data di PostgreSQL.
"""
import io
import re
from datetime import datetime, date, timedelta
from calendar import monthrange
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, select, delete, extract, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import require_role, get_current_user, CurrentUser, Roles
from app.models.employee import Employee, EmployeeUploadLog

router = APIRouter()

# ── Mapping kolom Excel → field model ─────────────────────────────────────────
# Index berdasarkan posisi kolom di baris 1 (0-based), sesuai export standar Talenta
# (sheet "Employee Data", header di baris 1, data mulai baris 2).
COL = {
    "user_id":                  0,
    "full_name":                1,
    "sex":                      2,
    "level":                    3,
    "department":               4,
    "division":                 5,
    "team":                     6,
    "job_title":                7,
    "work_placement":           8,
    "status":                   9,
    "date_of_joining":          10,
    "retire_date":              14,
    "pkwt_ke":                  15,
    "starting_pkwt":            16,
    "end_pkwt":                 17,
    "permanent_date":           18,
    "resign_date":              19,
    "place_of_birth":           20,
    "date_of_birth":            22,
    "no_bpjs_health":           26,
    "no_bpjs_employee":         27,
    "education_degree":         28,   # "Master, Chung-Ang University" — dipecah jadi degree + school
    "education_major":          29,
    "employee_grade":           30,
    "working_experience_years": 31,
    "previous_company":         32,
    "address":                  36,
    "marital_status":           39,
    "phone_number":             40,
    "emergency_phone":          41,
    "religion":                 42,
    "blood_type":               43,
    "npwp_number":               45,
    "bank_account_bca":         50,
    "bank_account_name":        51,
    "personal_email":           77,
    "company_email":            78,
}

DATE_FIELDS = {"date_of_joining", "retire_date", "starting_pkwt", "end_pkwt",
               "permanent_date", "resign_date", "date_of_birth"}

# Panjang maksimum kolom String() di model Employee — dipakai untuk truncate defensif
# supaya upload tidak crash kalau data dari Talenta lebih panjang dari perkiraan.
MAXLEN = {
    "user_id": 20, "full_name": 200, "sex": 1, "level": 80, "department": 100,
    "division": 100, "team": 100, "job_title": 200, "work_placement": 100,
    "status": 50, "pkwt_ke": 30, "place_of_birth": 100, "no_bpjs_health": 50,
    "no_bpjs_employee": 50, "education_degree": 50, "education_school": 200,
    "education_major": 200, "employee_grade": 20, "working_experience_years": 20,
    "previous_company": 200, "marital_status": 50, "phone_number": 50,
    "emergency_phone": 50, "religion": 50, "blood_type": 5, "npwp_number": 50,
    "bank_account_bca": 50, "bank_account_name": 200, "personal_email": 200,
    "company_email": 200, "employment_status": 20,
}

# ── New template (from 2026 onward): "Employee Data ... .xls/.xlsx" with two
# sheets — "EMP Active" (has DIVISION/TEAM columns) and "EMP Resign" (doesn't).
# Column *positions* differ from the legacy Talenta layout (COL above) and even
# between these two sheets, so instead of hardcoded indices this format is
# parsed by matching each sheet's own header row text → field name. This is
# robust to the two sheets having different column counts/order, and to minor
# future column insertions in the template.
NEW_TEMPLATE_SHEETS = {"EMP Active": "Active", "EMP Resign": "Resign"}

NEW_TEMPLATE_ALIASES = {
    "user_id":                  "user id",
    "full_name":                "full name",
    "resign_date":               "resign date",
    "sex":                      "sex",
    "level":                    "level",
    "department":               "department",
    "division":                 "division",
    "team":                     "team",
    "job_title":                "job title",
    "work_placement":           "work placement",
    "status":                   "status",
    "date_of_joining":          "doj",
    "retire_date":              "retired (age 55)",
    "pkwt_ke":                  "pkwt ke-",
    "starting_pkwt":            "starting pkwt",
    "end_pkwt":                 "end pkwt",
    "permanent_date":           "permanent date",
    "place_of_birth":           "place of birth",
    "date_of_birth":            "date of birth",
    "no_bpjs_health":           "no bpjs health insurance",
    "no_bpjs_employee":         "no bpjs employee benefits",
    "employee_grade":           "employee grade",
    "working_experience_years": "working experience",
    "previous_company":         "previous company",
    "address":                  "adress/resident employee",
    "marital_status":           "marital status",
    "phone_number":             "phone number/hp",
    "emergency_phone":          "phone number/hp (emergency)",
    "religion":                 "religion",
    "blood_type":               "blood type",
    "npwp_number":              "npwp number",
    "bank_account_bca":         "rekening number (bca)",
    "bank_account_name":        "rekening name (bca)",
    "personal_email":           "personal email",
    "company_email":            "company email",
}
# "LAST EDUCATION BACKROUND" (sic — matches the template's own spelling) is a
# merged header spanning two sub-columns ("Degree" directly beneath it, "Major"
# in the next column) — handled separately from the alias table above since it
# needs the header column index, not a second header-text lookup.
_EDU_HEADER = "last education backround"


def _norm_header(s) -> str:
    s = re.sub(r"\s+", " ", str(s or "")).strip().lower()
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    return s


def _find_header_row(rows: list, max_scan: int = 15) -> Optional[int]:
    """Locate the header row within the first `max_scan` rows — the one containing
    a "USER ID" cell. Sheets in this template start with a few summary/title rows."""
    for r in range(min(max_scan, len(rows))):
        if any(_norm_header(c) == "user id" for c in rows[r]):
            return r
    return None


def _build_header_map(header_row: tuple) -> dict:
    """normalized header text -> first matching column index. Several headers
    repeat (e.g. "sex"/"dob"/"year" for each family member column) — only the
    first (primary) occurrence is kept, which is always the one this parser uses."""
    m = {}
    for idx, v in enumerate(header_row):
        key = _norm_header(v)
        if key and key not in m:
            m[key] = idx
    return m


def _parse_new_template_sheet(rows: list, batch_id: str, employment_status: str) -> list:
    """Parse one sheet ("EMP Active" or "EMP Resign") of the new template into
    a list of field dicts, tagging every row with `employment_status`."""
    header_row_idx = _find_header_row(rows)
    if header_row_idx is None:
        return []
    header_map = _build_header_map(rows[header_row_idx])
    edu_col = header_map.get(_EDU_HEADER)
    data_start = header_row_idx + 2  # main header row + sub-header row (Degree/Major etc.)

    out = []
    for r in range(data_start, len(rows)):
        row = rows[r]

        def cell(col_idx):
            return row[col_idx] if col_idx is not None and col_idx < len(row) else None

        user_id = _to_str(cell(header_map.get("user id")), "user_id")
        if not user_id:
            continue

        data = {
            "upload_batch_id": batch_id, "uploaded_at": datetime.utcnow(),
            "user_id": user_id, "employment_status": employment_status,
        }
        for field, header_key in NEW_TEMPLATE_ALIASES.items():
            if field == "user_id":
                continue
            val = cell(header_map.get(header_key))
            if field in DATE_FIELDS:
                data[field] = _to_date(val)
            elif field == "marital_status":
                raw = _to_str(val, field)
                data[field] = _MARITAL_MAP.get(raw.lower(), raw) if raw else None
            else:
                data[field] = _to_str(val, field)

        if edu_col is not None:
            data["education_degree"] = _to_str(cell(edu_col), "education_degree")
            data["education_major"]  = _to_str(cell(edu_col + 1), "education_major")
        data.setdefault("education_school", None)  # no separate school/university column in this template

        out.append(data)
    return out


def _sheet_names(contents: bytes, filename: str) -> list:
    if filename.lower().endswith(".xls"):
        import xlrd
        return xlrd.open_workbook(file_contents=contents).sheet_names()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    return wb.sheetnames


def _read_sheet_rows(contents: bytes, filename: str, sheet_name: str) -> list:
    """All rows of one sheet as a list of value-tuples, regardless of .xls/.xlsx engine."""
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=contents)
        sheet = wb.sheet_by_name(sheet_name)
        return [tuple(sheet.row_values(r)) for r in range(sheet.nrows)]

    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb[sheet_name]
    return list(ws.iter_rows(values_only=True))


_EMPTY_MARKERS = ("", "-", "none", "n/a", "na", "null")

# Talenta exports mix Indonesian and English marital-status values depending on
# who entered the data — normalize to English so the dashboard is consistent
# (this is a foreign company, all UI/data labels must read in English).
_MARITAL_MAP = {
    "menikah": "Married",
    "belum menikah": "Single",
    "cerai": "Divorced",
    "cerai hidup": "Divorced",
    "cerai mati": "Widow",
    "duda": "Widower",
    "janda": "Widow",
}


def _to_str(val, field: str = None) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        s = str(int(val))
    else:
        s = str(val).strip()
    if s.lower() in _EMPTY_MARKERS:
        return None
    max_len = MAXLEN.get(field) if field else None
    return s[:max_len] if max_len else s


def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # Excel serial date (1900 date system) — fallback untuk cell yang belum
        # ter-format sebagai tanggal saat file di-convert dari .xls lama.
        try:
            return (datetime(1899, 12, 30) + timedelta(days=val)).date()
        except (OverflowError, ValueError):
            return None
    s = str(val).strip()
    if s.lower() in _EMPTY_MARKERS:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_row(row: tuple, batch_id: str) -> Optional[dict]:
    """Ubah satu baris Excel → dict field. Return None jika baris kosong."""
    def cell(idx):
        return row[idx] if idx < len(row) else None

    user_id = _to_str(cell(COL["user_id"]), "user_id")
    if not user_id:
        return None

    data = {"upload_batch_id": batch_id, "uploaded_at": datetime.utcnow(), "user_id": user_id}
    for field, col_idx in COL.items():
        if field == "user_id":
            continue
        val = cell(col_idx)
        if field in DATE_FIELDS:
            data[field] = _to_date(val)
        elif field == "education_degree":
            # Talenta menggabungkan "Degree, School" dalam satu sel.
            raw = _to_str(val)
            if raw and "," in raw:
                degree, school = raw.split(",", 1)
                data["education_degree"] = degree.strip()[:MAXLEN["education_degree"]]
                data["education_school"] = school.strip()[:MAXLEN["education_school"]]
            else:
                data["education_degree"] = (raw or "")[:MAXLEN["education_degree"]] or None
                data["education_school"] = None
        elif field == "marital_status":
            raw = _to_str(val, field)
            data["marital_status"] = _MARITAL_MAP.get(raw.lower(), raw) if raw else None
        else:
            data[field] = _to_str(val, field)
    return data


def _extract_rows(contents: bytes, filename: str) -> list:
    """Baca file Excel (.xls lama atau .xlsx/.xlsm) → list of row tuples, data mulai baris 2."""
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=contents)
        sheet = wb.sheet_by_name("Employee Data") if "Employee Data" in wb.sheet_names() else wb.sheet_by_index(0)
        return [tuple(sheet.row_values(r)) for r in range(1, sheet.nrows)]

    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb["Employee Data"] if "Employee Data" in wb.sheetnames else wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


# ── Organization Chart — derive "reports to" from level/department/division/team ──
# Talenta's export has no "Direct Supervisor" column, so instead of leaving every
# employee's supervisor blank after each REPLACE upload, infer it from the existing
# level/department/division/team fields: within each (department, division, team)
# group the most senior `level` becomes that group's lead, everyone else in the group
# reports to the lead, and each group's lead reports to the lead of its parent group
# (team → division → department → President Director). HR can still correct any one
# employee's supervisor afterward via PATCH /{user_id}/supervisor.
_LEVEL_RANK = {
    "Director": 0, "General Manager": 1, "Senior Manager": 2, "Manager": 3,
    "Assistant Manager": 4, "Supervisor": 5, "Senior Staff": 6,
    "Officer": 7, "Staff": 7, "Operator / Clerk": 8,
}


def _assign_supervisors(rows: list) -> None:
    """Mutates each row dict in `rows`, setting row['supervisor_id']. No-op if no
    'President Director' row is found (nothing recognizable to anchor the chart to)."""
    def rank(row):
        return _LEVEL_RANK.get(row.get("level"), 9)

    root = next((r for r in rows if (r.get("job_title") or "").strip().lower() == "president director"), None)
    if root is None:
        return
    root["supervisor_id"] = None

    # Plant's Director sits in a separate "Director" pseudo-department alongside the
    # President Director rather than in a (department="Plant", division="", team="")
    # bucket of its own, so it's keyed onto department "Plant" by hand.
    plant_director = next((r for r in rows if (r.get("job_title") or "").strip().lower() == "plant director"), None)
    dept_head_override = {}
    if plant_director:
        plant_director["supervisor_id"] = root["user_id"]
        dept_head_override["Plant"] = plant_director

    others = [r for r in rows if r is not root and r is not plant_director]

    def group_key(r):
        return (r.get("department") or "", r.get("division") or "", r.get("team") or "")

    groups: dict = {}
    for r in others:
        groups.setdefault(group_key(r), []).append(r)

    def leader_of(key):
        members = groups.get(key)
        return min(members, key=rank) if members else None

    for key, members in groups.items():
        dept, div, team = key
        leader = leader_of(key)
        for r in members:
            if r is not leader:
                r["supervisor_id"] = leader["user_id"]

        if team:
            parent_key = (dept, div, "")
        elif div:
            parent_key = (dept, "", "")
        else:
            parent_key = None  # this bucket already IS the department-head bucket

        if parent_key is None:
            leader["supervisor_id"] = root["user_id"]
        else:
            parent_leader = leader_of(parent_key) or dept_head_override.get(dept)
            leader["supervisor_id"] = (parent_leader or root)["user_id"]


# ── Upload endpoint ────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_employees(
    file:     UploadFile = File(...),
    notes:    str        = Form(""),
    db:       AsyncSession = Depends(get_db),
    user:     CurrentUser  = Depends(require_role(Roles.HR)),
):
    """
    Upload file Excel karyawan. Dua format didukung:
      - Template baru (2026+): sheet "EMP Active" + "EMP Resign" — kolom dibaca
        dinamis dari header text tiap sheet (posisi kolom beda antar sheet).
        employment_status diisi "Active"/"Resign" sesuai sheet asal baris.
      - Legacy (export standar Talenta): sheet "Employee Data", header di baris 1,
        data mulai baris 2. employment_status diturunkan dari ada/tidaknya resign_date.
    Logic: REPLACE — seluruh data karyawan lama dihapus, diganti data dari file baru.
    """
    if not file.filename.lower().endswith((".xls", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be .xls, .xlsx, or .xlsm format")

    contents = await file.read()
    batch_id = f"batch_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    try:
        sheet_names = _sheet_names(contents, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    new_format_sheets = [s for s in NEW_TEMPLATE_SHEETS if s in sheet_names]

    rows_parsed = []
    seen_ids = set()
    skipped = 0
    try:
        if new_format_sheets:
            for sheet_name in new_format_sheets:
                sheet_rows = _read_sheet_rows(contents, file.filename, sheet_name)
                for parsed in _parse_new_template_sheet(sheet_rows, batch_id, NEW_TEMPLATE_SHEETS[sheet_name]):
                    if parsed["user_id"] in seen_ids:
                        skipped += 1
                        continue
                    seen_ids.add(parsed["user_id"])
                    rows_parsed.append(parsed)
        elif "Employee Data" in sheet_names:
            data_rows = _extract_rows(contents, file.filename)
            for row in data_rows:
                parsed = _parse_row(row, batch_id)
                if parsed is None:
                    continue
                if parsed["user_id"] in seen_ids:
                    skipped += 1
                    continue
                parsed["employment_status"] = "Resign" if parsed.get("resign_date") else "Active"
                seen_ids.add(parsed["user_id"])
                rows_parsed.append(parsed)
        else:
            raise HTTPException(status_code=422, detail="Unrecognized template: expected sheets "
                                 "'EMP Active'/'EMP Resign', or legacy 'Employee Data'.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to read data rows: {e}")

    if not rows_parsed:
        raise HTTPException(status_code=422, detail="No employee data found in file. "
                            "Make sure the sheet is named 'EMP Active'/'EMP Resign' (new template) "
                            "or 'Employee Data' (legacy) with headers in the right row.")

    _assign_supervisors(rows_parsed)

    try:
        # REPLACE — delete all previous data, replace with the new file's data
        deleted_result = await db.execute(delete(Employee))
        replaced_count = deleted_result.rowcount or 0

        db.add_all(Employee(**data) for data in rows_parsed)
        await db.flush()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to save data to database: {e}")

    # Simpan log upload
    log = EmployeeUploadLog(
        batch_id    = batch_id,
        filename    = file.filename,
        total_rows  = len(rows_parsed),
        inserted    = len(rows_parsed),
        updated     = 0,
        skipped     = skipped,
        uploaded_by = user.username or "unknown",
        notes       = notes or None,
    )
    db.add(log)

    return {
        "batch_id":          batch_id,
        "filename":          file.filename,
        "total_rows":        len(rows_parsed),
        "inserted":          len(rows_parsed),
        "updated":           0,
        "skipped":           skipped,
        "replaced_previous": replaced_count,
        "message":           f"Upload successful: {len(rows_parsed)} employee records replaced {replaced_count} previous records.",
    }


# ── Query endpoints ────────────────────────────────────────────────────────────

@router.get("/birthdays-this-month")
async def get_birthdays_this_month(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(get_current_user),
):
    """Active employees whose birthday falls in the current month — powers the
    Application Center's Announcement panel. Open to any authenticated user
    (not just HR), unlike the rest of this router."""
    today = date.today()
    q = await db.execute(
        select(Employee.full_name, Employee.department, Employee.job_title, Employee.date_of_birth)
        .where(
            Employee.date_of_birth.isnot(None),
            extract("month", Employee.date_of_birth) == today.month,
            or_(Employee.employment_status == "Active",
                and_(Employee.employment_status.is_(None), Employee.resign_date.is_(None))),
        )
        .order_by(extract("day", Employee.date_of_birth))
    )
    return [
        {
            "name":       r[0],
            "department": r[1],
            "job_title":  r[2],
            "day":        r[3].day,
            "date":       r[3].isoformat(),
            "is_today":   r[3].month == today.month and r[3].day == today.day,
        }
        for r in q.fetchall()
    ]


@router.get("/summary")
async def get_employee_summary(
    join_month: Optional[int] = Query(None, ge=1, le=12),  # same cumulative cutoff as the list below —
    join_year:  Optional[int] = Query(None),                # keeps the KPI cards in sync with what's shown
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Statistik ringkasan untuk KPI cards.

    Accepts the same join_month/join_year cumulative cutoff as the list below
    it, and applies it to every count here — otherwise the cards (unfiltered)
    silently disagree with the filtered list (e.g. "Resign" card showing more
    than the list actually displays for the same cutoff)."""
    base = _apply_employee_filters(select(Employee), join_month=join_month, join_year=join_year)

    def counted(*conditions):
        q = base
        for c in conditions:
            q = q.where(c)
        return select(func.count()).select_from(q.subquery())

    total_q   = await db.execute(counted())
    total     = total_q.scalar() or 0

    perm_q    = await db.execute(counted(Employee.status == "Permanent"))
    permanent = perm_q.scalar() or 0

    contract_q = await db.execute(counted(Employee.status == "Contract"))
    contract  = contract_q.scalar() or 0

    probation_q = await db.execute(counted(Employee.status == "Probation"))
    probation = probation_q.scalar() or 0

    active_q = await db.execute(counted(Employee.employment_status == "Active"))
    active_count = active_q.scalar() or 0

    resign_q = await db.execute(counted(Employee.employment_status == "Resign"))
    resign_count = resign_q.scalar() or 0

    base_sq = base.subquery()

    by_dept_q = await db.execute(
        select(base_sq.c.department, func.count().label("total"))
        .group_by(base_sq.c.department)
        .order_by(func.count().desc())
    )
    by_dept = [{"department": r[0] or "—", "total": r[1]} for r in by_dept_q.fetchall()]

    by_level_q = await db.execute(
        select(base_sq.c.level, func.count().label("total"))
        .group_by(base_sq.c.level)
        .order_by(func.count().desc())
    )
    by_level = [{"level": r[0] or "—", "total": r[1]} for r in by_level_q.fetchall()]

    by_sex_q = await db.execute(
        select(base_sq.c.sex, func.count().label("total"))
        .group_by(base_sq.c.sex)
    )
    by_sex = {r[0]: r[1] for r in by_sex_q.fetchall()}

    return {
        "total":     total,
        "permanent": permanent,
        "contract":  contract,
        "probation": probation,
        "active":    active_count,
        "resign":    resign_count,
        "by_dept":   by_dept,
        "by_level":  by_level,
        "male":      by_sex.get("M", 0),
        "female":    by_sex.get("F", 0),
    }


def _emp_dict(e: Employee) -> dict:
    return {
        "id":               e.id,
        "user_id":          e.user_id,
        "full_name":        e.full_name,
        "sex":              e.sex,
        "level":            e.level,
        "department":       e.department,
        "division":         e.division,
        "team":             e.team,
        "job_title":        e.job_title,
        "work_placement":   e.work_placement,
        "status":           e.status,
        "employment_status": e.employment_status,
        "employee_grade":   e.employee_grade,
        "supervisor_id":    e.supervisor_id,
        "education_degree": e.education_degree,
        "education_school": e.education_school,
        "education_major":  e.education_major,
        "marital_status":   e.marital_status,
        "religion":         e.religion,
        "blood_type":       e.blood_type,
        "phone_number":     e.phone_number,
        "emergency_phone":  e.emergency_phone,
        "company_email":    e.company_email,
        "personal_email":   e.personal_email,
        "date_of_joining":  str(e.date_of_joining)  if e.date_of_joining  else None,
        "date_of_birth":    str(e.date_of_birth)    if e.date_of_birth    else None,
        "place_of_birth":   e.place_of_birth,
        "retire_date":      str(e.retire_date)       if e.retire_date      else None,
        "end_pkwt":         str(e.end_pkwt)          if e.end_pkwt         else None,
        "starting_pkwt":    str(e.starting_pkwt)     if e.starting_pkwt    else None,
        "pkwt_ke":          e.pkwt_ke,
        "permanent_date":   str(e.permanent_date)    if e.permanent_date   else None,
        "resign_date":      str(e.resign_date)        if e.resign_date      else None,
        "resign_reason":    e.resign_reason,
        "no_bpjs_health":   e.no_bpjs_health,
        "no_bpjs_employee": e.no_bpjs_employee,
        "working_experience_years": e.working_experience_years,
        "previous_company": e.previous_company,
        "address":          e.address,
        "npwp_number":      e.npwp_number,
        "bank_account_bca": e.bank_account_bca,
        "bank_account_name": e.bank_account_name,
    }


def _apply_employee_filters(
    q, *, search=None, department=None, division=None, status=None, employment_status=None,
    level=None, team=None, sex=None, education=None, position=None,
    marital_status=None, join_month=None, join_year=None,
    snapshot_month=None, snapshot_year=None,
):
    """Shared WHERE-clause builder — used by both the list endpoint and the
    Excel export, so the two can never silently drift apart on what counts
    as "matching" a filter."""
    if search:
        term = f"%{search}%"
        q = q.where(
            Employee.full_name.ilike(term)  |
            Employee.user_id.ilike(term)    |
            Employee.job_title.ilike(term)
        )
    if department:
        if department in DEPT_GROUPS:
            # One of the 4 canonical Employee Summary groups (e.g. drilling
            # down from the summary view) — match every raw department value
            # that rolls up into this group, not just an exact string match.
            # Needed both for case-duplicates ("Plant"/"PLANT") and for
            # misfiled raw values ("Director", "Validation", ...) that group
            # display labels don't literally match (e.g. "Strategy &
            # Development" vs the raw "Strategy Development").
            raw_uppers = [k for k, v in _DEPT_GROUP_MAP.items() if v == department]
            q = q.where(func.upper(Employee.department).in_(raw_uppers))
        else:
            # Case-insensitive — the source Excel has case duplicates for the
            # same department ("Plant" / "PLANT"); an exact match would
            # silently miss half of them.
            q = q.where(func.upper(Employee.department) == department.upper())
    if division:
        q = q.where(Employee.division == division)
    if sex:
        q = q.where(Employee.sex == sex)
    if status:
        q = q.where(Employee.status == status)
    if employment_status:
        q = q.where(Employee.employment_status == employment_status)
    if level:
        q = q.where(Employee.level == level)
    if team:
        q = q.where(Employee.team == team)
    if education:
        q = q.where(Employee.education_degree == education)
    if position:
        q = q.where(Employee.job_title == position)
    if marital_status:
        q = q.where(Employee.marital_status == marital_status)
    if join_year:
        # Cumulative "as of" cutoff — how many employees have joined up to
        # this month/year, regardless of whether they've since resigned.
        cutoff_month = join_month or 12
        cutoff_date = date(join_year, cutoff_month, monthrange(join_year, cutoff_month)[1])
        q = q.where(Employee.date_of_joining.isnot(None), Employee.date_of_joining <= cutoff_date)
    if snapshot_year:
        # "Active as of" a specific month-end — same windowing as
        # /summary/by-month, so drilling down from that report into this
        # list yields the exact same headcount.
        cutoff_month = snapshot_month or 12
        snapshot_date = date(snapshot_year, cutoff_month, monthrange(snapshot_year, cutoff_month)[1])
        q = q.where(
            Employee.date_of_joining.isnot(None), Employee.date_of_joining <= snapshot_date,
            (Employee.resign_date.is_(None) | (Employee.resign_date >= snapshot_date)),
        )
    return q


@router.get("")
async def list_employees(
    search:     str           = Query(""),
    department: Optional[str] = Query(None),
    division:   Optional[str] = Query(None),
    status:     Optional[str] = Query(None),
    employment_status: Optional[str] = Query(None),  # "Active" / "Resign"
    level:      Optional[str] = Query(None),
    team:       Optional[str] = Query(None),
    sex:        Optional[str] = Query(None),
    education:  Optional[str] = Query(None),  # education_degree, exact match
    position:   Optional[str] = Query(None),  # job_title, exact match
    marital_status: Optional[str] = Query(None),
    join_month: Optional[int] = Query(None, ge=1, le=12),  # cumulative cutoff: join date <= end of this month/year
    join_year:  Optional[int] = Query(None),
    snapshot_month: Optional[int] = Query(None, ge=1, le=12),  # "active as of" end of this month/year
    snapshot_year:  Optional[int] = Query(None),
    sort_by:    str           = Query("full_name"),
    sort_dir:   str           = Query("asc"),
    page:       int           = Query(1, ge=1),
    page_size:  int           = Query(25, ge=1, le=5000),  # higher cap lets the frontend fetch everything in one call for Excel export
    db:         AsyncSession  = Depends(get_db),
    user:       CurrentUser   = Depends(require_role(Roles.HR)),
):
    """List karyawan dengan search, filter, pagination, dan sorting."""
    q = _apply_employee_filters(
        select(Employee), search=search, department=department, division=division, status=status,
        employment_status=employment_status, level=level, team=team, sex=sex,
        education=education, position=position, marital_status=marital_status,
        join_month=join_month, join_year=join_year,
        snapshot_month=snapshot_month, snapshot_year=snapshot_year,
    )
    # Total count
    count_q  = await db.execute(select(func.count()).select_from(q.subquery()))
    total    = count_q.scalar() or 0

    # Dynamic sort
    _SORT_COLS = {
        "user_id":          Employee.user_id,
        "full_name":        Employee.full_name,
        "level":            Employee.level,
        "department":       Employee.department,
        "division":         Employee.division,
        "team":             Employee.team,
        "job_title":        Employee.job_title,
        "work_placement":   Employee.work_placement,
        "status":           Employee.status,
        "employment_status": Employee.employment_status,
        "sex":              Employee.sex,
        "employee_grade":   Employee.employee_grade,
        "education_degree": Employee.education_degree,
        "education_school": Employee.education_school,
        "education_major":  Employee.education_major,
        "marital_status":   Employee.marital_status,
        "date_of_joining":  Employee.date_of_joining,
        "end_pkwt":         Employee.end_pkwt,
        "place_of_birth":   Employee.place_of_birth,
        "date_of_birth":    Employee.date_of_birth,
        "religion":         Employee.religion,
        "blood_type":       Employee.blood_type,
        "supervisor_id":    Employee.supervisor_id,
        "pkwt_ke":          Employee.pkwt_ke,
        "starting_pkwt":    Employee.starting_pkwt,
        "permanent_date":   Employee.permanent_date,
        "resign_date":      Employee.resign_date,
        "resign_reason":    Employee.resign_reason,
        "retire_date":      Employee.retire_date,
        "working_experience_years": Employee.working_experience_years,
        "previous_company": Employee.previous_company,
        "phone_number":     Employee.phone_number,
        "emergency_phone":  Employee.emergency_phone,
        "company_email":    Employee.company_email,
        "personal_email":   Employee.personal_email,
        "address":          Employee.address,
        "no_bpjs_health":   Employee.no_bpjs_health,
        "no_bpjs_employee": Employee.no_bpjs_employee,
        "npwp_number":      Employee.npwp_number,
        "bank_account_bca": Employee.bank_account_bca,
        "bank_account_name": Employee.bank_account_name,
    }
    sort_col = _SORT_COLS.get(sort_by, Employee.full_name)
    q        = q.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())
    q        = q.offset((page - 1) * page_size).limit(page_size)
    result   = await db.execute(q)
    employees = result.scalars().all()

    return {
        "total":      total,
        "page":       page,
        "page_size":  page_size,
        "pages":      (total + page_size - 1) // page_size,
        "employees":  [_emp_dict(e) for e in employees],
    }


# ── Excel export — styled to match HR's reference template ────────────────────
# (key, label, width) — same 18 fields/order as the frontend's export picker
# (EMPLOYEE_COLS in HR.jsx), plus the column widths from the reference file
# employee_data_2026-07-23_output_template_excel.xlsx.
_EXPORT_COLUMNS = [
    ("user_id",           "NIK",               10),
    ("full_name",         "Name",              32),
    ("level",              "Level",            18),
    ("department",         "Department",       20),
    ("division",           "Division",         14),
    ("team",               "Team",             10),
    ("job_title",          "Position",         32),
    ("work_placement",     "Placement",        16),
    ("status",             "Status",           12),
    ("employment_status",  "Employment Status", 18),
    ("sex",                "Gender",           10),
    ("employee_grade",     "Grade",            10),
    ("education_degree",   "Education",        40),
    ("marital_status",     "Marital",          10),
    ("date_of_joining",    "Join Date",        14),
    ("resign_date",        "Resign Date",      14),
    ("end_pkwt",           "End PKWT",         14),
    ("phone_number",       "Phone",            16),
]
_EXPORT_DATE_KEYS = {"date_of_joining", "resign_date", "end_pkwt"}


def _export_cell_value(e: Employee, key: str):
    if key == "resign_date":
        return "" if e.employment_status == "Active" else _fmt_export_date(e.resign_date)
    if key in _EXPORT_DATE_KEYS:
        return _fmt_export_date(getattr(e, key))
    if key == "sex":
        return "Male" if e.sex == "M" else "Female" if e.sex == "F" else ""
    return getattr(e, key) or ""


def _fmt_export_date(d):
    return d.strftime("%b %d, %Y") if d else ""


@router.get("/export")
async def export_employees(
    search:     str           = Query(""),
    department: Optional[str] = Query(None),
    division:   Optional[str] = Query(None),
    status:     Optional[str] = Query(None),
    employment_status: Optional[str] = Query(None),
    level:      Optional[str] = Query(None),
    team:       Optional[str] = Query(None),
    sex:        Optional[str] = Query(None),
    education:  Optional[str] = Query(None),
    position:   Optional[str] = Query(None),
    marital_status: Optional[str] = Query(None),
    join_month: Optional[int] = Query(None, ge=1, le=12),
    join_year:  Optional[int] = Query(None),
    snapshot_month: Optional[int] = Query(None, ge=1, le=12),
    snapshot_year:  Optional[int] = Query(None),
    sort_by:    str           = Query("full_name"),
    sort_dir:   str           = Query("asc"),
    fields:     Optional[str] = Query(None),  # comma-separated _EXPORT_COLUMNS keys; default = all
    db:         AsyncSession  = Depends(get_db),
    user:       CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Excel export of the (filtered) Employee List, styled to match HR's
    reference template — bold centered header with fill + borders, thin
    borders on data rows, "MMM DD, YYYY" dates, auto column widths."""
    q = _apply_employee_filters(
        select(Employee), search=search, department=department, division=division, status=status,
        employment_status=employment_status, level=level, team=team, sex=sex,
        education=education, position=position, marital_status=marital_status,
        join_month=join_month, join_year=join_year,
        snapshot_month=snapshot_month, snapshot_year=snapshot_year,
    )
    sort_map = {k: getattr(Employee, k) for k, _, _ in _EXPORT_COLUMNS}
    sort_col = sort_map.get(sort_by, Employee.full_name)
    q = q.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())
    employees = (await db.execute(q)).scalars().all()

    selected_keys = set(fields.split(",")) if fields else None
    cols = [c for c in _EXPORT_COLUMNS if selected_keys is None or c[0] in selected_keys]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font   = Font(name="Calibri", size=12, bold=True)
    header_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align  = Alignment(horizontal="center", vertical="center")
    data_font     = Font(name="Calibri", size=12)
    thin          = Side(style="thin")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_key, label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, e in enumerate(employees, 2):
        for col_idx, (key, _label, _width) in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_export_cell_value(e, key))
            cell.font = data_font
            cell.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"employee_data_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/monthly-summary")
async def get_monthly_summary(
    month: Optional[int] = Query(None, ge=1, le=12),
    year:  Optional[int] = Query(None),
    db:    AsyncSession = Depends(get_db),
    user:  CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Monthly headcount trend + demographic breakdowns.

    month/year are optional — when given, the KPI/breakdown numbers reflect
    the active-employee roster as of the end of that period (same
    join/resign-date windowing used by /turnover-summary and
    /target-vs-achievement) instead of today's live roster. Attributes like
    department/status/marital status are still each employee's *current*
    value (the Employee table isn't historized), only *which* employees are
    counted is period-aware."""
    import traceback
    from datetime import date
    from calendar import monthrange
    from sqlalchemy import and_, cast, String, literal_column

    MN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    try:
        # Monthly joinings map — cast date to text for grouping.
        # group_by references the SELECT alias ("month") rather than repeating
        # the to_char(...) expression: with a bound literal format string,
        # Postgres sees the SELECT and GROUP BY occurrences as two distinct
        # parameters ($1/$2) and can't prove they're equal, raising GroupingError.
        joins_q = await db.execute(
            select(
                func.to_char(Employee.date_of_joining, "YYYY-MM").label("month"),
                func.count().label("cnt"),
            )
            .where(Employee.date_of_joining.isnot(None))
            .group_by(literal_column("month"))
        )
        joins_map = {r[0]: r[1] for r in joins_q.fetchall()}
    except Exception as e:
        raise HTTPException(500, f"joins_q error: {e}\n{traceback.format_exc()}")

    try:
        # All join/resign pairs for cumulative headcount
        emps_q = await db.execute(
            select(Employee.user_id, Employee.date_of_joining, Employee.resign_date)
            .where(Employee.date_of_joining.isnot(None))
        )
        emps_full = [(r[0], r[1], r[2]) for r in emps_q.fetchall()]
        emps = [(join, resign) for _uid, join, resign in emps_full]
    except Exception as e:
        raise HTTPException(500, f"emps_q error: {e}")

    today = date.today()

    def _month_back(base_year, base_month, n):
        m = base_month - n
        y = base_year
        while m <= 0:
            m += 12
            y -= 1
        return y, m

    # Headcount trend — last 36 months
    headcount_trend = []
    for i in range(35, -1, -1):
        y, m = _month_back(today.year, today.month, i)
        last_day  = date(y, m, monthrange(y, m)[1])
        first_day = date(y, m, 1)
        cnt = sum(
            1 for join, resign in emps
            if join <= last_day and (resign is None or resign >= first_day)
        )
        headcount_trend.append({"month": f"{y}-{m:02d}", "label": f"{MN[m-1]} '{str(y)[2:]}", "count": cnt})

    # Monthly joinings — last 24 months
    monthly_joins = []
    for i in range(23, -1, -1):
        y, m = _month_back(today.year, today.month, i)
        key = f"{y}-{m:02d}"
        monthly_joins.append({"month": key, "label": f"{MN[m-1]} '{str(y)[2:]}", "joins": joins_map.get(key, 0)})

    # Period filter — when month/year is given, breakdowns/KPIs reflect the
    # active roster as of the end of that period instead of today's roster.
    # Attribute values (department, status, ...) are still each employee's
    # *current* value; only which employees get counted is period-aware.
    active_ids = None
    if year:
        snapshot_date = date(year, month or 12, monthrange(year, month or 12)[1])
        active_ids = {
            uid for uid, join, resign in emps_full
            if join <= snapshot_date and (resign is None or resign >= snapshot_date)
        }
    else:
        snapshot_date = today

    # Generic breakdown helper — explicit AND to avoid any dialect issues
    async def _bd(col):
        try:
            conditions = [col.isnot(None), col != ""]
            if active_ids is not None:
                conditions.append(Employee.user_id.in_(active_ids))
            q = await db.execute(
                select(col, func.count().label("total"))
                .where(and_(*conditions))
                .group_by(col)
                .order_by(func.count().desc())
            )
            return [{"name": r[0], "total": r[1]} for r in q.fetchall()]
        except Exception:
            return []

    by_dept    = await _bd(Employee.department)
    by_level   = await _bd(Employee.level)
    by_edu     = await _bd(Employee.education_degree)
    by_marital = await _bd(Employee.marital_status)
    by_status  = await _bd(Employee.status)
    by_grade   = await _bd(Employee.employee_grade)
    by_religion= await _bd(Employee.religion)

    try:
        sex_q = select(Employee.sex, func.count().label("t"))
        if active_ids is not None:
            sex_q = sex_q.where(Employee.user_id.in_(active_ids))
        sex_q = sex_q.group_by(Employee.sex)
        sex_result = await db.execute(sex_q)
        sex_map = {r[0]: r[1] for r in sex_result.fetchall()}
    except Exception:
        sex_map = {}

    by_gender = [
        {"name": "Male",   "total": sex_map.get("M", 0)},
        {"name": "Female", "total": sex_map.get("F", 0)},
    ]

    period_total = len(active_ids) if active_ids is not None else headcount_trend[-1]["count"]
    if year and month:
        period_joins = joins_map.get(f"{year}-{month:02d}", 0)
    elif year:
        period_joins = sum(v for k, v in joins_map.items() if k.startswith(f"{year}-"))
    else:
        period_joins = None
    period_label = (
        f"{MN[month-1]} {year}" if year and month else
        f"Year {year}" if year else
        "Current"
    )

    return {
        "headcount_trend": headcount_trend,
        "monthly_joins":   monthly_joins,
        "by_dept":         by_dept,
        "by_level":        by_level,
        "by_education":    by_edu,
        "by_marital":      by_marital,
        "by_status":       by_status,
        "by_grade":        by_grade,
        "by_religion":     by_religion,
        "by_gender":       by_gender,
        "period": {
            "year":  year,
            "month": month,
            "label": period_label,
            "snapshot_date": str(snapshot_date),
            "total_employees": period_total,
            "joins_in_month":  period_joins,
        },
    }


@router.get("/turnover-summary")
async def get_turnover_summary(
    year:       Optional[int] = Query(None),
    month:      Optional[int] = Query(None, ge=1, le=12),
    department: Optional[str] = Query(None),
    team:       Optional[str] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Laporan turnover: tren resign bulanan (Jan-Des tahun terpilih, default tahun
    berjalan), turnover rate, breakdown per departemen/level. `month` mempersempit
    breakdown & avg tenure ke bulan itu saja; tanpa `month`, breakdown mencakup
    satu tahun penuh. `department`/`team` mempersempit seluruh laporan."""
    from datetime import date
    from calendar import monthrange

    MN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    today = date.today()
    target_year = year or today.year

    q = select(Employee.date_of_joining, Employee.resign_date, Employee.department,
               Employee.level, Employee.status, Employee.team).where(Employee.date_of_joining.isnot(None))
    if department:
        q = q.where(Employee.department == department)
    if team:
        q = q.where(Employee.team == team)
    emps_q = await db.execute(q)
    emps = emps_q.fetchall()

    # Tren resign + turnover rate bulanan — Januari s/d Desember target_year
    resign_trend = []
    for m in range(1, 13):
        last_day  = date(target_year, m, monthrange(target_year, m)[1])
        first_day = date(target_year, m, 1)
        resigns_in_month = sum(
            1 for join, resign, *_ in emps
            if resign is not None and first_day <= resign <= last_day
        )
        headcount_start = sum(
            1 for join, resign, *_ in emps
            if join < first_day and (resign is None or resign >= first_day)
        )
        headcount_end = sum(
            1 for join, resign, *_ in emps
            if join <= last_day and (resign is None or resign >= first_day)
        )
        avg_headcount = (headcount_start + headcount_end) / 2 if (headcount_start + headcount_end) > 0 else 0
        rate = round((resigns_in_month / avg_headcount) * 100, 2) if avg_headcount > 0 else 0
        resign_trend.append({
            "month":          f"{target_year}-{m:02d}",
            "label":          MN[m - 1],
            "resigns":        resigns_in_month,
            "avg_headcount":  round(avg_headcount, 1),
            "turnover_rate":  rate,
        })

    # Turnover rate tahunan (selalu dihitung atas 12 bulan penuh target_year,
    # terlepas dari filter `month` — itu hanya mempersempit breakdown di bawah)
    total_resigns_year = sum(r["resigns"] for r in resign_trend)
    avg_headcount_year = sum(r["avg_headcount"] for r in resign_trend) / 12
    annual_turnover_rate = round((total_resigns_year / avg_headcount_year) * 100, 2) if avg_headcount_year > 0 else 0

    # Breakdown karyawan resign — satu bulan (jika `month` diisi) atau satu tahun penuh
    if month:
        b_start = date(target_year, month, 1)
        b_end   = date(target_year, month, monthrange(target_year, month)[1])
    else:
        b_start = date(target_year, 1, 1)
        b_end   = date(target_year, 12, 31)
    resigned_in_scope = [row for row in emps if row[1] is not None and b_start <= row[1] <= b_end]

    def _bd(idx):
        counts = {}
        for row in resigned_in_scope:
            key = row[idx] or "—"
            counts[key] = counts.get(key, 0) + 1
        return sorted(
            [{"name": k, "total": v} for k, v in counts.items()],
            key=lambda x: x["total"], reverse=True,
        )

    by_dept   = _bd(2)
    by_level  = _bd(3)
    by_status = _bd(4)

    # Rata-rata masa kerja karyawan yang resign (dalam scope breakdown di atas)
    tenures = [
        (row[1] - row[0]).days / 365.25
        for row in resigned_in_scope
        if row[0] and row[1]
    ]
    avg_tenure_years = round(sum(tenures) / len(tenures), 1) if tenures else 0

    current_headcount = sum(1 for join, resign, *_ in emps if resign is None)

    return {
        "year":                 target_year,
        "month":                month,
        "resign_trend":         resign_trend,
        "annual_turnover_rate": annual_turnover_rate,
        "total_resigns_period": len(resigned_in_scope),
        "avg_tenure_years":     avg_tenure_years,
        "current_headcount":    current_headcount,
        "by_dept":              by_dept,
        "by_level":             by_level,
        "by_status":            by_status,
    }


def _normalize_dept(raw: Optional[str]) -> str:
    """Same value stored with different casing (e.g. "Plant" / "PLANT") is a
    known data-quality issue in the source Excel — group them together here
    rather than showing near-duplicate rows in a pivot table."""
    if not raw or raw.strip().isdigit():
        return "—"
    return raw.strip()


# The business tracks headcount by exactly 4 top-level departments. The raw
# Employee.department column has case-duplicates ("Plant" / "PLANT") plus a
# handful of misfiled values that are really sub-functions of one of the 4 —
# confirmed empirically by cross-checking job_title/team for each (query run
# 2026-07-29): "Director" (President Director / Plant Director) -> the
# corporate/admin function; "Mkt & BD" (Product Executive) -> Sales &
# Marketing; "RA & BD" (Business Development / Regulatory Affairs Manager,
# Supervisor, Staff) -> matches Strategy & Development's existing
# Business Development / Regulatory Affairs teams; "Validation" (Validation
# Supervisor) -> matches Plant > Quality Management > Validation team.
DEPT_GROUPS = ["Administration", "Sales & Marketing", "Strategy & Development", "Plant"]

_DEPT_GROUP_MAP = {
    "ADMINISTRATION":        "Administration",
    "DIRECTOR":              "Administration",
    "SALES & MARKETING":     "Sales & Marketing",
    "MKT & BD":              "Sales & Marketing",
    "STRATEGY DEVELOPMENT":  "Strategy & Development",
    "RA & BD":               "Strategy & Development",
    "PLANT":                 "Plant",
    "VALIDATION":            "Plant",
}


def _group_department(raw: Optional[str]) -> Optional[str]:
    """Raw Employee.department value -> one of the 4 canonical DEPT_GROUPS,
    or None to exclude (blank/numeric-corrupted rows — same exclusion
    _normalize_dept already applied)."""
    if not raw or raw.strip().isdigit():
        return None
    return _DEPT_GROUP_MAP.get(raw.strip().upper())


def _dept_display_labels(raw_values) -> dict:
    """UPPER(dept) -> best display label — prefers a mixed-case variant
    ("Plant") over an all-caps one ("PLANT") when both exist."""
    groups: dict = {}
    for v in raw_values:
        groups.setdefault(v.upper(), []).append(v)
    display = {}
    for key, variants in groups.items():
        non_caps = [v for v in variants if v != v.upper()]
        display[key] = non_caps[0] if non_caps else variants[0]
    return display


@router.get("/summary/by-year")
async def get_summary_by_year(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Headcount by department (grouped into the 4 canonical DEPT_GROUPS),
    Beginning/Ending per year — same "active as of a date" windowing used
    by /turnover-summary and /monthly-summary."""
    rows_q = await db.execute(
        select(Employee.department, Employee.date_of_joining, Employee.resign_date)
        .where(Employee.date_of_joining.isnot(None))
    )
    emps = [(_group_department(d), j, r) for d, j, r in rows_q.fetchall()]
    emps = [(d, j, r) for d, j, r in emps if d is not None]

    today = date.today()
    year_from = min((j.year for _d, j, _r in emps), default=today.year)
    years = list(range(year_from, today.year + 1))

    departments = DEPT_GROUPS

    def active_count(snapshot, dept_filter=None):
        return sum(
            1 for d, j, r in emps
            if j <= snapshot and (r is None or r >= snapshot)
            and (dept_filter is None or d == dept_filter)
        )

    rows = []
    for label in DEPT_GROUPS:
        by_year = {}
        for y in years:
            beg_snapshot = date(y, 1, 1)
            end_snapshot = min(date(y, 12, 31), today)
            by_year[y] = {
                "beginning": active_count(beg_snapshot, label),
                "ending":    active_count(end_snapshot, label),
            }
        rows.append({"department": label, "by_year": by_year})

    total = {}
    for y in years:
        beg_snapshot = date(y, 1, 1)
        end_snapshot = min(date(y, 12, 31), today)
        total[y] = {"beginning": active_count(beg_snapshot), "ending": active_count(end_snapshot)}

    growth = {}
    for i, y in enumerate(years):
        growth[y] = None if i == 0 else total[y]["ending"] - total[years[i - 1]]["ending"]

    return {"years": years, "departments": departments, "rows": rows, "total": total, "growth": growth}


@router.get("/summary/by-month")
async def get_summary_by_month(
    year: Optional[int] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Headcount by department (grouped into the 4 canonical DEPT_GROUPS) >
    division > team, end-of-month snapshot for each month of the given year
    (default: current year). Division is only populated for some departments
    (currently just Plant) — departments without it go straight from
    department to team rows. Department/division/team are each employee's
    *current* value — the Employee table isn't historized — same caveat as
    /monthly-summary."""
    target_year = year or date.today().year
    rows_q = await db.execute(
        select(Employee.department, Employee.division, Employee.team,
               Employee.date_of_joining, Employee.resign_date)
        .where(Employee.date_of_joining.isnot(None))
    )
    emps = [
        (_group_department(d), (v or "").strip() or None, (t or "").strip() or None, j, r)
        for d, v, t, j, r in rows_q.fetchall()
    ]
    emps = [(d, v, t, j, r) for d, v, t, j, r in emps if d is not None]

    today = date.today()
    months = list(range(1, 13))

    def active_count(snapshot, dept_filter=None, division_filter=None, team_filter=None):
        return sum(
            1 for d, v, t, j, r in emps
            if j <= snapshot and (r is None or r >= snapshot)
            and (dept_filter is None or d == dept_filter)
            and (division_filter is None or v == division_filter)
            and (team_filter is None or t == team_filter)
        )

    def snapshot_for(m):
        return min(date(target_year, m, monthrange(target_year, m)[1]), today)

    def by_month_for(dept_filter=None, division_filter=None, team_filter=None):
        return {m: active_count(snapshot_for(m), dept_filter, division_filter, team_filter) for m in months}

    rows = []
    for label in DEPT_GROUPS:
        rows.append({"department": label, "division": None, "team": None, "by_month": by_month_for(label)})

        divisions_in_dept = sorted({v for d, v, _t, _j, _r in emps if d == label and v})
        teams_direct = sorted({t for d, v, t, _j, _r in emps if d == label and not v and t})

        for division in divisions_in_dept:
            rows.append({
                "department": label, "division": division, "team": None,
                "by_month": by_month_for(label, division),
            })
            teams_in_division = sorted({
                t for d, v, t, _j, _r in emps
                if d == label and v == division and t
            })
            for team in teams_in_division:
                rows.append({
                    "department": label, "division": division, "team": team,
                    "by_month": by_month_for(label, division, team),
                })

        for team in teams_direct:
            rows.append({
                "department": label, "division": None, "team": team,
                "by_month": by_month_for(label, None, team),
            })

    total = by_month_for()

    return {"year": target_year, "months": months, "rows": rows, "total": total}


@router.get("/upload-logs")
async def get_upload_logs(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Riwayat upload file Excel."""
    result = await db.execute(
        select(EmployeeUploadLog)
        .order_by(EmployeeUploadLog.uploaded_at.desc())
        .limit(20)
    )
    logs = result.scalars().all()
    return [
        {
            "batch_id":    l.batch_id,
            "filename":    l.filename,
            "total_rows":  l.total_rows,
            "inserted":    l.inserted,
            "updated":     l.updated,
            "uploaded_by": l.uploaded_by,
            "uploaded_at": l.uploaded_at.isoformat() if l.uploaded_at else None,
            "notes":       l.notes,
        }
        for l in logs
    ]


@router.get("/join-years")
async def get_join_years(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar tahun (dari date_of_joining) untuk filter dropdown Year."""
    result = await db.execute(
        select(extract("year", Employee.date_of_joining).label("y"))
        .where(Employee.date_of_joining.isnot(None))
        .distinct().order_by(extract("year", Employee.date_of_joining).desc())
    )
    return [int(r[0]) for r in result.fetchall() if r[0]]


@router.get("/departments")
async def get_departments(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar department untuk filter dropdown, langsung dari data karyawan.
    Beberapa baris sumber data punya department yang rusak (mis. angka '15'
    dari pergeseran kolom di file Excel asal) — nilai yang bukan nama (murni
    angka) disaring dari daftar filter, meski tetap tersimpan apa adanya di
    record karyawan itu sendiri. Case-duplicates ("Plant" / "PLANT") juga
    digabung ke satu nama tampilan (lihat _dept_display_labels) — filter
    department di list/export sudah case-insensitive, jadi memilih "Plant"
    tetap menangkap baris "PLANT" juga."""
    result = await db.execute(
        select(Employee.department).distinct()
    )
    raw = [r[0] for r in result.fetchall() if r[0] and not r[0].strip().isdigit()]
    return sorted(_dept_display_labels(raw).values())


@router.get("/teams")
async def get_teams(
    department: Optional[str] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar team untuk filter dropdown, opsional difilter per department."""
    q = select(Employee.team).distinct().order_by(Employee.team)
    if department:
        q = q.where(Employee.department == department)
    result = await db.execute(q)
    return [r[0] for r in result.fetchall() if r[0]]


@router.get("/educations")
async def get_educations(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar education degree untuk filter dropdown Employee Summary."""
    result = await db.execute(select(Employee.education_degree).distinct().order_by(Employee.education_degree))
    return [r[0] for r in result.fetchall() if r[0]]


@router.get("/positions")
async def get_positions(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar job title (posisi) untuk filter dropdown Employee Summary."""
    result = await db.execute(select(Employee.job_title).distinct().order_by(Employee.job_title))
    return [r[0] for r in result.fetchall() if r[0] and not r[0].strip().isdigit()]


@router.get("/levels")
async def get_levels(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar level untuk filter dropdown Employee Summary."""
    result = await db.execute(select(Employee.level).distinct().order_by(Employee.level))
    return [r[0] for r in result.fetchall() if r[0] and not r[0].strip().isdigit()]


@router.get("/marital-statuses")
async def get_marital_statuses(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Daftar marital status untuk filter dropdown Employee Summary."""
    result = await db.execute(select(Employee.marital_status).distinct().order_by(Employee.marital_status))
    return [r[0] for r in result.fetchall() if r[0]]


@router.get("/names")
async def get_employee_names(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Lightweight active-employee list for LOV/multi-select fields (e.g. To Do List Assigned To)."""
    result = await db.execute(
        select(Employee.user_id, Employee.full_name, Employee.department)
        .where(
            Employee.full_name.isnot(None),
            or_(Employee.employment_status == "Active",
                and_(Employee.employment_status.is_(None), Employee.resign_date.is_(None))),
        )
        .order_by(Employee.full_name)
    )
    return [{"user_id": r[0], "full_name": r[1], "department": r[2]} for r in result.fetchall()]


class SupervisorUpdate(BaseModel):
    supervisor_id: Optional[str] = None


@router.patch("/{user_id}/supervisor")
async def set_employee_supervisor(
    user_id: str,
    body: SupervisorUpdate,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Manual override for one employee's direct supervisor (used by the Organization
    Chart tab). Gets overwritten the next time the master data is re-uploaded, since
    upload REPLACEs the whole table and re-derives supervisor_id from scratch."""
    target = await db.scalar(select(Employee).where(Employee.user_id == user_id))
    if not target:
        raise HTTPException(status_code=404, detail=f"Employee {user_id} not found")

    new_sup_id = body.supervisor_id or None
    if new_sup_id:
        if new_sup_id == user_id:
            raise HTTPException(status_code=400, detail="An employee cannot be their own supervisor")
        supervisor = await db.scalar(select(Employee).where(Employee.user_id == new_sup_id))
        if not supervisor:
            raise HTTPException(status_code=404, detail=f"Supervisor {new_sup_id} not found")

        # Walk the chain upward from the proposed supervisor — if it reaches
        # user_id again, this assignment would create a reporting-line loop.
        cursor = supervisor.supervisor_id
        seen = {user_id}
        hops = 0
        while cursor and hops < 200:
            if cursor in seen:
                raise HTTPException(status_code=400, detail="This assignment would create a reporting-line loop")
            seen.add(cursor)
            cursor = await db.scalar(select(Employee.supervisor_id).where(Employee.user_id == cursor))
            hops += 1

    target.supervisor_id = new_sup_id
    await db.flush()
    return {"user_id": user_id, "supervisor_id": new_sup_id}


# ── Single-employee create/edit/resign — Employee List's detail popup ──────────

def _parse_iso_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


async def _check_supervisor_loop(db: AsyncSession, user_id: str, new_sup_id: str):
    """Raise if assigning new_sup_id as user_id's supervisor would create a cycle."""
    if new_sup_id == user_id:
        raise HTTPException(status_code=400, detail="An employee cannot be their own supervisor")
    supervisor = await db.scalar(select(Employee).where(Employee.user_id == new_sup_id))
    if not supervisor:
        raise HTTPException(status_code=404, detail=f"Supervisor {new_sup_id} not found")
    cursor = supervisor.supervisor_id
    seen = {user_id}
    hops = 0
    while cursor and hops < 200:
        if cursor in seen:
            raise HTTPException(status_code=400, detail="This assignment would create a reporting-line loop")
        seen.add(cursor)
        cursor = await db.scalar(select(Employee.supervisor_id).where(Employee.user_id == cursor))
        hops += 1


_EMPLOYEE_UPSERT_DATE_FIELDS = {
    "date_of_joining", "retire_date", "starting_pkwt", "end_pkwt",
    "permanent_date", "resign_date", "date_of_birth",
}
_EMPLOYEE_UPSERT_TEXT_FIELDS = [
    "full_name", "sex", "level", "department", "division", "team", "job_title",
    "work_placement", "status", "employment_status", "pkwt_ke", "place_of_birth",
    "no_bpjs_health", "no_bpjs_employee", "education_degree", "education_school",
    "education_major", "employee_grade", "supervisor_id", "working_experience_years",
    "previous_company", "address", "marital_status", "phone_number", "emergency_phone",
    "religion", "blood_type", "npwp_number", "bank_account_bca", "bank_account_name",
    "personal_email", "company_email", "resign_reason",
]


class EmployeeUpsert(BaseModel):
    user_id: Optional[str] = None
    full_name: Optional[str] = None
    sex: Optional[str] = None
    level: Optional[str] = None
    department: Optional[str] = None
    division: Optional[str] = None
    team: Optional[str] = None
    job_title: Optional[str] = None
    work_placement: Optional[str] = None
    status: Optional[str] = None
    employment_status: Optional[str] = None
    date_of_joining: Optional[str] = None
    retire_date: Optional[str] = None
    pkwt_ke: Optional[str] = None
    starting_pkwt: Optional[str] = None
    end_pkwt: Optional[str] = None
    permanent_date: Optional[str] = None
    resign_date: Optional[str] = None
    resign_reason: Optional[str] = None
    place_of_birth: Optional[str] = None
    date_of_birth: Optional[str] = None
    no_bpjs_health: Optional[str] = None
    no_bpjs_employee: Optional[str] = None
    education_degree: Optional[str] = None
    education_school: Optional[str] = None
    education_major: Optional[str] = None
    employee_grade: Optional[str] = None
    supervisor_id: Optional[str] = None
    working_experience_years: Optional[str] = None
    previous_company: Optional[str] = None
    address: Optional[str] = None
    marital_status: Optional[str] = None
    phone_number: Optional[str] = None
    emergency_phone: Optional[str] = None
    religion: Optional[str] = None
    blood_type: Optional[str] = None
    npwp_number: Optional[str] = None
    bank_account_bca: Optional[str] = None
    bank_account_name: Optional[str] = None
    personal_email: Optional[str] = None
    company_email: Optional[str] = None


def _apply_upsert_fields(target: Employee, body: EmployeeUpsert):
    for field in _EMPLOYEE_UPSERT_TEXT_FIELDS:
        setattr(target, field, getattr(body, field) or None)
    for field in _EMPLOYEE_UPSERT_DATE_FIELDS:
        setattr(target, field, _parse_iso_date(getattr(body, field)))


@router.post("")
async def create_employee(
    body: EmployeeUpsert,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Add a single new employee by hand (outside the Excel upload/replace flow)."""
    user_id = (body.user_id or "").strip()
    if not user_id:
        raise HTTPException(status_code=400, detail="NIK (user_id) is required")
    if not (body.full_name or "").strip():
        raise HTTPException(status_code=400, detail="Full name is required")

    existing = await db.scalar(select(Employee).where(Employee.user_id == user_id))
    if existing:
        raise HTTPException(status_code=409, detail=f"Employee {user_id} already exists")

    if body.supervisor_id:
        await _check_supervisor_loop(db, user_id, body.supervisor_id)

    emp = Employee(user_id=user_id, employment_status=body.employment_status or "Active")
    _apply_upsert_fields(emp, body)
    emp.user_id = user_id  # _apply_upsert_fields doesn't touch user_id — set explicitly
    db.add(emp)
    await db.flush()
    return _emp_dict(emp)


@router.put("/{user_id}")
async def update_employee(
    user_id: str,
    body: EmployeeUpsert,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Full edit of one employee's record from the Employee List detail popup."""
    target = await db.scalar(select(Employee).where(Employee.user_id == user_id))
    if not target:
        raise HTTPException(status_code=404, detail=f"Employee {user_id} not found")

    if body.supervisor_id and body.supervisor_id != target.supervisor_id:
        await _check_supervisor_loop(db, user_id, body.supervisor_id)

    _apply_upsert_fields(target, body)
    await db.flush()
    return _emp_dict(target)


class ResignRequest(BaseModel):
    resign_date: str
    reason: Optional[str] = None


@router.post("/{user_id}/resign")
async def resign_employee(
    user_id: str,
    body: ResignRequest,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Mark an employee as resigned — sets resign_date/reason and flips employment_status."""
    target = await db.scalar(select(Employee).where(Employee.user_id == user_id))
    if not target:
        raise HTTPException(status_code=404, detail=f"Employee {user_id} not found")

    resign_date = _parse_iso_date(body.resign_date)
    if not resign_date:
        raise HTTPException(status_code=400, detail="resign_date must be in YYYY-MM-DD format")

    target.resign_date = resign_date
    target.resign_reason = body.reason or None
    target.employment_status = "Resign"
    await db.flush()
    return _emp_dict(target)


@router.get("/org-chart")
async def get_org_chart(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Full company hierarchy, nested by supervisor_id, for the Organization Chart tab.
    Always returns a single root node — employees with no resolvable supervisor chain
    (shouldn't normally happen) are bucketed under a synthetic 'Unassigned' node
    instead of becoming extra floating roots."""
    result = await db.execute(
        select(
            Employee.user_id, Employee.full_name, Employee.job_title, Employee.level,
            Employee.department, Employee.division, Employee.team, Employee.sex,
            Employee.supervisor_id,
        ).where(
            or_(Employee.employment_status == "Active",
                and_(Employee.employment_status.is_(None), Employee.resign_date.is_(None))),
        ).order_by(Employee.full_name)
    )
    rows = [dict(r._mapping) for r in result.fetchall()]
    if not rows:
        return {"total": 0, "root": None}

    by_id = {r["user_id"]: {**r, "children": []} for r in rows}
    roots = []
    for r in rows:
        node = by_id[r["user_id"]]
        sup_id = r["supervisor_id"]
        if sup_id and sup_id in by_id and sup_id != r["user_id"]:
            by_id[sup_id]["children"].append(node)
        else:
            roots.append(node)

    def subtree_size(node):
        return 1 + sum(subtree_size(c) for c in node["children"])

    def finalize(node):
        node["children"].sort(key=lambda c: (-subtree_size(c), c["full_name"] or ""))
        node["direct_count"] = len(node["children"])
        for c in node["children"]:
            finalize(c)

    main_root = next((n for n in roots if (n["job_title"] or "").strip().lower() == "president director"), None)
    if main_root is None:
        roots.sort(key=lambda n: -len(n["children"]))
        main_root = roots[0] if roots else None

    orphans = [n for n in roots if n is not main_root]
    if main_root and orphans:
        main_root["children"].append({
            "user_id": "__unassigned__", "full_name": "Unassigned", "job_title": "No supervisor set",
            "level": None, "department": None, "division": None, "team": None, "sex": None,
            "supervisor_id": main_root["user_id"], "children": orphans,
        })

    if main_root:
        finalize(main_root)

    return {"total": len(rows), "root": main_root}


# NOTE: declared last — a path-param route shadows any static route registered
# after it (FastAPI matches in declaration order), so this must stay below
# /departments, /teams, /names, /org-chart, and /{user_id}/supervisor.
@router.get("/{user_id}")
async def get_employee(
    user_id: str,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Single employee's full record — used by the Organization Chart tab when a node is clicked."""
    emp = await db.scalar(select(Employee).where(Employee.user_id == user_id))
    if not emp:
        raise HTTPException(status_code=404, detail=f"Employee {user_id} not found")
    return _emp_dict(emp)
