"""
HR Attendance Router
Route prefix : /api/v1/dashboard/hr/attendance

Four upload sources, combined for every report/graph below:
  - Intercom (POST /upload)         — daily physical check-in/out log
                                       (e.g. "Attendance JUN-2026-Intercom.xlsx")
  - Talenta  (POST /upload-talenta) — leave & business-trip days
                                       (e.g. "Attendance MAY-JUN-2026 Talenta.xlsx")
  - Plant    (POST /upload-plant)   — combined physical + leave log for plant
                                       employees, one workbook with a sheet per
                                       month (e.g. "Attendance Plant 2026.xlsx")
  - Office   (POST /upload-office)  — combined physical + leave log for
                                       non-Plant/office staff, no department
                                       column (e.g. "Attendance 1-22 July 2025.xls")

Combination rule for every attendance-rate calculation below: a Business Trip
(BT) day counts as present (worked off-site); a leave day
(SL/AL/ALAB/ML/EM/UL/ULBB/EL/HD) is excluded from both the numerator and
denominator of the rate entirely — it isn't a required working day for that
person that day. A Plant shift rest day (is_day_off) is excluded the same
way. Everything else falls back to the Intercom/Plant attendance_status (or,
for legacy rows with no status, whether actual_checkin was recorded).
"""
import io
from datetime import datetime, date, time, timedelta
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, Query
from pydantic import BaseModel
from sqlalchemy import select, update, func, case, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import require_role, CurrentUser, Roles
from app.services.department_taxonomy import normalize_department, clean_department_list
from app.models.attendance import AttendanceRecord, AttendanceRecordAuditLog, AttendanceUploadLog

router = APIRouter()

WEEKENDS       = ["Saturday", "Sunday"]
# "H" = Holiday, a per-record calendar-holiday marker Talenta puts on the
# employee's row (distinct from WorkingCalendarHoliday) — treated the same
# as leave: not a required working day, excluded from both plan and actual.
# "EL"/"HD" (Event Leave / Half Day Leave) come from the Plant source's
# REMARK column — same treatment, excluded entirely rather than partially
# scored, since the model has no concept of a fractional working day.
LEAVE_CODES    = ("SL", "AL", "ALAB", "ML", "EM", "UL", "ULBB", "H", "EL", "HD")
BT_CODE        = "BT"
PRESENT_STATUSES = ("W", "L", "E", "LE")

LEAVE_LABELS = {
    "SL": "Sick Leave", "AL": "Annual Leave", "ALAB": "Annual Leave",
    "ML": "Maternity Leave", "EM": "Employee Marriage",
    "UL": "Unpaid Leave", "ULBB": "Unpaid Leave", "BT": "Business Trip",
    "H": "Holiday", "EL": "Event Leave", "HD": "Half Day Leave",
}

IS_WEEKDAY = AttendanceRecord.week_day.notin_(WEEKENDS)


# ── Shared attendance-rate scoring (single table, no join needed) ─────────────

def _is_leave_code():
    return AttendanceRecord.leave_code.in_(LEAVE_CODES)


def _is_bt_code():
    return AttendanceRecord.leave_code == BT_CODE


def _is_present_intercom():
    return or_(
        AttendanceRecord.attendance_status.in_(PRESENT_STATUSES),
        and_(AttendanceRecord.attendance_status.is_(None), AttendanceRecord.actual_checkin.isnot(None)),
    )


def _plan_expr():
    """1 if this is a required working day (weekday, not a Plant shift rest
    day, not on excluded leave)."""
    return case(
        (AttendanceRecord.is_day_off.is_(True), 0),
        (and_(IS_WEEKDAY, _is_leave_code()), 0),
        (IS_WEEKDAY, 1),
        else_=0,
    )


def _actual_expr():
    """1 if counted as present — Business Trip, or Intercom-present — and not
    on excluded leave or a Plant shift rest day."""
    return case(
        (AttendanceRecord.is_day_off.is_(True), 0),
        (and_(IS_WEEKDAY, _is_leave_code()), 0),
        (and_(IS_WEEKDAY, _is_bt_code()), 1),
        (and_(IS_WEEKDAY, _is_present_intercom()), 1),
        else_=0,
    )


def _read_data_rows(contents: bytes) -> list:
    """Read every data row (from row 2 onward, values only) from an
    .xlsx/.xlsm file's first worksheet. Tries openpyxl first (normal, then
    read_only mode), then falls back to python-calamine — a Rust-based
    parser that only reads raw cell values and ignores styles/filters/data
    validation entirely — for files openpyxl can't parse at all. Real
    Talenta exports have hit both classes of failure: minor OOXML
    non-compliance, and a malformed AutoFilter definition openpyxl's strict
    parser refuses to load even in read_only mode ("Value must be either
    numerical or a string containing a wildcard")."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        return list(wb.active.iter_rows(min_row=2, values_only=True))
    except Exception:
        pass
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        return list(wb.active.iter_rows(min_row=2, values_only=True))
    except Exception:
        pass
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_filelike(io.BytesIO(contents))
    rows = wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
    return [tuple(r) for r in rows[1:]]


def _to_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("", "-", "None") else s


def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ── Intercom upload: daily physical check-in/out log ──────────────────────────
# Header row 1, data from row 2. Real export column layout (0-based):
#   Name | ID | Department | Team | Date | Week | Time Period |
#   Required Check-In Time | Required Check-Out Time |
#   Actual Check-In Time | Actual Check-Out Time | Attendance Records |
#   Required Work Hours | Total Work Hours | Attendance Status | ...
COL_NAME       = 0
COL_ID         = 1
COL_DEPT       = 2
COL_TEAM       = 3
COL_DATE       = 4
COL_WEEK       = 5
COL_PERIOD     = 6
COL_SCHED_IN   = 7
COL_SCHED_OUT  = 8
COL_ACTUAL_IN  = 9
COL_ACTUAL_OUT = 10
COL_STATUS     = 14  # W=Worked, L=Late, E=Early leave, LE=Late+Early, A=Absent


def _parse_intercom_row(row: tuple, batch_id: str) -> Optional[dict]:
    if len(row) <= COL_ID:
        return None
    employee_id = _to_str(row[COL_ID])
    if not employee_id:
        return None

    att_date = _to_date(row[COL_DATE]) if len(row) > COL_DATE else None
    if att_date is None:
        return None

    return {
        "employee_id":        employee_id,
        "employee_name":      _to_str(row[COL_NAME])      if len(row) > COL_NAME      else None,
        # The Intercom export's literal "Department" column is always the
        # placeholder "All Departments" — the real department name is in
        # "Team" instead, so that's what feeds our `department` field. The
        # literal Department column carries no real information, so `team`
        # (our model column) is left unset for Intercom-sourced rows.
        "department":         _to_str(row[COL_TEAM])      if len(row) > COL_TEAM      else None,
        "team":               None,
        "attendance_date":    att_date,
        "week_day":           _to_str(row[COL_WEEK])      if len(row) > COL_WEEK      else None,
        "time_period":        _to_str(row[COL_PERIOD])    if len(row) > COL_PERIOD    else None,
        "scheduled_checkin":  _to_str(row[COL_SCHED_IN])  if len(row) > COL_SCHED_IN  else None,
        "scheduled_checkout": _to_str(row[COL_SCHED_OUT]) if len(row) > COL_SCHED_OUT else None,
        "actual_checkin":     _to_str(row[COL_ACTUAL_IN]) if len(row) > COL_ACTUAL_IN else None,
        "actual_checkout":    _to_str(row[COL_ACTUAL_OUT])if len(row) > COL_ACTUAL_OUT else None,
        "attendance_status":  _to_str(row[COL_STATUS])    if len(row) > COL_STATUS    else None,
        "source":             "intercom",
        "upload_batch_id":    batch_id,
        "uploaded_at":        datetime.utcnow(),
    }


@router.post("/upload")
async def upload_attendance(
    file:  UploadFile    = File(...),
    notes: str           = Form(""),
    db:    AsyncSession  = Depends(get_db),
    user:  CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Upload the Intercom daily attendance export. Header row 1, data from
    row 2. UPSERT by employee_id + attendance_date."""
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xlsm format")

    contents = await file.read()
    batch_id = f"att_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    try:
        data_rows = _read_data_rows(contents)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {e}. Try opening it in Excel and doing "
                   f"File > Save As > .xlsx to normalize the file, then re-upload.",
        )

    rows_parsed = [
        r for r in (_parse_intercom_row(row, batch_id) for row in data_rows)
        if r is not None
    ]

    if not rows_parsed:
        raise HTTPException(
            status_code=422,
            detail="No attendance data found. Make sure the file matches the Intercom "
                   "export template (header row 1, data starting row 2).",
        )

    # Normalize department against the Employee master (source of truth) so
    # department names are consistent across Intercom/Talenta/reports —
    # otherwise casing differences between sources (e.g. "Administration" vs
    # "ADMINISTRATION") would show up as duplicate near-identical department
    # bars in every chart.
    from app.models.employee import Employee
    emp_dept_q = await db.execute(select(Employee.user_id, Employee.department))
    emp_dept_map = {r[0]: r[1] for r in emp_dept_q.fetchall() if r[1]}
    for data in rows_parsed:
        if data["employee_id"] in emp_dept_map:
            data["department"] = emp_dept_map[data["employee_id"]]
        else:
            normalized = normalize_department(data["department"])
            if normalized:
                data["department"] = normalized

    keys_result = await db.execute(select(AttendanceRecord.employee_id, AttendanceRecord.attendance_date))
    existing_keys = {(r[0], r[1]) for r in keys_result.fetchall()}

    inserted = updated = 0
    for data in rows_parsed:
        key = (data["employee_id"], data["attendance_date"])
        if key in existing_keys:
            await db.execute(
                update(AttendanceRecord)
                .where(AttendanceRecord.employee_id == key[0])
                .where(AttendanceRecord.attendance_date == key[1])
                .values(**data)
            )
            updated += 1
        else:
            db.add(AttendanceRecord(**data))
            inserted += 1

    await db.flush()

    log = AttendanceUploadLog(
        batch_id=batch_id, source="intercom", filename=file.filename,
        total_rows=len(rows_parsed), inserted=inserted, updated=updated, skipped=0,
        uploaded_by=user.username or "unknown", notes=notes or None,
    )
    db.add(log)

    return {
        "batch_id": batch_id, "filename": file.filename, "source": "intercom",
        "total_rows": len(rows_parsed), "inserted": inserted, "updated": updated, "skipped": 0,
        "message": f"Upload successful: {inserted} new records, {updated} updated.",
    }


# ── Talenta upload: leave & business-trip days ─────────────────────────────────
# Real export column layout (0-based):
#   Employee ID | Full Name | Branch | department | Job Position | Date | Shift |
#   Shift Code | Shift Label | Schedule Check In | Schedule Check Out |
#   Attendance Code | Time Off Code | ...
TCOL_ID       = 0
TCOL_NAME     = 1
TCOL_DEPT     = 3
TCOL_DATE     = 5
TCOL_ATT_CODE = 11
TCOL_TIMEOFF  = 12


def _parse_talenta_row(row: tuple, batch_id: str) -> Optional[dict]:
    if len(row) <= TCOL_ID:
        return None
    employee_id = _to_str(row[TCOL_ID])
    if not employee_id:
        return None

    att_date = _to_date(row[TCOL_DATE]) if len(row) > TCOL_DATE else None
    if att_date is None:
        return None

    attendance_code = _to_str(row[TCOL_ATT_CODE]) if len(row) > TCOL_ATT_CODE else None
    time_off_code   = _to_str(row[TCOL_TIMEOFF])  if len(row) > TCOL_TIMEOFF  else None
    leave_code = attendance_code or time_off_code
    if not leave_code:
        return None  # regular working day — nothing relevant to attendance-ratio scoring

    return {
        "employee_id":     employee_id,
        "employee_name":   _to_str(row[TCOL_NAME]) if len(row) > TCOL_NAME else None,
        "department":      _to_str(row[TCOL_DEPT]) if len(row) > TCOL_DEPT else None,
        "attendance_date": att_date,
        "leave_code":      leave_code,
        "source":          "talenta",
        "upload_batch_id": batch_id,
        "uploaded_at":     datetime.utcnow(),
    }


@router.post("/upload-talenta")
async def upload_attendance_talenta(
    file:  UploadFile    = File(...),
    notes: str           = Form(""),
    db:    AsyncSession  = Depends(get_db),
    user:  CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Upload the Talenta leave/business-trip export. Only rows with an
    Attendance Code or Time Off Code are stored — these override how that
    employee-day is scored in every attendance report (see module docstring).
    UPSERT by employee_id + attendance_date."""
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xlsm format")

    contents = await file.read()
    batch_id = f"tal_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    try:
        data_rows = _read_data_rows(contents)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {e}. Try opening it in Excel and doing "
                   f"File > Save As > .xlsx to normalize the file, then re-upload.",
        )

    rows_parsed = [
        r for r in (_parse_talenta_row(row, batch_id) for row in data_rows)
        if r is not None
    ]

    if not rows_parsed:
        raise HTTPException(
            status_code=422,
            detail="No leave/business-trip rows found. Make sure the file matches the "
                   "Talenta export template and has Attendance Code / Time Off Code filled in.",
        )

    # Same department normalization as the Intercom upload — Talenta's
    # department names are uppercase ("ADMINISTRATION"), which would
    # otherwise show up as a duplicate of the Employee-master-cased version.
    from app.models.employee import Employee
    emp_dept_q = await db.execute(select(Employee.user_id, Employee.department))
    emp_dept_map = {r[0]: r[1] for r in emp_dept_q.fetchall() if r[1]}
    for data in rows_parsed:
        if data["employee_id"] in emp_dept_map:
            data["department"] = emp_dept_map[data["employee_id"]]
        else:
            normalized = normalize_department(data["department"])
            if normalized:
                data["department"] = normalized

    # Same table as every other source now — if Intercom never covered this
    # date for this employee (e.g. they were on leave the whole time and
    # never had a badge scan), the row is created here instead of updated.
    ar_keys_result = await db.execute(select(AttendanceRecord.employee_id, AttendanceRecord.attendance_date))
    existing_keys = {(r[0], r[1]) for r in ar_keys_result.fetchall()}

    inserted = updated = 0
    for data in rows_parsed:
        key = (data["employee_id"], data["attendance_date"])
        if key in existing_keys:
            await db.execute(
                update(AttendanceRecord)
                .where(AttendanceRecord.employee_id == key[0])
                .where(AttendanceRecord.attendance_date == key[1])
                .values(**data)
            )
            updated += 1
        else:
            d = data["attendance_date"]
            db.add(AttendanceRecord(**data, week_day=d.strftime("%A")))
            existing_keys.add(key)
            inserted += 1

    await db.flush()

    log = AttendanceUploadLog(
        batch_id=batch_id, source="talenta", filename=file.filename,
        total_rows=len(rows_parsed), inserted=inserted, updated=updated, skipped=0,
        uploaded_by=user.username or "unknown", notes=notes or None,
    )
    db.add(log)

    return {
        "batch_id": batch_id, "filename": file.filename, "source": "talenta",
        "total_rows": len(rows_parsed), "inserted": inserted, "updated": updated, "skipped": 0,
        "message": f"Upload successful: {inserted} new leave/BT records, {updated} updated.",
    }


# ── Plant upload: daily attendance + leave, combined in one row ────────────────
# Real export layout: one workbook, one sheet per month (JAN..DEC), title in
# row 1, header in row 2, data from row 3. Every sheet's DATE column carries
# a full date, so all sheets in the workbook are read and merged — no need
# to know which sheet is "current". Column layout (0-based, within a row):
#   NO | TEAM | Employee ID | Employee Name | Day | Date | On Duty | Off Duty |
#   Start Overtime | Clock In | Clock Out | Late (duration) | Late (status) |
#   Remark | ...(unrelated join/leaver announcements in the trailing columns)
PCOL_TEAM         = 1
PCOL_ID           = 2
PCOL_NAME         = 3
PCOL_DAY          = 4
PCOL_DATE         = 5
PCOL_ON_DUTY      = 6
PCOL_OFF_DUTY     = 7
PCOL_CLOCK_IN     = 9
PCOL_CLOCK_OUT    = 10
PCOL_LATE_STATUS  = 12
PCOL_REMARK       = 13

# REMARK text -> the same short leave/BT codes Talenta uses, so Plant rows
# plug into the exact same _is_leave_code()/_is_bt_code() scoring above.
# COLLECTIVE LEAVE maps to "H" — the same per-record calendar-holiday code
# Talenta already uses (see LEAVE_CODES above). LATE ATTEND, OVERTIME, BACK
# TO HOME EARLY, and RDO are handled separately below (status / day-off,
# not a leave event).
PLANT_REMARK_LEAVE_MAP = {
    "ANNUAL LEAVE":     "AL",
    "SICK LEAVE":       "SL",
    "UNPAID LEAVE":     "UL",
    "BUSINESS TRIP":    "BT",
    "EVENT LEAVE":      "EL",
    "HALF DAY LEAVE":   "HD",
    "COLLECTIVE LEAVE": "H",
}


def _read_plant_rows(contents: bytes) -> list:
    """Read every data row (from row 3 onward — row 1 title, row 2 header)
    across every worksheet in the Plant workbook (one sheet per month)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        return [r for ws in wb.worksheets for r in ws.iter_rows(min_row=3, values_only=True)]
    except Exception:
        pass
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        return [r for ws in wb.worksheets for r in ws.iter_rows(min_row=3, values_only=True)]
    except Exception:
        pass
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_filelike(io.BytesIO(contents))
    rows = []
    for name in wb.sheet_names:
        sheet_rows = wb.get_sheet_by_name(name).to_python()
        rows.extend(tuple(r) for r in sheet_rows[2:])  # skip title + header rows
    return rows


def _plant_time_str(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        val = val.time()
    if isinstance(val, time):
        return val.strftime("%H:%M")
    s = str(val).strip()
    return s or None


def _is_off_marker(val) -> bool:
    return isinstance(val, str) and val.strip().upper() == "OFF"


def _parse_plant_row(row: tuple, batch_id: str) -> Optional[dict]:
    if len(row) <= PCOL_ID:
        return None
    employee_id = _to_str(row[PCOL_ID])
    if not employee_id:
        return None

    att_date = _to_date(row[PCOL_DATE]) if len(row) > PCOL_DATE else None
    if att_date is None:
        return None

    on_duty_raw   = row[PCOL_ON_DUTY]   if len(row) > PCOL_ON_DUTY   else None
    clock_in_raw  = row[PCOL_CLOCK_IN]  if len(row) > PCOL_CLOCK_IN  else None
    clock_out_raw = row[PCOL_CLOCK_OUT] if len(row) > PCOL_CLOCK_OUT else None
    late_status   = _to_str(row[PCOL_LATE_STATUS]) if len(row) > PCOL_LATE_STATUS else None
    remark        = _to_str(row[PCOL_REMARK])       if len(row) > PCOL_REMARK       else None

    remark_upper = remark.upper() if remark else None
    # RDO (Rostered Day Off) is a scheduled rest day communicated only via
    # REMARK, not the On Duty column — same treatment as an explicit "OFF".
    is_day_off = _is_off_marker(on_duty_raw) or remark_upper == "RDO"
    actual_checkin  = None if _is_off_marker(clock_in_raw)  else _plant_time_str(clock_in_raw)
    actual_checkout = None if _is_off_marker(clock_out_raw) else _plant_time_str(clock_out_raw)

    remark_code = PLANT_REMARK_LEAVE_MAP.get(remark_upper) if remark_upper else None

    if is_day_off:
        attendance_status = None
    elif late_status and late_status.upper() == "LATE ATTEND":
        attendance_status = "L"
    elif remark_upper == "BACK TO HOME EARLY":
        attendance_status = "E"
    elif remark_code is None and actual_checkin is None and actual_checkout is None:
        attendance_status = "A"
    else:
        attendance_status = None

    ar_data = {
        "employee_id":        employee_id,
        "employee_name":      _to_str(row[PCOL_NAME]) if len(row) > PCOL_NAME else None,
        "department":         _to_str(row[PCOL_TEAM]) if len(row) > PCOL_TEAM else None,
        "team":               None,
        "attendance_date":    att_date,
        "week_day":           _to_str(row[PCOL_DAY]) if len(row) > PCOL_DAY else None,
        "time_period":        None,
        "scheduled_checkin":  None if is_day_off else _plant_time_str(on_duty_raw),
        "scheduled_checkout": None if is_day_off else _plant_time_str(row[PCOL_OFF_DUTY] if len(row) > PCOL_OFF_DUTY else None),
        "actual_checkin":     actual_checkin,
        "actual_checkout":    actual_checkout,
        "attendance_status":  attendance_status,
        "notes":              f"Plant remark: {remark}" if remark else None,
        "is_day_off":         is_day_off,
        "leave_code":         remark_code,
        "source":             "plant",
        "upload_batch_id":    batch_id,
        "uploaded_at":        datetime.utcnow(),
    }

    return ar_data


@router.post("/upload-plant")
async def upload_attendance_plant(
    file:  UploadFile    = File(...),
    notes: str           = Form(""),
    db:    AsyncSession  = Depends(get_db),
    user:  CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Upload the Plant attendance workbook (one sheet per month, all sheets
    read and merged). Each row carries both the physical check-in/out AND
    the leave/BT remark at once, written to AttendanceRecord in a single
    upsert. UPSERT by employee_id + attendance_date."""
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xlsm format")

    contents = await file.read()
    batch_id = f"plt_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    try:
        data_rows = _read_plant_rows(contents)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {e}. Try opening it in Excel and doing "
                   f"File > Save As > .xlsx to normalize the file, then re-upload.",
        )

    rows_parsed = [_parse_plant_row(row, batch_id) for row in data_rows]
    rows_parsed = [r for r in rows_parsed if r is not None]

    if not rows_parsed:
        raise HTTPException(
            status_code=422,
            detail="No attendance rows found. Make sure the file matches the Plant "
                   "export template (title row 1, header row 2, data from row 3).",
        )

    leave_rows = sum(1 for r in rows_parsed if r["leave_code"])

    # Same department normalization as Intercom/Talenta.
    from app.models.employee import Employee
    emp_dept_q = await db.execute(select(Employee.user_id, Employee.department))
    emp_dept_map = {r[0]: r[1] for r in emp_dept_q.fetchall() if r[1]}
    for data in rows_parsed:
        if data["employee_id"] in emp_dept_map:
            data["department"] = emp_dept_map[data["employee_id"]]
        else:
            normalized = normalize_department(data["department"])
            if normalized:
                data["department"] = normalized

    ar_keys_result = await db.execute(select(AttendanceRecord.employee_id, AttendanceRecord.attendance_date))
    existing_keys = {(r[0], r[1]) for r in ar_keys_result.fetchall()}

    inserted = updated = 0
    for data in rows_parsed:
        key = (data["employee_id"], data["attendance_date"])
        if key in existing_keys:
            await db.execute(
                update(AttendanceRecord)
                .where(AttendanceRecord.employee_id == key[0])
                .where(AttendanceRecord.attendance_date == key[1])
                .values(**data)
            )
            updated += 1
        else:
            db.add(AttendanceRecord(**data))
            existing_keys.add(key)
            inserted += 1

    await db.flush()

    log = AttendanceUploadLog(
        batch_id=batch_id, source="plant", filename=file.filename,
        total_rows=len(rows_parsed), inserted=inserted, updated=updated, skipped=0,
        uploaded_by=user.username or "unknown", notes=notes or None,
    )
    db.add(log)

    return {
        "batch_id": batch_id, "filename": file.filename, "source": "plant",
        "total_rows": len(rows_parsed), "inserted": inserted, "updated": updated, "skipped": 0,
        "leave_rows": leave_rows,
        "message": f"Upload successful: {inserted} new attendance records, {updated} updated "
                   f"(including {leave_rows} leave/BT days).",
    }


# ── Office upload: combined attendance + leave, no department/team column ──────
# Real export layout: title/header vary slightly per sheet — every sheet in
# the workbook is read and its own header row 1 is used to locate columns
# (real files have shown both an 9-10 column layout with a "Timetable"
# column, and an 8-column layout without it, e.g. a per-employee addendum
# sheet). Data from row 2. Columns identified by header text (case-
# insensitive), not fixed position:
#   No. | Name | Date | [Timetable] | On duty | Off duty | Clock In | Clock Out | Remarks
# No Department/Team column at all — department always comes from the
# Employee master (this source never has a raw value to fall back to).
# Remarks are free-text English phrases, not short codes — mapped below.
# Real files seen so far are legacy .xls, which openpyxl cannot open at
# all, so python-calamine is used directly (not as a last-resort fallback
# like the other upload endpoints).
OFFICE_REMARK_EXACT = {
    "ANNUAL LEAVE": "AL",
    "SICK LEAVE":   "SL",
    "UNPAID LEAVE": "UL",
    "MATERNITY LEAVE": "ML",
    "BT": "BT",
}


def _office_remark_code(remark_upper: Optional[str]):
    """Returns (leave_code_or_None, is_day_off)."""
    if not remark_upper:
        return None, False
    if "HALF DAY" in remark_upper or "AFTER LUNCH" in remark_upper or "BEFORE LUNCH" in remark_upper:
        return "HD", False
    if remark_upper == "REPLACEMENT DAY OFF":
        return None, True
    return OFFICE_REMARK_EXACT.get(remark_upper), False


def _office_find_col(header: list, name: str) -> Optional[int]:
    name = name.lower()
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def _read_office_rows(contents: bytes) -> list:
    """Returns a list of (row, col_map) pairs — col_map is per-sheet since
    real files have shown sheets with different column layouts."""
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_filelike(io.BytesIO(contents))
    rows = []
    for name in wb.sheet_names:
        sheet_rows = wb.get_sheet_by_name(name).to_python()
        if not sheet_rows:
            continue
        header = [str(h).strip().lower() if h is not None else "" for h in sheet_rows[0]]
        col = {
            "id":          _office_find_col(header, "no."),
            "name":        _office_find_col(header, "name"),
            "date":        _office_find_col(header, "date"),
            "period":      _office_find_col(header, "timetable"),
            "sched_in":    _office_find_col(header, "on duty"),
            "sched_out":   _office_find_col(header, "off duty"),
            "actual_in":   _office_find_col(header, "clock in"),
            "actual_out":  _office_find_col(header, "clock out"),
            "remark":      _office_find_col(header, "remarks"),
        }
        if col["id"] is None or col["date"] is None:
            continue  # not a recognizable data sheet
        for row in sheet_rows[1:]:
            rows.append((row, col))
    return rows


def _to_office_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # Excel serial date (1899-12-30 epoch, incl. Excel's leap-year bug)
        try:
            return date(1899, 12, 30) + timedelta(days=int(val))
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_office_time_str(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        return val.time().strftime("%H:%M")
    s = str(val).strip()
    if not s or s.upper() == "OFF":
        return None
    return s


def _parse_office_row(row: tuple, col: dict, batch_id: str) -> Optional[dict]:
    def get(key):
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    employee_id = _to_str(get("id"))
    if not employee_id:
        return None

    att_date = _to_office_date(get("date"))
    if att_date is None:
        return None

    remark = _to_str(get("remark"))
    remark_upper = remark.upper() if remark else None
    leave_code, is_day_off = _office_remark_code(remark_upper)

    actual_checkin  = _to_office_time_str(get("actual_in"))
    actual_checkout = _to_office_time_str(get("actual_out"))

    if is_day_off or leave_code:
        attendance_status = None
    elif not actual_checkin and not actual_checkout and not remark:
        attendance_status = "A"
    else:
        attendance_status = None

    notes = f"Office remark: {remark}" if (remark and not leave_code and not is_day_off) else None

    return {
        "employee_id":        employee_id,
        "employee_name":      _to_str(get("name")),
        "department":         None,
        "team":               None,
        "attendance_date":    att_date,
        "week_day":           att_date.strftime("%A"),
        "time_period":        _to_str(get("period")),
        "scheduled_checkin":  _to_office_time_str(get("sched_in")),
        "scheduled_checkout": _to_office_time_str(get("sched_out")),
        "actual_checkin":     actual_checkin,
        "actual_checkout":    actual_checkout,
        "attendance_status":  attendance_status,
        "notes":              notes,
        "is_day_off":         is_day_off,
        "leave_code":         leave_code,
        "source":             "office",
        "upload_batch_id":    batch_id,
        "uploaded_at":        datetime.utcnow(),
    }


@router.post("/upload-office")
async def upload_attendance_office(
    file:  UploadFile    = File(...),
    notes: str           = Form(""),
    db:    AsyncSession  = Depends(get_db),
    user:  CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Upload the Office attendance export (no Department/Team column —
    department always comes from the Employee master). UPSERT by
    employee_id + attendance_date."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls format")

    contents = await file.read()
    batch_id = f"off_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    try:
        data_rows = _read_office_rows(contents)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file: {e}. Try opening it in Excel and doing "
                   f"File > Save As > .xlsx to normalize the file, then re-upload.",
        )

    rows_parsed = []
    skipped_no_id = 0
    skipped_names = set()
    for row, col in data_rows:
        parsed = _parse_office_row(row, col, batch_id)
        if parsed is None:
            skipped_no_id += 1
            id_idx = col.get("id")
            has_id = id_idx is not None and id_idx < len(row) and _to_str(row[id_idx])
            if not has_id:
                name_idx = col.get("name")
                if name_idx is not None and name_idx < len(row):
                    nm = _to_str(row[name_idx])
                    if nm:
                        skipped_names.add(nm)
        else:
            rows_parsed.append(parsed)

    if not rows_parsed:
        raise HTTPException(
            status_code=422,
            detail="No attendance rows found. Make sure the file matches the Office "
                   "export template (header row 1, data from row 2, Employee ID in the 'No.' column).",
        )

    # This source has no department/team column at all — always from the
    # Employee master, no raw fallback to normalize.
    from app.models.employee import Employee
    emp_dept_q = await db.execute(select(Employee.user_id, Employee.department))
    emp_dept_map = {r[0]: r[1] for r in emp_dept_q.fetchall() if r[1]}
    for data in rows_parsed:
        if data["employee_id"] in emp_dept_map:
            data["department"] = emp_dept_map[data["employee_id"]]

    ar_keys_result = await db.execute(select(AttendanceRecord.employee_id, AttendanceRecord.attendance_date))
    existing_keys = {(r[0], r[1]) for r in ar_keys_result.fetchall()}

    inserted = updated = 0
    for data in rows_parsed:
        key = (data["employee_id"], data["attendance_date"])
        if key in existing_keys:
            await db.execute(
                update(AttendanceRecord)
                .where(AttendanceRecord.employee_id == key[0])
                .where(AttendanceRecord.attendance_date == key[1])
                .values(**data)
            )
            updated += 1
        else:
            db.add(AttendanceRecord(**data))
            existing_keys.add(key)
            inserted += 1

    await db.flush()

    log = AttendanceUploadLog(
        batch_id=batch_id, source="office", filename=file.filename,
        total_rows=len(rows_parsed), inserted=inserted, updated=updated, skipped=skipped_no_id,
        uploaded_by=user.username or "unknown", notes=notes or None,
    )
    db.add(log)

    skip_note = f", {skipped_no_id} row(s) skipped (missing Employee ID)" if skipped_no_id else ""
    return {
        "batch_id": batch_id, "filename": file.filename, "source": "office",
        "total_rows": len(rows_parsed), "inserted": inserted, "updated": updated, "skipped": skipped_no_id,
        "skipped_names": sorted(skipped_names),
        "message": f"Upload successful: {inserted} new records, {updated} updated{skip_note}.",
    }


# ── Upload logs ────────────────────────────────────────────────────────────────

@router.get("/upload-logs")
async def get_upload_logs(
    source: Optional[str] = Query(None),
    db:     AsyncSession  = Depends(get_db),
    user:   CurrentUser   = Depends(require_role(Roles.HR)),
):
    q = select(AttendanceUploadLog)
    if source:
        q = q.where(AttendanceUploadLog.source == source)
    q = q.order_by(AttendanceUploadLog.uploaded_at.desc()).limit(20)
    result = await db.execute(q)
    logs = result.scalars().all()
    return [
        {
            "batch_id":    l.batch_id,
            "source":      l.source,
            "filename":    l.filename,
            "total_rows":  l.total_rows,
            "inserted":    l.inserted,
            "updated":     l.updated,
            "uploaded_by": l.uploaded_by,
            "uploaded_at": l.uploaded_at.isoformat() if l.uploaded_at else None,
            "notes":       l.notes,
        }
        for l in logs
    ]


# ── List attendance records ────────────────────────────────────────────────────

@router.get("")
async def list_attendance(
    employee_id: Optional[str] = Query(None),
    department:  Optional[str] = Query(None),
    date_from:   Optional[str] = Query(None),
    date_to:     Optional[str] = Query(None),
    page:        int           = Query(1, ge=1),
    page_size:   int           = Query(50, ge=1, le=200),
    db:          AsyncSession  = Depends(get_db),
    user:        CurrentUser   = Depends(require_role(Roles.HR)),
):
    q = select(AttendanceRecord)

    if employee_id:
        q = q.where(AttendanceRecord.employee_id == employee_id)
    if department:
        q = q.where(AttendanceRecord.department == department)
    if date_from:
        try:
            q = q.where(AttendanceRecord.attendance_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
    if date_to:
        try:
            q = q.where(AttendanceRecord.attendance_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass

    count_q = await db.execute(select(func.count()).select_from(q.subquery()))
    total   = count_q.scalar() or 0

    q       = q.order_by(AttendanceRecord.attendance_date.desc(), AttendanceRecord.employee_name)
    q       = q.offset((page - 1) * page_size).limit(page_size)
    result  = await db.execute(q)
    records = result.scalars().all()

    def _rec(r: AttendanceRecord) -> dict:
        return {
            "id":                 r.id,
            "employee_id":        r.employee_id,
            "employee_name":      r.employee_name,
            "department":         r.department,
            "team":               r.team,
            "attendance_date":    str(r.attendance_date) if r.attendance_date else None,
            "week_day":           r.week_day,
            "time_period":        r.time_period,
            "scheduled_checkin":  r.scheduled_checkin,
            "scheduled_checkout": r.scheduled_checkout,
            "actual_checkin":     r.actual_checkin,
            "actual_checkout":    r.actual_checkout,
            "attendance_status":  r.attendance_status,
            "notes":              r.notes,
            "leave_code":         r.leave_code,
            "is_day_off":         r.is_day_off,
            "source":             r.source,
        }

    return {
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "pages":     (total + page_size - 1) // page_size,
        "records":   [_rec(r) for r in records],
    }


# ── Manual data correction ─────────────────────────────────────────────────────
# Every source above (Intercom/Talenta/Plant) upserts into AttendanceRecord;
# this lets HR do the same thing by hand — fixing a wrong value, or creating
# a day that was never covered by any upload at all (e.g. a manual leave
# entry). A later Excel upload for the same employee+date is still free to
# overwrite whatever is set here — there's no lock — but every manual change
# is logged to AttendanceRecordAuditLog so the correction is traceable even
# after it's superseded.

class AttendanceEditRequest(BaseModel):
    attendance_status: Optional[str] = None
    actual_checkin:    Optional[str] = None
    actual_checkout:   Optional[str] = None
    leave_code:        Optional[str] = None
    is_day_off:        Optional[bool] = None
    notes:             Optional[str] = None
    reason:            Optional[str] = None  # why this edit was made — logged, not stored on the row


@router.patch("/{employee_id}/{attendance_date}")
async def edit_attendance_day(
    employee_id:     str,
    attendance_date: str,
    body:            AttendanceEditRequest,
    db:              AsyncSession = Depends(get_db),
    user:            CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Upsert one employee's attendance/leave day. Diffs every changed field
    against a field-level audit log, mirroring update_employee() in
    hr_employees.py."""
    att_date = _to_date(attendance_date)
    if att_date is None:
        raise HTTPException(status_code=400, detail="attendance_date must be in YYYY-MM-DD format")

    updates = body.model_dump(exclude={"reason"}, exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    target = await db.scalar(
        select(AttendanceRecord)
        .where(AttendanceRecord.employee_id == employee_id)
        .where(AttendanceRecord.attendance_date == att_date)
    )

    if target is None:
        from app.models.employee import Employee
        emp = await db.scalar(select(Employee).where(Employee.user_id == employee_id))
        if not emp:
            raise HTTPException(status_code=404, detail=f"Employee {employee_id} not found in Employee master")
        target = AttendanceRecord(
            employee_id=employee_id, employee_name=emp.full_name, department=emp.department,
            attendance_date=att_date, week_day=att_date.strftime("%A"),
            source="manual", uploaded_at=datetime.utcnow(),
        )
        db.add(target)
        for field, value in updates.items():
            setattr(target, field, value)
            db.add(AttendanceRecordAuditLog(
                employee_id=employee_id, attendance_date=att_date, field=field,
                old_value=None, new_value=str(value) if value is not None else None,
                changed_by=user.username, reason=body.reason or None,
            ))
    else:
        for field, value in updates.items():
            old_val = getattr(target, field)
            if (old_val if old_val is not None else None) != value:
                db.add(AttendanceRecordAuditLog(
                    employee_id=employee_id, attendance_date=att_date, field=field,
                    old_value=str(old_val) if old_val is not None else None,
                    new_value=str(value) if value is not None else None,
                    changed_by=user.username, reason=body.reason or None,
                ))
                setattr(target, field, value)
        target.source = "manual"

    await db.flush()
    return {
        "employee_id":       employee_id,
        "attendance_date":   str(att_date),
        "attendance_status": target.attendance_status,
        "actual_checkin":    target.actual_checkin,
        "actual_checkout":   target.actual_checkout,
        "leave_code":        target.leave_code,
        "is_day_off":        target.is_day_off,
        "notes":             target.notes,
        "source":            target.source,
    }


@router.get("/{employee_id}/{attendance_date}/history")
async def get_attendance_day_history(
    employee_id:     str,
    attendance_date: str,
    db:              AsyncSession = Depends(get_db),
    user:            CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Field-level manual-edit history for one employee's day, most recent first."""
    att_date = _to_date(attendance_date)
    if att_date is None:
        raise HTTPException(status_code=400, detail="attendance_date must be in YYYY-MM-DD format")

    result = await db.execute(
        select(AttendanceRecordAuditLog)
        .where(AttendanceRecordAuditLog.employee_id == employee_id)
        .where(AttendanceRecordAuditLog.attendance_date == att_date)
        .order_by(AttendanceRecordAuditLog.changed_at.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "id":         r.id,
            "field":      r.field,
            "old_value":  r.old_value,
            "new_value":  r.new_value,
            "changed_by": r.changed_by,
            "changed_at": r.changed_at.isoformat() if r.changed_at else None,
            "reason":     r.reason,
        }
        for r in rows
    ]


# ── Kehadiran hari ini (per department) ───────────────────────────────────────

@router.get("/today")
async def get_today_attendance(
    target_date: Optional[str] = Query(None),
    db:          AsyncSession  = Depends(get_db),
    user:        CurrentUser   = Depends(require_role(Roles.HR)),
):
    """
    Kehadiran untuk tanggal tertentu (default: hari ini), dikelompokkan per department.
    Jika tidak ada data untuk hari ini, otomatis tampilkan tanggal terakhir yang ada.
    """
    today = date.today()

    if target_date:
        try:
            q_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        except ValueError:
            q_date = today
    else:
        q_date = today

    count_q = await db.execute(
        select(func.count()).select_from(AttendanceRecord)
        .where(AttendanceRecord.attendance_date == q_date)
    )
    count = count_q.scalar() or 0

    actual_date = q_date
    if count == 0:
        latest_q = await db.execute(select(func.max(AttendanceRecord.attendance_date)))
        latest = latest_q.scalar()
        if not latest:
            return {"requested_date": str(q_date), "actual_date": str(q_date),
                    "is_today": False, "has_data": False, "summary": {}, "data": []}
        actual_date = latest

    q = select(
        AttendanceRecord.department,
        func.sum(_plan_expr()).label("total"),
        func.sum(_actual_expr()).label("hadir"),
    ).where(AttendanceRecord.attendance_date == actual_date)
    result = await db.execute(q.group_by(AttendanceRecord.department).order_by(AttendanceRecord.department))
    rows = result.fetchall()

    data = [
        {
            "department": r[0] or "—",
            "total":  int(r[1] or 0),
            "hadir":  int(r[2] or 0),
            "absen":  int(r[1] or 0) - int(r[2] or 0),
            "rate":   round(int(r[2] or 0) / int(r[1]) * 100, 1) if r[1] else 0,
        }
        for r in rows if r[1]  # skip departments with zero required working days that day (e.g. everyone on leave)
    ]

    total_all   = sum(d["total"]  for d in data)
    total_hadir = sum(d["hadir"]  for d in data)
    total_absen = sum(d["absen"]  for d in data)

    return {
        "requested_date": str(q_date),
        "actual_date":    str(actual_date),
        "is_today":       actual_date == today,
        "has_data":       True,
        "summary": {
            "total":          total_all,
            "hadir":          total_hadir,
            "absen":          total_absen,
            "attendance_rate": round(total_hadir / total_all * 100, 1) if total_all > 0 else 0,
        },
        "data": data,
    }


# ── Monthly attendance rate ────────────────────────────────────────────────────

@router.get("/departments")
async def get_attendance_departments(
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.HR)),
):
    """All departments that exist — union of the Employee master list and
    whatever department names show up in uploaded attendance records, so the
    filter always covers every department even if attendance hasn't been
    uploaded for all of them yet. Cleaned via clean_department_list: drops
    numeric-junk values (e.g. '15' from a known Excel column-shift bug still
    pending manual HR correction on a couple of Employee rows) and collapses
    case-only duplicates ("Plant" / "PLANT") to one display label."""
    from app.models.employee import Employee

    emp_result = await db.execute(
        select(Employee.department).where(Employee.department.isnot(None)).distinct()
    )
    att_result = await db.execute(
        select(AttendanceRecord.department).where(AttendanceRecord.department.isnot(None)).distinct()
    )
    names = {r[0] for r in emp_result.all() if r[0]} | {r[0] for r in att_result.all() if r[0]}
    return clean_department_list(names)


@router.get("/coverage")
async def get_attendance_coverage(
    years: int = Query(10, ge=1, le=30),
    db:    AsyncSession = Depends(get_db),
    user:  CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Year x month grid of how much attendance data exists, so HR can spot
    months nobody ever uploaded. A month only counts as a "gap" if it falls
    inside the range that already has data somewhere — before the first
    ever upload isn't a gap, it's just outside the system's operating
    history."""
    from sqlalchemy import extract

    per_source_result = await db.execute(
        select(
            extract("year",  AttendanceRecord.attendance_date).label("year"),
            extract("month", AttendanceRecord.attendance_date).label("month"),
            AttendanceRecord.source,
            func.count().label("total"),
        )
        .group_by(
            extract("year",  AttendanceRecord.attendance_date),
            extract("month", AttendanceRecord.attendance_date),
            AttendanceRecord.source,
        )
    )
    per_month_result = await db.execute(
        select(
            extract("year",  AttendanceRecord.attendance_date).label("year"),
            extract("month", AttendanceRecord.attendance_date).label("month"),
            func.count(func.distinct(AttendanceRecord.employee_id)).label("employees"),
        )
        .group_by(
            extract("year",  AttendanceRecord.attendance_date),
            extract("month", AttendanceRecord.attendance_date),
        )
    )

    by_ym = {}
    for yr, mo, source, total in per_source_result.fetchall():
        key = (int(yr), int(mo))
        cell = by_ym.setdefault(key, {"total": 0, "employees": 0, "by_source": {}})
        cell["total"] += int(total)
        cell["by_source"][source or "unknown"] = int(total)
    for yr, mo, employees in per_month_result.fetchall():
        key = (int(yr), int(mo))
        by_ym.setdefault(key, {"total": 0, "employees": 0, "by_source": {}})["employees"] = int(employees)

    covered_ym = sorted(k for k, v in by_ym.items() if v["total"] > 0)
    observed_from = covered_ym[0] if covered_ym else None
    observed_to   = covered_ym[-1] if covered_ym else None

    today  = date.today()
    cur_ym = (today.year, today.month)
    year_list = list(range(today.year - years + 1, today.year + 1))
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    grid = []
    for yr in year_list:
        for mo in range(1, 13):
            ym = (yr, mo)
            cell = by_ym.get(ym, {"total": 0, "employees": 0, "by_source": {}})
            if cell["total"] > 0:
                status = "data"
            elif observed_from and observed_from <= ym <= observed_to and ym <= cur_ym:
                status = "gap"
            else:
                status = "outside"
            grid.append({
                "year": yr, "month": mo, "period": f"{MONTHS[mo-1]} {yr}",
                "total": cell["total"], "employees": cell["employees"],
                "by_source": cell["by_source"], "status": status,
            })

    log_result = await db.execute(
        select(AttendanceUploadLog).order_by(AttendanceUploadLog.uploaded_at.desc())
    )
    last_upload_by_source = {}
    for log in log_result.scalars().all():
        if log.source not in last_upload_by_source:
            last_upload_by_source[log.source] = {
                "uploaded_at": log.uploaded_at.isoformat() if log.uploaded_at else None,
                "filename":    log.filename,
                "batch_id":    log.batch_id,
            }

    return {
        "years": year_list,
        "grid":  grid,
        "last_upload_by_source": last_upload_by_source,
        "observed_range": {
            "from": f"{observed_from[0]}-{observed_from[1]:02d}" if observed_from else None,
            "to":   f"{observed_to[0]}-{observed_to[1]:02d}"     if observed_to   else None,
        },
    }


@router.get("/monthly-rate")
async def get_monthly_attendance_rate(
    department: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Attendance rate per bulan."""
    from sqlalchemy import extract

    q = select(
        extract("year",  AttendanceRecord.attendance_date).label("year"),
        extract("month", AttendanceRecord.attendance_date).label("month"),
        func.sum(_plan_expr()).label("working"),
        func.sum(_actual_expr()).label("hadir"),
    )
    if department:
        q = q.where(AttendanceRecord.department == department)
    if year:
        q = q.where(extract("year", AttendanceRecord.attendance_date) == year)

    result = await db.execute(
        q.group_by(
            extract("year",  AttendanceRecord.attendance_date),
            extract("month", AttendanceRecord.attendance_date),
        )
        .order_by(
            extract("year",  AttendanceRecord.attendance_date).desc(),
            extract("month", AttendanceRecord.attendance_date).desc(),
        )
        .limit(12)
    )
    rows = result.fetchall()

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data = []
    for r in rows:
        year    = int(r[0])
        month   = int(r[1])
        working = int(r[2] or 0)
        hadir   = int(r[3] or 0)
        absen   = working - hadir
        rate    = round(hadir / working * 100, 1) if working > 0 else 0
        data.append({
            "period":   f"{MONTHS[month-1]} {year}",
            "year":     year,
            "month":    month,
            "working":  working,
            "hadir":    hadir,
            "absen":    absen,
            "rate":     rate,
        })
    return data


# ── Target vs Achievement (Attendance Ratio) ──────────────────────────────────

@router.get("/target-vs-achievement")
async def get_target_vs_achievement(
    year: Optional[int] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """
    Target vs Achievement per bulan (Attendance Ratio).

    Target (Man-Days)  = Total Employees (aktif bulan itu) x Effective Working Days
    Achievement         = man-days hadir aktual (Intercom present OR Talenta
                           Business Trip; leave days excluded from Achievement,
                           same combination rule as every other report here)
    """
    from sqlalchemy import extract
    from app.models.employee import Employee
    from app.models.working_calendar import WorkingCalendarHoliday

    yr = year or date.today().year
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def _month_range(m: int):
        first = date(yr, m, 1)
        last  = date(yr, 12, 31) if m == 12 else date(yr, m + 1, 1) - timedelta(days=1)
        return first, last

    # 1) Effective Working Days per bulan — kalender kerja (exclude weekend + holiday)
    holidays_q = await db.execute(
        select(WorkingCalendarHoliday.holiday_date)
        .where(extract("year", WorkingCalendarHoliday.holiday_date) == yr)
    )
    holiday_dates = {r[0] for r in holidays_q.fetchall()}

    def _effective_working_days(m: int) -> int:
        first, last = _month_range(m)
        working = 0
        d = first
        while d <= last:
            if d.weekday() < 5 and d not in holiday_dates:
                working += 1
            d += timedelta(days=1)
        return working

    # 2) Total Employees aktif per bulan — sama seperti /employees/monthly-summary
    emps_q = await db.execute(
        select(Employee.date_of_joining, Employee.resign_date)
        .where(Employee.date_of_joining.isnot(None))
    )
    emps = emps_q.fetchall()

    def _active_headcount(m: int) -> int:
        first, last = _month_range(m)
        return sum(
            1 for join, resign in emps
            if join <= last and (resign is None or resign >= first)
        )

    # 3) Achievement — man-days hadir aktual per bulan (Intercom + Talenta combined)
    achievement_q = await db.execute(
        select(
            extract("month", AttendanceRecord.attendance_date).label("month"),
            func.sum(_actual_expr()).label("achievement"),
        )
        .where(extract("year", AttendanceRecord.attendance_date) == yr)
        .group_by(extract("month", AttendanceRecord.attendance_date))
    )
    achievement_map = {int(r[0]): int(r[1] or 0) for r in achievement_q.fetchall()}

    result = []
    for m in range(1, 13):
        headcount    = _active_headcount(m)
        working_days = _effective_working_days(m)
        target       = headcount * working_days
        achievement  = achievement_map.get(m, 0)
        rate         = round(achievement / target * 100, 1) if target > 0 else 0
        result.append({
            "period":       f"{MONTHS[m-1]} {yr}",
            "month":        m,
            "year":         yr,
            "headcount":    headcount,
            "working_days": working_days,
            "target":       target,
            "achievement":  achievement,
            "rate":         rate,
        })
    return {"year": yr, "months": result}


# ── Department summary (plan vs actual) ───────────────────────────────────────

@router.get("/dept-summary")
async def get_dept_summary(
    department: Optional[str] = Query(None),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Attendance Plan vs Actual per department.

    Always includes every department from the Employee master list (as
    0%) when no department filter is applied, not just departments that
    already have attendance records — otherwise a department with no
    attendance uploaded yet silently disappears from the chart even
    though it's a real department (visible in the filter dropdown)."""
    from sqlalchemy import extract
    from app.models.employee import Employee

    q = select(
        AttendanceRecord.department,
        func.sum(_plan_expr()).label("plan"),
        func.sum(_actual_expr()).label("actual"),
    )
    if department:
        q = q.where(AttendanceRecord.department == department)
    if month:
        q = q.where(extract("month", AttendanceRecord.attendance_date) == month)
    if year:
        q = q.where(extract("year", AttendanceRecord.attendance_date) == year)

    result = await db.execute(q.group_by(AttendanceRecord.department))
    att_map = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in result.fetchall() if r[0]}

    if department:
        dept_names = {department}
    else:
        emp_result = await db.execute(
            select(Employee.department).where(Employee.department.isnot(None)).distinct()
        )
        dept_names = {r[0] for r in emp_result.fetchall() if r[0]} | set(att_map.keys())

    rows = []
    for dept in sorted(dept_names):
        plan, actual = att_map.get(dept, (0, 0))
        rows.append({
            "department": dept,
            "plan":   plan,
            "actual": actual,
            "rate":   round(actual / max(plan, 1) * 100) if plan else 0,
        })
    return rows


# ── Daftar karyawan untuk filter hadir/absen/semua pada tanggal tertentu ───────

@router.get("/today/employees")
async def get_today_employees(
    target_date: Optional[str] = Query(None),
    filter:      str           = Query("all"),   # all | hadir | absen
    db:          AsyncSession  = Depends(get_db),
    user:        CurrentUser   = Depends(require_role(Roles.HR)),
):
    """Daftar nama karyawan (hadir/absen/semua) untuk tanggal yang ditampilkan di /today."""
    today  = date.today()
    q_date = today

    if target_date:
        try:
            q_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    count_q = await db.execute(
        select(func.count()).select_from(AttendanceRecord)
        .where(AttendanceRecord.attendance_date == q_date)
    )
    if (count_q.scalar() or 0) == 0:
        latest_q = await db.execute(select(func.max(AttendanceRecord.attendance_date)))
        latest   = latest_q.scalar()
        if latest:
            q_date = latest

    q = select(AttendanceRecord).where(AttendanceRecord.attendance_date == q_date).where(IS_WEEKDAY)

    if filter == "hadir":
        q = q.where(or_(_is_bt_code(), _is_present_intercom())).where(~_is_leave_code())
    elif filter == "absen":
        q = q.where(~_is_leave_code()).where(~_is_bt_code()).where(~_is_present_intercom())

    q = q.order_by(AttendanceRecord.department, AttendanceRecord.employee_name)
    result = await db.execute(q)
    rows = result.scalars().all()

    def _reason(rec):
        if rec.leave_code:
            return LEAVE_LABELS.get(rec.leave_code, rec.leave_code)
        return rec.notes

    return {
        "date":      str(q_date),
        "filter":    filter,
        "employees": [
            {
                "id":         r.employee_id,
                "name":       r.employee_name,
                "department": r.department,
                "checkin":    r.actual_checkin,
                "checkout":   r.actual_checkout,
                "notes":      _reason(r),
            }
            for r in rows
        ],
    }


# ── Who's off (absent/leave/business trip on latest date) ─────────────────────

@router.get("/whos-off")
async def get_whos_off(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Karyawan yang tidak hadir fisik pada tanggal terakhir yang tersedia —
    termasuk cuti dan business trip (dengan alasannya), bukan hanya absen."""
    latest_q = await db.execute(select(func.max(AttendanceRecord.attendance_date)))
    latest   = latest_q.scalar()
    if not latest:
        return {"date": None, "data": []}

    result = await db.execute(
        select(
            AttendanceRecord.employee_name,
            AttendanceRecord.department,
            AttendanceRecord.leave_code,
        )
        .where(AttendanceRecord.attendance_date == latest)
        .where(IS_WEEKDAY)
        .where(or_(~_is_present_intercom(), _is_leave_code(), _is_bt_code()))
        .order_by(AttendanceRecord.employee_name)
        .limit(15)
    )
    rows = result.fetchall()
    data = []
    for name, dept, code in rows:
        data.append({
            "name": name or "—",
            "department": dept or "—",
            "reason": LEAVE_LABELS.get(code, code) if code else "Absent",
        })
    return {"date": str(latest), "data": data}


# ── Workforce stats (gender + work location) ───────────────────────────────────

@router.get("/workforce-stats")
async def get_workforce_stats(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Attendance rate berdasarkan gender dan work_placement (join dengan tabel employees)."""
    from app.models.employee import Employee

    def _rate(plan, actual):
        p, a = int(plan or 0), int(actual or 0)
        return round(a / p * 100) if p > 0 else 0

    gender_q = await db.execute(
        select(
            Employee.sex,
            func.sum(_plan_expr()).label("plan"),
            func.sum(_actual_expr()).label("actual"),
        )
        .join(Employee, AttendanceRecord.employee_id == Employee.user_id)
        .where(Employee.sex.isnot(None))
        .group_by(Employee.sex)
    )
    gender_rows = gender_q.fetchall()

    loc_q = await db.execute(
        select(
            Employee.work_placement,
            func.sum(_plan_expr()).label("plan"),
            func.sum(_actual_expr()).label("actual"),
        )
        .join(Employee, AttendanceRecord.employee_id == Employee.user_id)
        .where(Employee.work_placement.isnot(None))
        .group_by(Employee.work_placement)
        .order_by(Employee.work_placement)
    )
    loc_rows = loc_q.fetchall()

    GENDER = {"M": "Male", "F": "Female"}
    return {
        "by_gender": [
            {"label": GENDER.get(r[0], r[0]), "plan": int(r[1] or 0),
             "actual": int(r[2] or 0), "rate": _rate(r[1], r[2])}
            for r in gender_rows if r[0]
        ],
        "by_location": [
            {"label": r[0], "plan": int(r[1] or 0),
             "actual": int(r[2] or 0), "rate": _rate(r[1], r[2])}
            for r in loc_rows if r[0]
        ],
    }


# ── Department + Team summary (rekap per tim) ─────────────────────────────────

@router.get("/dept-team-summary")
async def get_dept_team_summary(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Rekap Plan vs Actual per department dan team (join dengan tabel employees)."""
    from app.models.employee import Employee
    from collections import defaultdict

    result = await db.execute(
        select(
            AttendanceRecord.department,
            Employee.team,
            func.count(func.distinct(AttendanceRecord.employee_id)).label("employees"),
            func.sum(_plan_expr()).label("plan"),
            func.sum(_actual_expr()).label("actual"),
        )
        .join(Employee, AttendanceRecord.employee_id == Employee.user_id, isouter=True)
        .group_by(AttendanceRecord.department, Employee.team)
        .order_by(AttendanceRecord.department, Employee.team)
    )
    rows = result.fetchall()

    depts: dict = defaultdict(list)
    for r in rows:
        dept   = r[0] or "—"
        team   = r[1] or "—"
        plan   = int(r[3] or 0)
        actual = int(r[4] or 0)
        if plan == 0 and actual == 0:
            continue
        depts[dept].append({
            "team":      team,
            "employees": int(r[2] or 0),
            "plan":      plan,
            "actual":    actual,
            "rate":      round(actual / plan * 100) if plan > 0 else 0,
        })

    result_list = []
    grand = {"employees": 0, "plan": 0, "actual": 0}

    for dept, teams in depts.items():
        tot = {
            "employees": sum(t["employees"] for t in teams),
            "plan":      sum(t["plan"]      for t in teams),
            "actual":    sum(t["actual"]    for t in teams),
        }
        tot["rate"] = round(tot["actual"] / tot["plan"] * 100) if tot["plan"] > 0 else 0
        result_list.append({"department": dept, "teams": teams, "total": tot})
        grand["employees"] += tot["employees"]
        grand["plan"]      += tot["plan"]
        grand["actual"]    += tot["actual"]

    grand["rate"] = round(grand["actual"] / grand["plan"] * 100) if grand["plan"] > 0 else 0
    return {"departments": result_list, "grand_total": grand}


# ── Employee search ────────────────────────────────────────────────────────────

@router.get("/search-employees")
async def search_employees(
    q:    str          = Query(..., min_length=2),
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(require_role(Roles.HR)),
):
    result = await db.execute(
        select(
            AttendanceRecord.employee_id,
            AttendanceRecord.employee_name,
            AttendanceRecord.department,
        )
        .where(AttendanceRecord.employee_name.ilike(f"%{q}%"))
        .where(AttendanceRecord.employee_id.isnot(None))
        .distinct(AttendanceRecord.employee_id)
        .limit(10)
    )
    rows = result.fetchall()
    return [{"id": r[0], "name": r[1], "department": r[2]} for r in rows if r[0]]


# ── Individual employee detail ─────────────────────────────────────────────────

@router.get("/employee/{employee_id}/detail")
async def get_employee_detail(
    employee_id: str,
    year:        Optional[int] = Query(None, description="Batasi chart bulanan + absence ke 1 tahun ini saja; kosong = semua tahun"),
    db:          AsyncSession = Depends(get_db),
    user:        CurrentUser  = Depends(require_role(Roles.HR)),
):
    """Monthly attendance + absence records untuk satu karyawan."""
    from app.models.employee import Employee
    from sqlalchemy import extract

    emp_q = await db.execute(select(Employee).where(Employee.user_id == employee_id))
    emp   = emp_q.scalar_one_or_none()

    monthly_q = await db.execute(
        select(
            extract("year",  AttendanceRecord.attendance_date).label("year"),
            extract("month", AttendanceRecord.attendance_date).label("month"),
            func.sum(_plan_expr()).label("plan"),
            func.sum(_actual_expr()).label("actual"),
        )
        .where(AttendanceRecord.employee_id == employee_id)
        .where(True if year is None else extract("year", AttendanceRecord.attendance_date) == year)
        .group_by(
            extract("year",  AttendanceRecord.attendance_date),
            extract("month", AttendanceRecord.attendance_date),
        )
        .order_by(
            extract("year",  AttendanceRecord.attendance_date).asc(),
            extract("month", AttendanceRecord.attendance_date).asc(),
        )
    )
    monthly_rows = monthly_q.fetchall()

    # Only genuinely unexplained absences — not on leave, not on business
    # trip, not present per Intercom.
    absence_q = await db.execute(
        select(AttendanceRecord)
        .where(AttendanceRecord.employee_id == employee_id)
        .where(IS_WEEKDAY)
        .where(~_is_leave_code())
        .where(~_is_bt_code())
        .where(~_is_present_intercom())
        .where(True if year is None else extract("year", AttendanceRecord.attendance_date) == year)
        .order_by(AttendanceRecord.attendance_date.desc())
        .limit(30)
    )
    absences = absence_q.scalars().all()

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    first_absence = absences[0] if absences else None
    return {
        "employee": {
            "id":             employee_id,
            "name":           emp.full_name        if emp else (first_absence.employee_name if first_absence else None),
            "department":     emp.department       if emp else (first_absence.department    if first_absence else None),
            "team":           emp.team             if emp else None,
            "work_placement": emp.work_placement   if emp else None,
            "sex":            emp.sex              if emp else None,
        },
        "monthly": [
            {
                "period": f"{MONTHS[int(r[1])-1]} {int(r[0])}",
                "plan":   int(r[2] or 0),
                "actual": int(r[3] or 0),
                "rate":   round(int(r[3] or 0) / max(int(r[2] or 1), 1) * 100),
            }
            for r in monthly_rows
        ],
        "absences": [
            {"date": str(a.attendance_date), "reason": a.notes or "Absent"}
            for a in absences
        ],
    }
