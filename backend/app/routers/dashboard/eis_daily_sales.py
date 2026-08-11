import json
import io
import re
import datetime
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.dependencies import get_current_user
from app.database import get_db
from app.models.sales_plan import SalesPlan

router = APIRouter()

DATA_FILE = Path(__file__).parent.parent.parent / "data" / "daily_sales.json"
DEFAULT_FILE = Path(__file__).parent.parent.parent / "data" / "daily_sales_default.json"

MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december"
]


def _load_store() -> dict:
    """The on-disk file is a dict keyed by fiscal year (as a string), e.g.
    {"2025": {...}, "2026": {...}} -- each value has the same shape the old
    single-blob format used (year/month/rows/...). Keyed storage is what
    lets the Year selector actually change what's displayed; previously
    the whole file WAS one year's data, so switching the dropdown had no
    effect and a new upload silently overwrote whatever year was there
    before, regardless of which year the data was actually for."""
    for path in (DATA_FILE, DEFAULT_FILE):
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                return data if isinstance(data, dict) else {}
            except Exception:
                continue
    return {}


async def _pac_monthly_business_plan(db: AsyncSession, year: int) -> tuple[list[float], bool]:
    """Company-wide Sales Business Plan per month (Jan-Dec, in Million IDR)
    from the PAC dashboard's Sales Plan module — the same figures PAC uses
    for its own Gross Sales Report, now the single source of truth for the
    Daily Sales BP card too (previously a manually-typed cell in the Daily
    Sales Excel's "Daily Sales Performance" sheet, which the currently-used
    upload templates don't even have, leaving it permanently blank).

    Each PAC team stores its plan as one pre-aggregated "Total" document
    (content.meta.area == "Total", meta.type blank) plus one document per
    market/customer segment that sums up to that Total — group by
    (department, team_code) and prefer each team's own Total doc so a team
    is counted exactly once; fall back to summing its segment docs only if
    it never submitted a Total rollup, so that team isn't dropped entirely."""
    result = await db.execute(
        select(SalesPlan).where(SalesPlan.plan_year == year, SalesPlan.plan_type == "value")
    )
    groups: dict = {}
    for plan in result.scalars().all():
        content = plan.content or {}
        meta = content.get("meta", {})
        is_total = meta.get("area") == "Total" and not meta.get("type")
        key = (plan.department, plan.team_code)
        bucket = groups.setdefault(key, {"total": None, "segments": []})
        rows = content.get("rows", [])
        if is_total:
            bucket["total"] = rows
        else:
            bucket["segments"].append(rows)

    monthly = [0.0] * 12
    has_data = False
    for bucket in groups.values():
        rows = bucket["total"] if bucket["total"] is not None else [r for seg in bucket["segments"] for r in seg]
        for row in rows:
            # Row shape: [no, country, product, jan..dec, total_value, total_unit, price] (18 items) —
            # months start at index 3, not 2, since a "Country" column was added before "Product".
            if len(row) < 15:
                continue
            has_data = True
            for i in range(12):
                v = row[3 + i]
                if isinstance(v, (int, float)):
                    monthly[i] += v
    return [round(v / 1_000_000, 3) for v in monthly], has_data


def _month_progress(rows: list, month: str) -> tuple[float, int]:
    """(last_acc, wd_actual_count) for one month out of the Daily Sales
    Chart sheet's WD rows — how much has been sold so far, and how many
    working days that covers."""
    last_acc, wd_actual = 0.0, 0
    for row in rows:
        cell = row.get(month) or {}
        if cell.get("acc") is not None:
            wd_actual += 1
            last_acc = cell["acc"]
    return last_acc, wd_actual


def _expectation_closing(last_acc: float, wd_actual: int, wd_total: int) -> float:
    """Run-rate projection: (Acc so far / WD elapsed) x WD in the whole
    month. Floored at last_acc so a month can never appear to be closing
    *below* what's already been sold."""
    if wd_actual == 0:
        return 0.0
    if wd_total == 0:
        return round(last_acc, 3)
    return round(max(last_acc, last_acc / wd_actual * wd_total), 3)


def _daily_sales_kpi_by_month(rows: list) -> dict:
    """Actual-to-date and run-rate-projected closing for every month, from
    the Daily Sales WD rows alone, each tagged with a status so the
    frontend can color real vs. estimated figures differently:
      - "actual": a closed month (not the most recent one with data) — its
        own WD count already IS its total, no projection needed.
      - "projected": the single "in progress" month (the most recent one
        with any data) — run-rate projected using the average WD count of
        this year's already-closed months as its estimated total. Can't use
        the sheet's own Target column for that estimate — it turns out to
        be filled in lockstep with Acc/Sales (row by row, as each day is
        entered), not pre-filled for the whole month, so it always equals
        "days reported so far". An HR Working Calendar (Mon-Fri minus
        holidays) was tried instead and rejected too: every closed month
        runs 3-4 WD *higher* than that calendar (e.g. Dec 2025: 25 WD rows
        vs. 23 Mon-Fri days that month) — evidently Daily Sales' own WD
        scheme isn't a plain weekday count (most likely some Saturdays
        count as working days), so plain office attendance doesn't apply.
      - "carried_forward": a month with no data yet, chronologically after
        the in-progress month (e.g. picking "August" when data only goes
        through July) — shows the running year-to-date total through the
        last real month instead of a bare, alarming-looking 0.
      - "no_data": no month this year has any data at all yet.
    wd_total is also returned per month (0 for carried_forward/no_data) so
    callers that need to project individual WD rows — not just the
    month-end total — can reuse the same "how many WD will this month have"
    estimate (see _project_current_month_rows)."""
    progress = {m: _month_progress(rows, m) for m in MONTHS}
    months_with_data = [m for m in MONTHS if progress[m][1] > 0]
    current_month = months_with_data[-1] if months_with_data else None
    closed_months = months_with_data[:-1]
    avg_wd_closed = (
        round(sum(progress[m][1] for m in closed_months) / len(closed_months))
        if closed_months else None
    )

    result = {}
    ytd_running = 0.0
    for m in MONTHS:
        last_acc, wd_actual = progress[m]
        if wd_actual > 0:
            if m == current_month and avg_wd_closed is not None:
                wd_total, status = avg_wd_closed, "projected"
            else:
                wd_total, status = wd_actual, "actual"
            exp = _expectation_closing(last_acc, wd_actual, wd_total)
            ytd_running += exp
            result[m] = {"last_acc": last_acc, "wd_actual": wd_actual, "wd_total": wd_total, "expectation_closing": exp, "status": status}
        elif current_month is not None:
            result[m] = {"last_acc": 0.0, "wd_actual": 0, "wd_total": 0, "expectation_closing": round(ytd_running, 3), "status": "carried_forward"}
        else:
            result[m] = {"last_acc": 0.0, "wd_actual": 0, "wd_total": 0, "expectation_closing": 0.0, "status": "no_data"}
    return result


def _project_current_month_rows(rows: list, kpi_by_month: dict) -> None:
    """Mutates `rows` in place: for the single "projected" (in-progress)
    month, fills its Acc/Sales past the last real WD with a straight-line
    projection toward that month's run-rate expectation_closing, tagged
    `projected: True` so the frontend can render them in a visibly
    different color from real data — the Daily Sales Detail table
    otherwise just leaves those cells blank until the real upload
    catches up, which reads as missing/broken rather than "not yet due"."""
    current_month = next((m for m, info in kpi_by_month.items() if info["status"] == "projected"), None)
    if current_month is None:
        return
    info = kpi_by_month[current_month]
    wd_actual, wd_total = info["wd_actual"], info["wd_total"]
    if wd_total <= wd_actual:
        return
    remaining = wd_total - wd_actual
    increment = (info["expectation_closing"] - info["last_acc"]) / remaining
    for row in rows:
        wd = row.get("wd")
        if wd is None or not (wd_actual < wd <= wd_total):
            continue
        cell = row.get(current_month) or {}
        if cell.get("acc") is not None:
            continue
        row[current_month] = {
            "target": cell.get("target"),
            "acc": round(info["last_acc"] + increment * (wd - wd_actual), 3),
            "sales": round(increment, 3),
            "projected": True,
        }


def _save_store(store: dict):
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    DATA_FILE.write_text(json.dumps(store, ensure_ascii=False), encoding="utf-8")


def _parse_excel(content: bytes, filename: str = "") -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    chart_idx = next((i for i, s in enumerate(sheet_names_lower) if "chart" in s), None)
    perf_idx = next((i for i, s in enumerate(sheet_names_lower) if "daily sales" in s or "performance" in s), None)

    if chart_idx is None:
        raise HTTPException(status_code=422, detail="Sheet 'Chart' tidak ditemukan di file Excel")

    ws = wb[wb.sheetnames[chart_idx]]
    rows = list(ws.iter_rows(min_row=1, max_row=35, values_only=True))

    if len(rows) < 2:
        raise HTTPException(status_code=422, detail="Sheet 'Chart' tidak memiliki baris header yang valid")

    # Kolom pertama boleh berisi "YEAR" (opsional) diikuti WD, Target, Acc, Sales
    # per bulan. Tanpa kolom YEAR, urutannya WD, Target, Acc, Sales per bulan.
    header_row = rows[1]
    has_year_col = isinstance(header_row[0], str) and header_row[0].strip().lower() == "year"
    wd_col = 1 if has_year_col else 0
    month_starts = [(2 if has_year_col else 1) + i * 3 for i in range(12)]

    table_rows = []
    detected_year = None
    for row in rows[2:]:
        if row[wd_col] is None or not isinstance(row[wd_col], (int, float)):
            continue
        if has_year_col and detected_year is None and isinstance(row[0], (int, float)):
            detected_year = int(row[0])
        wd = int(row[wd_col])
        entry = {"wd": wd}
        for i, month in enumerate(MONTHS):
            ms = month_starts[i]
            entry[month] = {
                "target": round(float(row[ms]), 3) if row[ms] is not None else None,
                "acc": round(float(row[ms + 1]), 3) if row[ms + 1] is not None else None,
                "sales": round(float(row[ms + 2]), 3) if row[ms + 2] is not None else None,
            }
        table_rows.append(entry)

    if not table_rows:
        raise HTTPException(
            status_code=422,
            detail="Tidak ada baris data valid di sheet 'Chart'. Pastikan kolom WD berisi angka.",
        )

    month_targets = {}
    last_month_with_data = None
    for i, month in enumerate(MONTHS):
        ms = month_starts[i]
        # Target is a single BP figure repeated down every WD row for the
        # month, not a per-day curve. Taking the *first* non-blank cell used
        # to mean a stray 0/placeholder in an early row (before the plan was
        # finalized) silently became "the" target, which zeroed out the
        # dashed reference line for months that otherwise had real target
        # data further down the column. Taking the max survives that.
        values = [round(float(row[ms]), 3) for row in rows[2:] if row[ms] is not None]
        if values:
            month_targets[month] = max(values)
        if any(r[month]["acc"] is not None for r in table_rows):
            last_month_with_data = month

    bp = 0.0
    exp_closing = 0.0
    ach_pct = 0.0
    as_of = ""
    year_from_perf = None
    month_label = (last_month_with_data or "december").capitalize()

    if perf_idx is not None:
        ws2 = wb[wb.sheetnames[perf_idx]]
        perf_rows = list(ws2.iter_rows(min_row=1, max_row=15, values_only=True))

        # Locate the "Business Plan" / "Expectation Closing" / "Achievement"
        # labels (spelling varies across templates, e.g. "Bussiness Plan",
        # "Achievment") and read the number that sits a couple of rows below
        # in the same column. This replaces guessing which of the sheet's
        # numeric cells is which by magnitude (previously: BP had to fall
        # between 1000-50000 and achievement between 0-200%), which broke
        # silently — leaving BP/Expectation Closing/Achievement all at 0 —
        # whenever a real figure fell outside those hardcoded ranges.
        label_aliases = {
            "business_plan": ("business plan", "bussiness plan"),
            "expectation_closing": ("expectation closing",),
            "achievement_pct": ("achievement", "achievment"),
        }
        label_positions = {}
        for r_idx, row in enumerate(perf_rows):
            for col_idx, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                low = cell.strip().lower()
                for key, aliases in label_aliases.items():
                    if key not in label_positions and low in aliases:
                        label_positions[key] = (r_idx, col_idx)

        if len(label_positions) == len(label_aliases):
            for key, (r_idx, col_idx) in label_positions.items():
                val = None
                for row in perf_rows[r_idx + 1: r_idx + 6]:
                    if col_idx < len(row) and isinstance(row[col_idx], (int, float)):
                        val = row[col_idx]
                        break
                if val is None:
                    continue
                if key == "business_plan":
                    bp = round(val, 3)
                elif key == "expectation_closing":
                    exp_closing = round(val, 3)
                elif key == "achievement_pct":
                    ach_pct = round(val * 100, 2)
        else:
            # Fallback for templates where the labels above aren't found.
            for row in perf_rows:
                nums = [v for v in row if isinstance(v, (int, float))]
                if len(nums) >= 3:
                    bp_c, exp_c, ach_c = nums[0], nums[1], nums[2]
                    if 1000 < bp_c < 50000 and 0 < ach_c < 2:
                        bp = round(bp_c, 3)
                        exp_closing = round(exp_c, 3)
                        ach_pct = round(ach_c * 100, 2)
                        break

        for row in perf_rows:
            dates = [v for v in row if isinstance(v, datetime.datetime)]
            if dates:
                as_of = dates[0].strftime("%Y-%m-%d")
                year_from_perf = dates[0].year
                break

    # Prioritas sumber tahun: kolom YEAR di sheet Chart > tanggal di sheet
    # performance/summary > angka tahun pada nama file > tahun berjalan.
    year = detected_year or year_from_perf
    if year is None:
        m = re.search(r"(20\d{2})", filename or "")
        if m:
            year = int(m.group(1))
    if year is None:
        year = datetime.date.today().year

    return {
        "detected_year": year,
        "month": month_label,
        "as_of": as_of,
        "business_plan": bp,
        "expectation_closing": exp_closing,
        "achievement_pct": ach_pct,
        "month_targets": month_targets,
        "rows": table_rows,
    }


@router.get("/years")
async def get_daily_sales_years(user = Depends(get_current_user)):
    """Years that actually have uploaded Daily Sales data, newest first —
    lets the Year filter only offer years worth picking instead of a fixed
    rolling 5-year window that includes years nobody has uploaded yet."""
    store = _load_store()
    return sorted((int(y) for y in store.keys()), reverse=True)


@router.get("/kpi")
async def get_daily_sales_kpi(
    year: int = Query(..., description="Fiscal year"),
    month: Optional[str] = Query(None, description="Lowercase month key (e.g. 'december') for a Monthly view; omit for Yearly"),
    db: AsyncSession = Depends(get_db),
    user = Depends(get_current_user),
):
    """Business Plan / Expectation Closing / Achievement for one card,
    Yearly (month omitted) or Monthly (month given). Business Plan comes
    from PAC's Sales Plan; Expectation Closing and Achievement are derived
    from the Daily Sales WD data already uploaded — see
    _pac_monthly_business_plan / _expectation_closing docstrings for the
    exact formulas."""
    if month is not None and month not in MONTHS:
        raise HTTPException(status_code=422, detail=f"Invalid month '{month}'. Expected one of {MONTHS}.")

    monthly_bp, has_pac_data = await _pac_monthly_business_plan(db, year)

    store = _load_store()
    entry = store.get(str(year))
    rows = entry.get("rows", []) if entry else []
    kpi_by_month = _daily_sales_kpi_by_month(rows)

    months_to_use = [month] if month else MONTHS
    total_bp = sum(monthly_bp[i] for i, m in enumerate(MONTHS) if m in months_to_use)
    total_actual = sum(kpi_by_month[m]["last_acc"] for m in months_to_use)

    if month:
        # Monthly: show this one month's own figure, carry-forward fallback included.
        total_expectation = kpi_by_month[month]["expectation_closing"]
        expectation_status = kpi_by_month[month]["status"]
    else:
        # Yearly: sum only real contributions (actual/projected) — a
        # carried_forward month repeats the same YTD figure as a display
        # fallback for its own card, so summing it into the year total
        # would double-count months already counted individually.
        counted = [kpi_by_month[m] for m in months_to_use if kpi_by_month[m]["status"] in ("actual", "projected")]
        total_expectation = sum(c["expectation_closing"] for c in counted)
        expectation_status = "projected" if any(c["status"] == "projected" for c in counted) else ("actual" if counted else "no_data")

    achievement_pct = round(total_actual / total_bp * 100, 2) if total_bp > 0 else 0.0

    return {
        "year": year,
        "month": month,
        "business_plan": round(total_bp, 3),
        "expectation_closing": round(total_expectation, 3),
        "expectation_status": expectation_status,
        "achievement_pct": achievement_pct,
        "has_pac_data": has_pac_data,
        "has_daily_sales_data": entry is not None,
    }


@router.get("/data")
async def get_daily_sales(
    year: int = Query(..., description="Fiscal year to display"),
    user = Depends(get_current_user),
):
    store = _load_store()
    entry = store.get(str(year))
    if entry is None:
        return {"data": None}
    rows = entry.get("rows", [])
    _project_current_month_rows(rows, _daily_sales_kpi_by_month(rows))
    return {"data": {**entry, "year": year}}


@router.post("/upload")
async def upload_daily_sales(
    year: int = Query(..., description="Fiscal year this file's data belongs to — required so it's stored under the right year instead of overwriting whatever was there before"),
    file: UploadFile = File(...),
    user = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="File harus berformat .xlsx atau .xls")
    content = await file.read()
    parsed = _parse_excel(content, file.filename)
    detected_year = parsed.pop("detected_year", None)

    store = _load_store()
    store[str(year)] = parsed
    _save_store(store)

    result = {**parsed, "year": year}
    message = f"Data berhasil diupload untuk tahun {year}"
    if detected_year and detected_year != year:
        message += (
            f" (perhatian: tahun yang terbaca dari file adalah {detected_year}, "
            f"berbeda dari tahun {year} yang dipilih — periksa kembali file atau pilihan tahunnya)"
        )
    return {"message": message, "detected_year": detected_year, "data": result}


@router.delete("/data")
async def delete_daily_sales(
    year: int = Query(..., description="Fiscal year to remove"),
    user = Depends(get_current_user),
):
    """Remove one year's uploaded Daily Sales data — lets a bad upload be
    cleared without having to overwrite it with another file first."""
    store = _load_store()
    if str(year) not in store:
        raise HTTPException(status_code=404, detail=f"Tidak ada data Daily Sales untuk tahun {year}")
    del store[str(year)]
    _save_store(store)
    return {"message": f"Data Daily Sales tahun {year} berhasil dihapus"}
