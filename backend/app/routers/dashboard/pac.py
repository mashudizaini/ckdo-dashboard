"""
PAC Dashboard Router
Route prefix : /api/v1/dashboard/pac
Required role: pac_staff OR admin

Endpoints:
  GET  /summary
  GET  /budget-usage              — Actual vs Budget per period/cost-center
  GET  /lov/ledgers               — GL ledger LOV
  GET  /business-plans            — List business plan documents
  GET  /business-plans/{id}       — Get single document
  POST /business-plans            — Create / update document (upsert)
  DELETE /business-plans/{id}     — Delete document
"""
import asyncio
import io
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
from pydantic import BaseModel
from app.dependencies import require_role, CurrentUser, Roles
from app.database import get_db
from sqlalchemy.ext.asyncio import AsyncSession
from app.services.pac_service import PACService
from app.services.business_plan_service import BusinessPlanService
from app.services.business_plan_setup_service import BusinessPlanSetupService
from app.services.outlook_material_service import OutlookMaterialService
from app.services.sales_plan_service import SalesPlanService
from app.services.purchase_plan_service import PurchasePlanService
from app.services.personnel_plan_service import PersonnelPlanService
from app.services.manufacture_plan_service import ManufacturePlanService
from app.services.investment_plan_service import InvestmentPlanService
from app.services.opex_plan_service import OpexPlanService
from app.services.ai_service import AIService
from app.services import exchange_rate_service
from app.services import user_api_key_service
from app.config import get_settings

settings = get_settings()

router = APIRouter()


@router.get("/summary")
async def get_summary(user: CurrentUser = Depends(require_role(Roles.PAC))):
    return {"module": "pac", "status": "ready"}


@router.get("/budget-usage")
async def get_budget_usage(
    year:         Optional[int] = Query(None),
    month:        Optional[int] = Query(None),
    cost_center:  Optional[str] = Query(None),
    account_type: Optional[str] = Query(None),
    ledger_id:    Optional[int] = Query(None),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Actual vs Business Plan (Budget) per cost center per GL period."""
    filters = dict(year=year, month=month, cost_center=cost_center,
                   account_type=account_type, ledger_id=ledger_id)
    return await PACService().get_budget_usage(filters)


@router.get("/lov/ledgers")
async def get_ledgers(user: CurrentUser = Depends(require_role(Roles.PAC))):
    return await PACService().get_ledgers()


# ── Business Plan ─────────────────────────────────────────────────────────────

class BusinessPlanPayload(BaseModel):
    id:          Optional[int]  = None
    doc_type:    str            = "strategy_plan"   # managerial_obj | strategy_plan
    plan_year:   int
    department:  Optional[str]  = "ALL"
    team_code:   Optional[str]  = ""
    team_name:   Optional[str]  = ""
    plan_role:   Optional[str]  = ""
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/business-plans")
async def list_business_plans(
    plan_year:  Optional[int] = Query(None),
    doc_type:   Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    return await BusinessPlanService().list_plans(db, plan_year, doc_type, department)


@router.get("/business-plans/{plan_id}")
async def get_business_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    return await BusinessPlanService().get_plan(db, plan_id)


@router.post("/business-plans")
async def upsert_business_plan(
    body: BusinessPlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    return await BusinessPlanService().upsert_plan(db, body.model_dump(), user.username)


@router.post("/business-plans/upload")
async def upload_business_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import a Strategy & Action Plan from an Excel file matching the
    "Strategy_Action Plan - Mashudi.xlsx" template."""
    content = await file.read()
    return await BusinessPlanService().import_excel(db, content, plan_year, user.username)


@router.delete("/business-plans/{plan_id}")
async def delete_business_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    return await BusinessPlanService().delete_plan(db, plan_id)


# ── Business Plan Setup (Schedule / Guideline / Outlook) ───────────────────────

class SetupPayload(BaseModel):
    id:           Optional[int]  = None
    setup_module: str            = "schedule"   # schedule | guideline | outlook
    plan_year:    int
    content:      dict           = {}
    status:       Optional[str]  = "draft"


@router.get("/setup-modules")
async def list_setup_modules(
    setup_module: Optional[str] = Query(None),
    plan_year:    Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List setup modules (schedule, guideline, outlook) filtered by year/module."""
    return await BusinessPlanSetupService().list_setup(db, setup_module, plan_year)


@router.get("/setup-modules/{setup_id}")
async def get_setup_module(
    setup_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single setup module document."""
    return await BusinessPlanSetupService().get_setup(db, setup_id)


@router.post("/setup-modules")
async def upsert_setup_module(
    body: SetupPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update setup module document."""
    return await BusinessPlanSetupService().upsert_setup(db, body.model_dump(), user.username)


@router.delete("/setup-modules/{setup_id}")
async def delete_setup_module(
    setup_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete setup module document."""
    return await BusinessPlanSetupService().delete_setup(db, setup_id)


@router.get("/setup-modules/schedule/export")
async def export_schedule_excel(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Business Plan Schedule to Excel, matching the layout of
    Business plan schedule.xlsx (extended with the Submission Date
    From/To split and the auto-computed Actual Date From/To columns)."""
    return await BusinessPlanSetupService().export_schedule_excel(db, plan_year)


@router.get("/setup-modules/guideline/export")
async def export_guideline_ppt(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Business Plan Guideline to PowerPoint, matching the
    layout of Business plan guideline.xlsx: a title slide plus one table
    slide per section, each column-split into Current Year / Previous
    Year."""
    return await BusinessPlanSetupService().export_guideline_ppt(db, plan_year)


@router.post("/setup-modules/outlook/materials")
async def upload_outlook_materials(
    plan_year: int = Query(...),
    category: str = Query("material", pattern="^(material|format)$"),
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Upload one or more reference files for plan_year. category="material"
    (default) = source data (economic reports, market data, etc.) that
    informs the Outlook write-up. category="format" = example/template
    files defining the desired output format/structure — there can be
    several, and all of them are used as reference when generating the
    Outlook report."""
    return await OutlookMaterialService().save_files(db, plan_year, files, user.username, category=category)


@router.get("/setup-modules/outlook/materials")
async def list_outlook_materials(
    plan_year: Optional[int] = Query(None),
    category:  Optional[str] = Query(None, pattern="^(material|format)$"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List uploaded Outlook reference files, optionally filtered by year and category."""
    return await OutlookMaterialService().list_materials(db, plan_year, category)


async def _resolve_gemini_key(db: AsyncSession, user: CurrentUser) -> str:
    """User's own saved Gemini key if they set one, else the shared company key."""
    user_key = await user_api_key_service.get_user_key(db, user.username, "gemini")
    return user_key or settings.gemini_api_key


@router.post("/setup-modules/outlook/materials/{material_id}/convert")
async def convert_outlook_material(
    material_id: int,
    provider: str = Query("onprem", pattern="^(onprem|gemini|anthropic)$"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Extract text from the file and summarize it into a structured
    Markdown brief — done once per file, then reused by generate_outlook
    on every generation instead of re-reading the raw file each time.
    provider: "onprem" (local, free — default), "gemini", or "anthropic" (Claude)."""
    gemini_key = await _resolve_gemini_key(db, user) if provider == "gemini" else None
    result = await OutlookMaterialService().convert_material(db, material_id, provider, gemini_key)
    if result.get("data") is None:
        raise HTTPException(404, result.get("error", "Material tidak ditemukan"))
    return result


@router.get("/setup-modules/outlook/materials/{material_id}/download")
async def download_outlook_material(
    material_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Download a previously uploaded Outlook reference material."""
    service = OutlookMaterialService()
    row = await service.get_material(db, material_id)
    if not row:
        raise HTTPException(404, "Material tidak ditemukan")
    path = service.storage_path(row.filename)
    if not os.path.exists(path):
        raise HTTPException(404, "File tidak ditemukan di server")
    with open(path, "rb") as f:
        content = f.read()
    return StreamingResponse(
        io.BytesIO(content),
        media_type=row.content_type or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{row.original_name}"'},
    )


@router.delete("/setup-modules/outlook/materials/{material_id}")
async def delete_outlook_material(
    material_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete an uploaded Outlook reference material (file + row)."""
    return await OutlookMaterialService().delete_material(db, material_id)


class GenerateOutlookRequest(BaseModel):
    year: int
    context: Optional[str] = None
    provider: str = "onprem"  # "onprem" (local, free — default), "gemini", or "anthropic" (Claude)


@router.post("/setup-modules/generate-outlook")
async def generate_outlook(
    body: GenerateOutlookRequest,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """
    Generate Business Plan Outlook content using AI based on the selected year.
    Reads the already-converted briefs of uploaded reference materials/format
    files (not the raw files themselves — see convert_outlook_material) as
    grounding context, so generation is fast and consistent across reruns.
    Returns structured outlook data that can be saved via upsert.
    """
    if body.provider not in ("onprem", "gemini", "anthropic"):
        raise HTTPException(400, 'Invalid provider — use "onprem", "gemini", or "anthropic"')
    gemini_key = await _resolve_gemini_key(db, user) if body.provider == "gemini" else None

    mat_result = await OutlookMaterialService().list_materials(db, body.year)
    materials = mat_result.get("data", [])
    material_briefs = [m for m in materials if m["category"] == "material" and m["brief_status"] == "done" and m["brief_text"]]
    format_briefs = [m for m in materials if m["category"] == "format" and m["brief_status"] == "done" and m["brief_text"]]
    not_converted = sum(1 for m in materials if m["brief_status"] != "done")

    context_parts = []
    if material_briefs:
        context_parts.append("## Ringkasan Bahan Referensi (sumber data)\n" + "\n\n".join(
            f"### {m['original_name']}\n{m['brief_text']}" for m in material_briefs
        ))
    if format_briefs:
        context_parts.append("## Ringkasan Contoh Format Laporan (acuan struktur)\n" + "\n\n".join(
            f"### {m['original_name']}\n{m['brief_text']}" for m in format_briefs
        ))
    reference_context = "\n\n".join(context_parts)

    ai = AIService()
    use_web_search = body.provider == "anthropic"

    expert_persona = (
        "You are a senior macroeconomic and financial markets analyst — the kind "
        "of outside consultant a pharmaceutical manufacturer's board would retain "
        "to write its annual Business Plan outlook. Write with an expert's "
        "judgment: synthesize the data into a point of view, not just a list of "
        "figures. Each section should read like a page from a professional "
        "economic/industry outlook report — substantive and directly useful for "
        "management decision-making, not a superficial summary."
    )
    search_instruction = (
        "\n\nYou have web search available (max 8 searches — budget it "
        "carefully). Use it only for the 2-4 most important, time-sensitive "
        "figures per section (the current policy rate, latest inflation print, "
        "a key market-size number) — general trend commentary and structural "
        "analysis can rely on your own knowledge. Prefer primary/official "
        "sources (IMF, World Bank, Bank Indonesia, BPS, Kemenkes/BPOM, "
        "established industry reports) over secondary summaries. Do not "
        "narrate your search process in visible text — search silently, then "
        "respond with your final answer only."
        if use_web_search else ""
    )

    prompt = f"""{expert_persona}{search_instruction}

Generate a comprehensive Business Plan Outlook for PT CKD OTTO Pharmaceuticals for year {body.year}.
{f'Additional context: {body.context}' if body.context else ''}

{"Ground your answer in the reference material below — prefer these figures/trends over generic knowledge whenever they're relevant, and follow the structural cues from the format examples if given." if reference_context else "No converted reference materials are available yet for this year — generate a realistic, well-grounded outlook (upload and convert reference files first for an even more accurate result)."}
{f"{chr(10)}{chr(10)}{reference_context}" if reference_context else ""}

For each of the 3 sections below, write the content as a single Markdown
string: 1-2 short paragraphs of analyst framing/narrative FIRST, THEN a
"- " bullet list of the specific supporting figures and facts with key
terms and numbers in **bold**. This should read as expert commentary, not
a bare bullet dump — freeform text the user can edit directly, not a fixed
list of fields. Aim for real depth: 5-8 substantive bullets per section
minimum, each with a concrete number, trend, or specific policy/event —
not vague statements.

Return ONLY valid JSON with this exact structure:
{{
  "global_economic": {{
    "title": "I. Global Economic Outlook",
    "text": "Brief analyst framing paragraph.\\n\\n- **Global GDP Forecast**: ...\\n- **Key Factor 1**: ...\\n- **Fed Interest Rate**: ...\\n- **Global Inflation**: ..."
  }},
  "indonesia_economic": {{
    "title": "II. Indonesia Economic Outlook",
    "text": "Brief analyst framing paragraph.\\n\\n- **GDP Forecast**: ...\\n- **Inflation**: ...\\n- **Interest Rate**: ...\\n- **Exchange Rate**: ...\\n- **Government Focus**: ..."
  }},
  "pharmaceutical": {{
    "title": "III. Pharmaceutical Industry",
    "text": "Brief analyst framing paragraph.\\n\\n- **Global Market**: ...\\n- **Indonesia Market**: ...\\n- **Oncology**: ...\\n- **CKD OTTO Strategy**: ..."
  }}
}}

Make it realistic and current for {body.year}, with specific numbers, trends, and sources where relevant."""

    try:
        response = await ai.complete(
            "You are a senior financial/economic analyst producing a management report. Return only valid JSON, no markdown code fences, no explanations.",
            prompt,
            num_ctx=16384,
            provider=body.provider,
            gemini_api_key=gemini_key,
            web_search=use_web_search,
        )
        json_text = response.strip()
        if json_text.startswith("```"):
            json_text = json_text.strip("`")
            if json_text.lower().startswith("json"):
                json_text = json_text[4:]
        content = json.loads(json_text)
        return {
            "success": True,
            "data": {"setup_module": "outlook", "plan_year": body.year, "content": content, "status": "draft"},
            "materials_used": len(material_briefs),
            "format_examples_used": len(format_briefs),
            "not_converted": not_converted,
        }
    except Exception as e:
        return {"success": False, "error": f"Failed to generate outlook: {str(e)}", "data": None}


@router.get("/setup-modules/outlook/export")
async def export_outlook_ppt(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Business Plan Outlook (Global/Indonesia Economic +
    Pharmaceutical Industry) to PowerPoint — one slide per section, rendered
    from each section's Markdown text."""
    return await BusinessPlanSetupService().export_outlook_ppt(db, plan_year)


# ── Sales Plan ────────────────────────────────────────────────────────────────

class SalesPlanPayload(BaseModel):
    id:          Optional[int]  = None
    plan_year:   int
    department:  str            = ""
    team_code:   str            = ""
    team_name:   str            = ""
    plan_type:   str            = "value"   # value | unit
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/sales-plans")
async def list_sales_plans(
    plan_year:  Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    team_code:  Optional[str] = Query(None),
    plan_type:  Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List sales plans filtered by year/department/team/type."""
    return await SalesPlanService().list_sales_plans(db, plan_year, department, team_code, plan_type)


@router.get("/sales-plans/gross-sales-report")
async def export_gross_sales_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Gross Sales Report — same column layout and calculations
    (Sales Quantity = Sales Amount / Price) as sumber/output_grossales2026.xlsx,
    built entirely from Sales Plan Data (Value) rows for the given year.
    Registered before /sales-plans/{plan_id} — otherwise Starlette's string
    path matching would swallow this literal route as plan_id="gross-sales-report"
    and only fail with 422 at the int-conversion step."""
    result = await SalesPlanService().get_gross_sales_report_data(db, plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Sales Plan Data")
    try:
        return _build_gross_sales_report_xlsx(result["data"], plan_year)
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {e}")


@router.get("/sales-plans/sales-summary")
async def export_sales_summary(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Sales Summary — mirrors V5.Sales Estimation2026.xlsx's
    "Summary" sheet (business-unit breakdown with formulas referencing a raw
    Gross Sales data sheet in the same workbook), built entirely from Sales
    Plan Data for the given year. Registered before /sales-plans/{plan_id}
    for the same string-vs-int path matching reason as gross-sales-report."""
    result = await SalesPlanService().get_gross_sales_report_data(db, plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Sales Plan Data")
    try:
        return _build_sales_summary_xlsx(result["data"], plan_year)
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {e}")


@router.get("/sales-plans/report/export")
async def export_sales_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download of Sales Plan Simulation Data, formatted like
    "2-1.Sales plan_product_Value" in the Business Plan Report workbook.
    See _build_sales_plan_report_xlsx for what's deliberately left out
    (therapeutic-area split, Export's Product-vs-Freight split, no
    prior-year comparison). Registered before /sales-plans/{plan_id} for
    the same reason as /sales-plans/gross-sales-report."""
    result = await SalesPlanService().list_sales_plans(db, plan_year=plan_year, plan_type="value")
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Sales Plan data")
    return _build_sales_plan_report_xlsx(result["data"], plan_year)


@router.get("/sales-plans/{plan_id}")
async def get_sales_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single sales plan."""
    return await SalesPlanService().get_sales_plan(db, plan_id)


@router.post("/sales-plans")
async def upsert_sales_plan(
    body: SalesPlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update sales plan."""
    return await SalesPlanService().upsert_sales_plan(db, body.model_dump(), user.username)


@router.delete("/sales-plans/{plan_id}")
async def delete_sales_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete sales plan."""
    return await SalesPlanService().delete_sales_plan(db, plan_id)


@router.delete("/sales-plans")
async def delete_sales_plans_by_year(
    plan_year: int = Query(...),
    plan_type: Optional[str] = Query(None, description="value | unit — omit to delete both"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Bulk-delete every Sales Plan for a year (optionally narrowed to one
    plan_type) — for cleansing stale/duplicate data before a fresh
    re-upload, e.g. after a fix to how plans are matched/scaled on import."""
    return await SalesPlanService().delete_sales_plans_by_year(db, plan_year, plan_type)


@router.post("/sales-plans/{plan_id}/export")
async def export_sales_plan_excel(
    plan_id: int,
    plan_type: str = Query(..., description="value atau unit"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export Sales Plan ke Excel (S1 Value atau S2 Unit)."""
    return await SalesPlanService().export_excel(db, plan_id, plan_type, user.username)


@router.post("/sales-plans/upload")
async def upload_sales_plan_excel(
    plan_year: int = Query(...),
    sheets: Optional[str] = Query(None, description="Which sheets (1-based, in tab order) to process, e.g. '1', '1-2', '1,3,5-7' — omit to process every recognized sheet"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import a Sales Plan from an Excel file matching the "(S1) Sales
    plan_Value.xlsx" template — creates/overwrites the plan for whatever
    department/team the file's own meta section specifies."""
    content = await file.read()
    return await SalesPlanService().import_excel(db, content, plan_year, user.username, sheets)


def _write_gross_sales_sheet(ws, lines: list, plan_year: int) -> tuple:
    """Populate `ws` with the Gross Sales Report layout — mirrors
    output_grossales2026.xlsx: title row, a totals row, a two-row grouped
    header, then one data row per Sales Plan Data product. Columns
    requiring external product-master data the app doesn't track (Tech
    Transfer partner, active ingredient, product type, prior-year actuals)
    are intentionally left out — only what's derivable from Sales Plan Data
    (Market/Customer/Product/Price/Quantity/Amount) is populated; "Actual
    Sales" columns are kept for layout parity but always 0 since this app
    has no actuals feed. Returns (data_start_row, last_data_row) so callers
    (e.g. the Sales Summary sheet) can SUMIFS over this sheet by name."""
    COL_NO, COL_MARKET, COL_CUSTOMER, COL_PRODUCT = 2, 3, 4, 5
    COL_PRICE_ORIG_25, COL_PRICE_ORIG_26 = 6, 7
    COL_PRICE_IDR_25, COL_PRICE_IDR_26 = 8, 9
    COL_QTY_START, COL_QTY_END = 10, 22          # Jan..Dec + Total
    COL_AMT_START, COL_AMT_END = 23, 35          # Jan..Dec + Total
    COL_PRICE_UNIT = 37
    COL_ACT26_START, COL_ACT26_END = 38, 50      # Jan..Dec + Total
    COL_ACT25_START, COL_ACT25_END = 51, 63      # Jan..Dec + Total
    LAST_COL = COL_ACT25_END

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=13)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")

    def merge(r1, c1, r2, c2, value=None, font=None, align=center):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        if font: cell.font = font
        if align: cell.alignment = align
        return cell

    merge(1, 1, 1, 20, f"PT CKD OTTO Pharmaceuticals - Gross Sales Report {plan_year}", title_font, Alignment(horizontal="left"))

    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    HR1, HR2 = 3, 4

    def vgroup(col, label):
        merge(HR1, col, HR2, col, label, bold)
        ws.cell(row=HR1, column=col).fill = fill_hdr

    vgroup(COL_NO, "No")
    vgroup(COL_MARKET, "Market")
    vgroup(COL_CUSTOMER, "Customer")
    vgroup(COL_PRODUCT, "Products")
    merge(HR1, COL_PRICE_ORIG_25, HR1, COL_PRICE_ORIG_26, "Price (Original curr)", bold)
    ws.cell(row=HR2, column=COL_PRICE_ORIG_25, value="2025").font = bold
    ws.cell(row=HR2, column=COL_PRICE_ORIG_26, value=plan_year).font = bold
    merge(HR1, COL_PRICE_IDR_25, HR1, COL_PRICE_IDR_26, "Price (IDR)", bold)
    ws.cell(row=HR2, column=COL_PRICE_IDR_25, value="2025").font = bold
    ws.cell(row=HR2, column=COL_PRICE_IDR_26, value=plan_year).font = bold
    merge(HR1, COL_QTY_START, HR1, COL_QTY_END, "Sales Quantity", bold)
    for i, m in enumerate(MONTHS):
        ws.cell(row=HR2, column=COL_QTY_START + i, value=m).font = bold
    ws.cell(row=HR2, column=COL_QTY_END, value="Total").font = bold
    merge(HR1, COL_AMT_START, HR1, COL_AMT_END, "Sales Amount", bold)
    for i, m in enumerate(MONTHS):
        ws.cell(row=HR2, column=COL_AMT_START + i, value=m).font = bold
    ws.cell(row=HR2, column=COL_AMT_END, value="Total").font = bold
    vgroup(COL_PRICE_UNIT, f"Price / unit ({plan_year})")
    merge(HR1, COL_ACT26_START, HR1, COL_ACT26_END, f"Actual Sales {plan_year} (Qty)", bold)
    for i, m in enumerate(MONTHS):
        ws.cell(row=HR2, column=COL_ACT26_START + i, value=m).font = bold
    ws.cell(row=HR2, column=COL_ACT26_END, value="Total").font = bold
    merge(HR1, COL_ACT25_START, HR1, COL_ACT25_END, "Actual Sales 2025 (Value)", bold)
    for i, m in enumerate(MONTHS):
        ws.cell(row=HR2, column=COL_ACT25_START + i, value=m).font = bold
    ws.cell(row=HR2, column=COL_ACT25_END, value="Total").font = bold

    for c in range(1, LAST_COL + 1):
        ws.cell(row=HR1, column=c).fill = fill_hdr
        ws.cell(row=HR2, column=c).fill = fill_hdr

    NUMFMT = "#,##0"
    r0 = HR2 + 1              # totals row (row 5)
    data_start = r0 + 1       # first data row (row 6)
    last_row = data_start + len(lines) - 1 if lines else data_start

    for r_idx, line in enumerate(lines, start=data_start):
        no = r_idx - data_start + 1
        price = line["price"] or 0
        ws.cell(row=r_idx, column=COL_NO, value=no)
        ws.cell(row=r_idx, column=COL_MARKET, value=line["market"])
        ws.cell(row=r_idx, column=COL_CUSTOMER, value=line["customer"])
        ws.cell(row=r_idx, column=COL_PRODUCT, value=line["product"])
        ws.cell(row=r_idx, column=COL_PRICE_ORIG_26, value=price)
        ws.cell(row=r_idx, column=COL_PRICE_IDR_26, value=price)
        price_col = get_column_letter(COL_PRICE_IDR_26)
        # Sales Plan Data's monthly cells are Sales VALUE (Rp), not quantity
        # — Amount is the real entered figure, Quantity is derived from it.
        for i, amt in enumerate(line["amounts"]):
            ws.cell(row=r_idx, column=COL_AMT_START + i, value=amt)
        amt_first, amt_last = get_column_letter(COL_AMT_START), get_column_letter(COL_AMT_END - 1)
        ws.cell(row=r_idx, column=COL_AMT_END, value=f"=SUM({amt_first}{r_idx}:{amt_last}{r_idx})")
        for i in range(12):
            amt_col = get_column_letter(COL_AMT_START + i)
            ws.cell(row=r_idx, column=COL_QTY_START + i,
                    value=f"=IFERROR({amt_col}{r_idx}/${price_col}${r_idx},0)")
        qty_first, qty_last = get_column_letter(COL_QTY_START), get_column_letter(COL_QTY_END - 1)
        ws.cell(row=r_idx, column=COL_QTY_END, value=f"=SUM({qty_first}{r_idx}:{qty_last}{r_idx})")
        ws.cell(row=r_idx, column=COL_PRICE_UNIT, value=f"={price_col}{r_idx}")
        for c in range(COL_ACT26_START, COL_ACT25_END + 1):
            ws.cell(row=r_idx, column=c, value=0)
        for c in range(COL_PRICE_ORIG_25, LAST_COL + 1):
            ws.cell(row=r_idx, column=c).number_format = NUMFMT

    # Totals row — SUM over the data range, positioned above the header like
    # the reference file's SUBTOTAL row.
    ws.cell(row=r0, column=COL_PRODUCT, value="GRAND TOTAL").font = bold
    if lines:
        for c in list(range(COL_QTY_START, COL_AMT_END + 1)) + [COL_ACT26_END, COL_ACT25_END]:
            col = get_column_letter(c)
            cell = ws.cell(row=r0, column=c, value=f"=SUM({col}{data_start}:{col}{last_row})")
            cell.font = bold
            cell.number_format = NUMFMT

    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 28
    for c in range(6, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = f"{get_column_letter(COL_PRODUCT + 1)}{data_start}"

    return data_start, last_row


def _build_gross_sales_report_xlsx(lines: list, plan_year: int) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Grosssales_{plan_year}"
    _write_gross_sales_sheet(ws, lines, plan_year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Gross_Sales_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


_MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _build_manufacture_plan_detail_xlsx(rows: list, plan_year: int) -> StreamingResponse:
    """rows: raw ManufacturePlan.content.rows lists, shape matches
    manufacture_plan_service.HEADERS:
      [No, Customer, ItemCode, Name, BatchSize, Yield%, Jan..Dec(12),
       TotalBatch, QtyBeforeYield, QtyAfterYield, SalesQty, Coverage]

    Total Batch / Qty Before Yield / Qty After Yield / Coverage are
    RECOMPUTED here from the raw inputs (monthly batches, batch size,
    yield%, sales qty) rather than trusting the stored derived columns —
    the manufacture plan editor only keeps "Total Batch" in sync when a
    monthly cell changes (see PAC.jsx's updateCell), not the downstream
    yield/coverage columns, so a report built off stale stored values
    could silently disagree with its own inputs. Formulas:
      Total Batch            = SUM(Jan..Dec)
      Total Qty Before Yield = Total Batch * Batch Size
      Total Qty After Yield  = Total Qty Before Yield * Yield%
      Coverage                = Total Qty After Yield / Sales Quantity
    — same chain as sumber/'BOM, Manufacturing plan.xlsx' ('Detail
    Manufacturing plan_All': W = S*T*V, Coverage = After-Yield / Sales Plan).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Detail_Manufacturing_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_total = PatternFill("solid", fgColor="F2F2F2")

    headers = (
        ["No", "Business Type", "Item Code", "Product Name"]
        + _MONTH_LABELS
        + ["Total Batch", "Batch Size (Vial)", "Total Qty Before Yield", "Yield (%)",
           "Total Production After Yield (Vial)", "Sales Quantity (Vial)", "Coverage"]
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=f"Manufacturing Plan Detail — {plan_year}").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    # sort by Business Type then Item Code, mirroring the reference file's
    # grouping (products clustered by Local/Export/CMO)
    def sort_key(r):
        return (str(r[1] if len(r) > 1 else ""), str(r[2] if len(r) > 2 else ""))

    sorted_rows = sorted(rows, key=sort_key)

    r = HR + 1
    totals = {"batch": 0, "before": 0.0, "after": 0.0, "sales": 0.0}
    for i, row in enumerate(sorted_rows, 1):
        customer   = row[1] if len(row) > 1 else ""
        item_code  = row[2] if len(row) > 2 else ""
        name       = row[3] if len(row) > 3 else ""
        batch_size = float(row[4] or 0) if len(row) > 4 else 0.0
        yield_pct  = float(row[5] or 0) if len(row) > 5 else 0.0
        months     = [float(row[6 + m] or 0) if len(row) > 6 + m else 0.0 for m in range(12)]
        sales_qty  = float(row[21] or 0) if len(row) > 21 else 0.0

        total_batch = sum(months)
        qty_before  = total_batch * batch_size
        qty_after   = qty_before * yield_pct
        coverage    = (qty_after / sales_qty) if sales_qty else None

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=customer)
        ws.cell(row=r, column=3, value=item_code)
        ws.cell(row=r, column=4, value=name)
        for m in range(12):
            ws.cell(row=r, column=5 + m, value=months[m] or None)
        ws.cell(row=r, column=17, value=total_batch)
        ws.cell(row=r, column=18, value=batch_size)
        ws.cell(row=r, column=19, value=qty_before)
        c_yield = ws.cell(row=r, column=20, value=yield_pct)
        c_yield.number_format = "0.0%"
        ws.cell(row=r, column=21, value=qty_after)
        ws.cell(row=r, column=22, value=sales_qty)
        c_cov = ws.cell(row=r, column=23, value=coverage)
        if coverage is not None:
            c_cov.number_format = "0.00"

        totals["batch"]  += total_batch
        totals["before"] += qty_before
        totals["after"]  += qty_after
        totals["sales"]  += sales_qty
        r += 1

    last_row = r - 1
    ws.cell(row=r, column=4, value="TOTAL").font = bold
    ws.cell(row=r, column=17, value=totals["batch"]).font = bold
    ws.cell(row=r, column=19, value=totals["before"]).font = bold
    ws.cell(row=r, column=21, value=totals["after"]).font = bold
    ws.cell(row=r, column=22, value=totals["sales"]).font = bold
    total_cov = (totals["after"] / totals["sales"]) if totals["sales"] else None
    if total_cov is not None:
        c = ws.cell(row=r, column=23, value=total_cov)
        c.font, c.number_format = bold, "0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).fill = fill_total

    ws.freeze_panes = ws.cell(row=HR + 1, column=5)
    ws.column_dimensions["D"].width = 26
    for col_letter in ["B", "C"]:
        ws.column_dimensions[col_letter].width = 14
    for m in range(12):
        ws.column_dimensions[get_column_letter(5 + m)].width = 6
    for col_letter in ["Q", "R", "S", "T", "U", "V", "W"]:
        ws.column_dimensions[col_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Manufacturing_Plan_Detail_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_manufacture_plan_report_xlsx(report: dict, plan_year: int) -> StreamingResponse:
    """Reporting > Manufacturing Plan — same hierarchical layout (indented
    Total -> Local/CMO/Export -> customer sub-group -> Liquid/Freeze Dry ->
    product) as the in-app report table built by
    ManufacturePlanService.get_report(), just rendered to .xlsx instead of
    HTML. `report` is that method's return value."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Manufacturing_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    row_fill = {
        "total":    PatternFill("solid", fgColor="D9C6F2"),
        "group":    PatternFill("solid", fgColor="F2F2F2"),
        "subtotal": PatternFill("solid", fgColor="FAFAFA"),
    }
    row_font = {"total": bold, "group": bold, "subtotal": Font(italic=True)}

    columns = report["columns"]
    headers = ["Product / Group"] + columns
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Manufacturing Plan Report — {plan_year} (unit: vial) — from Simulation Data > Manufacture Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    r = HR + 1
    for row in report["rows"]:
        label_cell = ws.cell(row=r, column=1, value=("  " * (row.get("level") or 0)) + row["label"])
        for ci, v in enumerate(row.get("values") or [], 2):
            ws.cell(row=r, column=ci, value=v if v else None)
        font, fill = row_font.get(row["type"]), row_fill.get(row["type"])
        if font:
            label_cell.font = font
            for c in range(2, ncols + 1):
                ws.cell(row=r, column=c).font = font
        if fill:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    ws.freeze_panes = ws.cell(row=HR + 1, column=2)
    ws.column_dimensions["A"].width = 32
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Manufacturing_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_investment_plan_report_xlsx(plans: list, plan_year: int) -> StreamingResponse:
    """Reporting format reference: sumber/01.V3.2026 Business_Plan_Report_
    Dec20_2025.xlsx, sheet "5. Investment Plan" — grouped by Classification
    (each item's Clarification field), Q1-Q4 + annual Total per item and
    per group, Team from each plan's own team_code.

    Deliberately NOT reproduced: the reference sheet's Acquisition Month /
    Depreciation Charged (COGS·OPEX) / Depreciation Amount columns and its
    bottom depreciation-summary block. Investment Plan Simulation Data has
    no "acquisition month" field to derive them from reliably — reverse-
    engineering it from which Jan-Dec column holds the entered value gave
    inconsistent results against two known reference rows (Head Space
    Analyzer implied depreciation starting the month AFTER acquisition;
    Technical Transfer implied starting the SAME month) — a real accounting
    policy nuance (or manual override) this data doesn't capture, so
    fabricating a depreciation figure risked being confidently wrong on a
    number people would actually rely on. Lifetime (Year) is included as
    entered; Acquisition/Depreciation are left for Accounting to add from
    the source workbook until there's a reliable way to derive them.
    """
    from openpyxl.utils import get_column_letter as _col_letter

    def _num(v):
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    rows_by_clarification: dict = {}
    for plan in plans:
        team = plan.get("team_code") or plan.get("team_name") or plan.get("department") or ""
        for row in (plan.get("content") or {}).get("rows") or []:
            if len(row) < 21:
                continue
            clarification = str(row[1] or "").strip()
            item = str(row[3] or "").strip()
            if not clarification or not item:
                continue
            months = [_num(row[8 + m]) for m in range(12)]
            quarters = [sum(months[0:3]), sum(months[3:6]), sum(months[6:9]), sum(months[9:12])]
            lifetime = row[7]
            notes = row[21] if len(row) > 21 else ""
            rows_by_clarification.setdefault(clarification, []).append({
                "item": item, "team": team, "quarters": quarters, "total": sum(quarters),
                "lifetime": lifetime, "notes": notes,
            })

    if not rows_by_clarification:
        raise HTTPException(404, f"Tidak ada data Investment Plan untuk tahun {plan_year}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Investment_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_group = PatternFill("solid", fgColor="F2F2F2")
    fill_total = PatternFill("solid", fgColor="D9C6F2")

    headers = ["No", "Classification / Item", "Team", "Q1", "Q2", "Q3", "Q4", "Total", "Lifetime (Year)", "Notes"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Investment Plan — {plan_year} (unit: mil Rp) — from Simulation Data > Investment Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    r = HR + 1
    grand_q = [0.0, 0.0, 0.0, 0.0]
    for no, clarification in enumerate(sorted(rows_by_clarification), 1):
        items = rows_by_clarification[clarification]
        group_q = [sum(it["quarters"][i] for it in items) for i in range(4)]
        ws.cell(row=r, column=1, value=no).font = bold
        ws.cell(row=r, column=2, value=clarification).font = bold
        for i in range(4):
            ws.cell(row=r, column=4 + i, value=group_q[i] or None).font = bold
        ws.cell(row=r, column=8, value=sum(group_q) or None).font = bold
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = fill_group
        for i in range(4):
            grand_q[i] += group_q[i]
        r += 1

        for it in items:
            ws.cell(row=r, column=2, value=it["item"])
            ws.cell(row=r, column=3, value=it["team"])
            for i in range(4):
                ws.cell(row=r, column=4 + i, value=it["quarters"][i] or None)
            ws.cell(row=r, column=8, value=it["total"] or None)
            ws.cell(row=r, column=9, value=it["lifetime"] or None)
            ws.cell(row=r, column=10, value=it["notes"] or None)
            r += 1

    ws.cell(row=r, column=2, value="Total Investment").font = bold
    for i in range(4):
        ws.cell(row=r, column=4 + i, value=grand_q[i] or None).font = bold
    ws.cell(row=r, column=8, value=sum(grand_q) or None).font = bold
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = fill_total

    ws.freeze_panes = ws.cell(row=HR + 1, column=3)
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    for c in range(4, 9):
        ws.column_dimensions[_col_letter(c)].width = 13
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Investment_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_purchase_plan_report_xlsx(plans_this: list, plans_prior: list, plan_year: int) -> StreamingResponse:
    """Reporting format reference: sumber/01.V3.2026 Business_Plan_Report_
    Dec20_2025.xlsx, sheet "6-1.Purchase_Plan_Value" — grouped by Type
    (API/Excipient/Primary Packaging/...), with a prior-year comparison and
    growth % per item, matching the reference's "2025(E)"/"2026(P)"/"% G/R"
    columns.

    Deliberately NOT reproduced: Vendor (not captured anywhere in Purchase
    Plan Simulation Data), and the reference's separate "Delivery 2026" /
    "Outstanding PO 2027" QTY+Value columns — this input only distinguishes
    planned Order vs planned Received quantities, which isn't reliably the
    same concept as a delivery schedule or PO-slippage carryover into next
    year; fabricating those would risk presenting a guess as a real supply
    figure. Prior-year ("2025(E)") figures come from this same app's
    Purchase Plan Simulation Data for plan_year-1, if any was entered —
    not from a separate actuals source.
    """
    def _num(v):
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    def _item_key(item: dict):
        code = str(item.get("item_code_no") or "").strip()
        name = str(item.get("item_code_name") or "").strip()
        return (str(item.get("type") or "").strip(), code or name)

    prior_by_key: dict = {}
    for plan in plans_prior:
        for item in (plan.get("content") or {}).get("items") or []:
            k = _item_key(item)
            agg = prior_by_key.setdefault(k, {"qty": 0.0, "value": 0.0})
            agg["qty"] += _num(item.get("order_total"))
            agg["value"] += _num(item.get("total_price"))

    rows_by_type: dict = {}
    for plan in plans_this:
        for item in (plan.get("content") or {}).get("items") or []:
            name = str(item.get("item_code_name") or "").strip()
            if not name:
                continue
            item_type = str(item.get("type") or "").strip() or "(Unclassified)"
            prior = prior_by_key.get(_item_key(item), {"qty": 0.0, "value": 0.0})
            rows_by_type.setdefault(item_type, []).append({
                "name": name, "uom": item.get("uom") or "",
                "unit_price_idr": _num(item.get("unit_price_idr")),
                "qty_prior": prior["qty"], "value_prior": prior["value"],
                "qty_this": _num(item.get("order_total")), "value_this": _num(item.get("total_price")),
            })

    if not rows_by_type:
        raise HTTPException(404, f"Tidak ada data Purchase Plan untuk tahun {plan_year}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Purchase_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_group = PatternFill("solid", fgColor="F2F2F2")
    fill_total = PatternFill("solid", fgColor="D9C6F2")

    headers = ["No", "Type / Material Name", "UOM", "Price (IDR)",
               f"{plan_year - 1}(E) QTY", f"{plan_year - 1}(E) Value", f"{plan_year}(P) QTY", f"{plan_year}(P) Value", "% G/R"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Purchase Plan — {plan_year} (unit: mil Rp) — from Simulation Data > Purchase Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    def _growth(prior_v, this_v):
        return (this_v - prior_v) / prior_v if prior_v else None

    r = HR + 1
    grand = {"qty_prior": 0.0, "value_prior": 0.0, "qty_this": 0.0, "value_this": 0.0}
    for no, item_type in enumerate(sorted(rows_by_type), 1):
        items = rows_by_type[item_type]
        g = {k: sum(it[k] for it in items) for k in ("qty_prior", "value_prior", "qty_this", "value_this")}
        ws.cell(row=r, column=1, value=no).font = bold
        ws.cell(row=r, column=2, value=item_type).font = bold
        ws.cell(row=r, column=5, value=g["qty_prior"] or None).font = bold
        ws.cell(row=r, column=6, value=g["value_prior"] or None).font = bold
        ws.cell(row=r, column=7, value=g["qty_this"] or None).font = bold
        ws.cell(row=r, column=8, value=g["value_this"] or None).font = bold
        gr = _growth(g["value_prior"], g["value_this"])
        gr_cell = ws.cell(row=r, column=9, value=gr)
        gr_cell.font, gr_cell.number_format = bold, "0.0%"
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = fill_group
        for k in grand:
            grand[k] += g[k]
        r += 1

        for it in items:
            ws.cell(row=r, column=2, value=it["name"])
            ws.cell(row=r, column=3, value=it["uom"])
            ws.cell(row=r, column=4, value=it["unit_price_idr"] or None)
            ws.cell(row=r, column=5, value=it["qty_prior"] or None)
            ws.cell(row=r, column=6, value=it["value_prior"] or None)
            ws.cell(row=r, column=7, value=it["qty_this"] or None)
            ws.cell(row=r, column=8, value=it["value_this"] or None)
            gr = _growth(it["value_prior"], it["value_this"])
            if gr is not None:
                ws.cell(row=r, column=9, value=gr).number_format = "0.0%"
            r += 1

    ws.cell(row=r, column=2, value="Total Purchase Plan").font = bold
    ws.cell(row=r, column=5, value=grand["qty_prior"] or None).font = bold
    ws.cell(row=r, column=6, value=grand["value_prior"] or None).font = bold
    ws.cell(row=r, column=7, value=grand["qty_this"] or None).font = bold
    ws.cell(row=r, column=8, value=grand["value_this"] or None).font = bold
    gr_cell = ws.cell(row=r, column=9, value=_growth(grand["value_prior"], grand["value_this"]))
    gr_cell.font, gr_cell.number_format = bold, "0.0%"
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = fill_total

    ws.freeze_panes = ws.cell(row=HR + 1, column=3)
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 10
    for c in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Purchase_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _classify_business_unit(market: str) -> str:
    """Business-unit bucket rule mirrored from V5.Sales Estimation2026.xlsx's
    Summary sheet: Local = Public/Private, CMO/Export are their own units,
    everything else (Service Agreement, Dossier Fee, ...) falls into Others —
    fully data-driven off whatever Market values actually appear, no
    hardcoded customer names."""
    m = (market or "").strip().lower()
    if m in ("public", "private"):
        return "Local"
    if m == "cmo":
        return "CMO"
    if m == "export":
        return "Export"
    return "Others"


def _esc(text: str) -> str:
    return str(text or "").replace('"', '""')


def _build_sales_summary_xlsx(lines: list, plan_year: int) -> StreamingResponse:
    """Sales Summary — mirrors V5.Sales Estimation2026.xlsx's "Summary"
    sheet: a business-unit breakdown (Local/CMO/Others/Export, each broken
    into its Market or Customer sub-rows) with SUMIFS formulas pointing at
    a raw data sheet in the SAME workbook — same cross-sheet-formula
    structure as the reference file's Summary -> Grosssales_2026 link.
    Scope note: only the current plan year's Business Plan figures are
    shown (Value in Rp mil, Quantity) — the reference file's 2024 Actual /
    2025 Estimation / Achievement / CAGR columns depend on prior-year
    actuals this app has no data source for, so they're intentionally
    left out rather than faked as blank/zero."""
    wb = openpyxl.Workbook()
    ws_gross = wb.create_sheet(f"Grosssales_{plan_year}")
    data_start, last_row = _write_gross_sales_sheet(ws_gross, lines, plan_year)
    gsheet = ws_gross.title

    ws = wb.active
    ws.title = "Summary"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_unit = PatternFill("solid", fgColor="F2F2F2")

    def merge(r1, c1, r2, c2, value=None, font=None, align=center):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        if font: cell.font = font
        if align: cell.alignment = align
        return cell

    merge(1, 1, 1, 5, f"PT CKD OTTO Pharmaceuticals - {plan_year} Sales Summary", title_font, Alignment(horizontal="left"))
    merge(2, 1, 2, 5, "Summary Sales by Business Unit — derived from Gross Sales Report / Sales Plan Data", Font(italic=True, size=10), Alignment(horizontal="left"))

    HR = 4
    headers = ["Business Unit", "", f"{plan_year} Business Plan Value (Rp mil)", f"{plan_year} Business Plan Quantity", "% of Total Value"]
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    # Group lines into business-unit -> sub-group (Market for Local/Others,
    # Customer for CMO/Export) — same bucket rule as the reference file's
    # own Local/CMO/Others/Export sections.
    UNIT_ORDER = ["Local", "CMO", "Others", "Export"]
    units: dict = {u: {} for u in UNIT_ORDER}
    for line in lines:
        unit = _classify_business_unit(line["market"])
        if unit in ("Local", "Others"):
            key = line["market"] or "(unknown)"
            units[unit].setdefault(key, {"market": line["market"], "customer": None})
        else:
            key = line["customer"] or "(no customer)"
            units[unit].setdefault(key, {"market": line["market"], "customer": line["customer"]})

    # Pass 1 — assign row numbers before writing any formulas, since the
    # unit header row's SUM needs to reference its sub-rows' final range.
    row_plan = []  # (unit, header_row, [sub_rows...])
    r = HR + 1
    for unit in UNIT_ORDER:
        if not units[unit]:
            continue
        header_row = r
        r += 1
        sub_rows = list(range(r, r + len(units[unit])))
        r += len(units[unit])
        row_plan.append((unit, header_row, sub_rows))
    grand_total_row = r

    NUMFMT = "#,##0"
    PCTFMT = "0.0%"

    def write_row(row, label, value_formula, qty_formula, is_bold, fill=None, indent=False):
        c1 = ws.cell(row=row, column=1, value=label)
        if indent:
            c1.alignment = Alignment(indent=1)
        c3 = ws.cell(row=row, column=3, value=value_formula)
        c4 = ws.cell(row=row, column=4, value=qty_formula)
        c5 = ws.cell(row=row, column=5, value=f"=IFERROR(C{row}/$C${grand_total_row},0)")
        c3.number_format = NUMFMT
        c4.number_format = NUMFMT
        c5.number_format = PCTFMT
        if is_bold:
            for cell in (c1, c3, c4, c5):
                cell.font = bold
        if fill:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = fill

    for unit, header_row, sub_rows in row_plan:
        keys = list(units[unit].keys())
        for sub_row, key in zip(sub_rows, keys):
            info = units[unit][key]
            market_f = _esc(info["market"])
            if info["customer"]:
                customer_f = _esc(info["customer"])
                value_formula = f'=SUMIFS({gsheet}!$AI:$AI,{gsheet}!$C:$C,"{market_f}",{gsheet}!$D:$D,"{customer_f}")/1000000'
                qty_formula = f'=SUMIFS({gsheet}!$V:$V,{gsheet}!$C:$C,"{market_f}",{gsheet}!$D:$D,"{customer_f}")'
            else:
                value_formula = f'=SUMIFS({gsheet}!$AI:$AI,{gsheet}!$C:$C,"{market_f}")/1000000'
                qty_formula = f'=SUMIFS({gsheet}!$V:$V,{gsheet}!$C:$C,"{market_f}")'
            write_row(sub_row, key, value_formula, qty_formula, is_bold=False, indent=True)

        if sub_rows:
            v_first, v_last = f"C{sub_rows[0]}", f"C{sub_rows[-1]}"
            q_first, q_last = f"D{sub_rows[0]}", f"D{sub_rows[-1]}"
            write_row(header_row, unit, f"=SUM({v_first}:{v_last})", f"=SUM({q_first}:{q_last})", is_bold=True, fill=fill_unit)
        else:
            write_row(header_row, unit, 0, 0, is_bold=True, fill=fill_unit)

    if row_plan:
        header_rows = [hr for _, hr, _ in row_plan]
        v_terms = "+".join(f"C{r}" for r in header_rows)
        q_terms = "+".join(f"D{r}" for r in header_rows)
        write_row(grand_total_row, "GRAND TOTAL", f"={v_terms}", f"={q_terms}", is_bold=True, fill=fill_unit)
    else:
        write_row(grand_total_row, "GRAND TOTAL", 0, 0, is_bold=True, fill=fill_unit)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = f"A{HR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Sales_Summary_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Purchase Plan (Material) ────────────────────────────────────────────────────

class PurchasePlanPayload(BaseModel):
    id:            Optional[int]  = None
    plan_year:     int
    plan_category: str            = "Local"   # Summary | Local | CMO | Export
    department:    str            = ""
    team_code:     str            = ""
    team_name:     str            = ""
    content:       dict           = {}
    status:        Optional[str]  = "draft"


@router.get("/purchase-plans")
async def list_purchase_plans(
    plan_year:     Optional[int] = Query(None),
    department:    Optional[str] = Query(None),
    team_code:     Optional[str] = Query(None),
    plan_category: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List purchase plans filtered by year/department/team/category."""
    return await PurchasePlanService().list_purchase_plans(db, plan_year, department, team_code, plan_category)


@router.get("/purchase-plans/report/export")
async def export_purchase_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download of Purchase Plan Simulation Data, formatted like
    "6-1.Purchase_Plan_Value" in the Business Plan Report workbook. See
    _build_purchase_plan_report_xlsx for what's deliberately left out
    (Vendor, Delivery/Outstanding PO columns).
    Registered before /purchase-plans/{plan_id} for the same reason as
    /manufacture-plans/detail-report."""
    this_year = await PurchasePlanService().list_purchase_plans(db, plan_year=plan_year)
    if not this_year.get("success"):
        raise HTTPException(400, this_year.get("error") or "Failed to load Purchase Plan data")
    prior_year = await PurchasePlanService().list_purchase_plans(db, plan_year=plan_year - 1)
    return _build_purchase_plan_report_xlsx(this_year["data"], prior_year.get("data") or [], plan_year)


@router.get("/purchase-plans/{plan_id}")
async def get_purchase_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single purchase plan."""
    return await PurchasePlanService().get_purchase_plan(db, plan_id)


@router.post("/purchase-plans")
async def upsert_purchase_plan(
    body: PurchasePlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update purchase plan."""
    return await PurchasePlanService().upsert_purchase_plan(db, body.model_dump(), user.username)


@router.delete("/purchase-plans/{plan_id}")
async def delete_purchase_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete purchase plan."""
    return await PurchasePlanService().delete_purchase_plan(db, plan_id)


@router.post("/purchase-plans/upload")
async def upload_purchase_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import Purchase Plan (Material) data from an Excel file matching the
    "(P1-M) Purchase plan_Material.xlsx" template — one plan created/updated
    per recognized data sheet (Summary/Local/CMO/Export)."""
    content = await file.read()
    return await PurchasePlanService().import_excel(db, content, plan_year, user.username)


# ── Personnel Plan ("Personal Plan Data") ──────────────────────────────────────

class PersonnelPlanPayload(BaseModel):
    id:          Optional[int]  = None
    plan_year:   int
    department:  str            = ""
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/personnel-plans")
async def list_personnel_plans(
    plan_year:  Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List personnel plans filtered by year/department."""
    return await PersonnelPlanService().list_personnel_plans(db, plan_year, department)


@router.get("/personnel-plans/report/export")
async def export_personnel_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download of Personnel Plan Simulation Data, formatted like the
    "2) Personnel Plan" section of sheet "9.Personnel plan" in the
    Business Plan Report workbook — the Organization Chart section is
    intentionally excluded. See _build_personnel_plan_report_xlsx for the
    one grouping difference (Level, not Team/Function).
    Registered before /personnel-plans/{plan_id} for the same reason as
    /manufacture-plans/detail-report."""
    result = await PersonnelPlanService().list_personnel_plans(db, plan_year=plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Personnel Plan data")
    return _build_personnel_plan_report_xlsx(result["data"], plan_year)


@router.get("/personnel-plans/{plan_id}")
async def get_personnel_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single personnel plan."""
    return await PersonnelPlanService().get_personnel_plan(db, plan_id)


@router.post("/personnel-plans")
async def upsert_personnel_plan(
    body: PersonnelPlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update personnel plan."""
    return await PersonnelPlanService().upsert_personnel_plan(db, body.model_dump(), user.username)


@router.delete("/personnel-plans/{plan_id}")
async def delete_personnel_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete personnel plan."""
    return await PersonnelPlanService().delete_personnel_plan(db, plan_id)


@router.post("/personnel-plans/upload")
async def upload_personnel_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import Personnel Plan data from an Excel file matching the "Personal
    plan template.xlsx" layout (headcount by level + recruitment schedules)."""
    content = await file.read()
    return await PersonnelPlanService().import_excel(db, content, plan_year, user.username)


def _build_sales_plan_report_xlsx(plans: list, plan_year: int) -> StreamingResponse:
    """Reporting format reference: sumber/01.V3.2026 Business_Plan_Report_
    Dec20_2025.xlsx, sheet "2-1.Sales plan_product_Value" — Total ->
    Local (Public/Private) / CMO & Others (CMO/Service Agreement) / Export
    (per country) -> product, with a Jan-Dec monthly breakdown + annual
    Total, matching that sheet's layout. Local/CMO/Export and Public/
    Private/CMO/Service-Agreement all come from each plan's Excel tab name
    (content.meta.sheet_name) — same substring convention already used by
    eis_summary.py's Sales Closing Estimation card — Export's per-country
    split comes from each row's own Country field.

    Deliberately NOT reproduced: the reference sheet's further Oncology /
    Immunosuppressant therapeutic-area split (no such classification
    exists anywhere in Sales Plan Simulation Data — it only has Country/
    Customer/Product) and its Export rows' Product-vs-Freight value split
    (Sales Plan stores one blended Total Value per product, not a separate
    freight charge). No prior-year ("2025(E)") comparison either — Sales
    Plan Simulation Data is entered fresh each cycle from the S1 template,
    not carried forward year to year the way Manufacture Plan is.
    """
    MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def _classify(sheet_name):
        s = str(sheet_name or "").strip().lower().replace("_", " ")
        if "national public" in s:
            return ("Local", "Public")
        if "national private" in s:
            return ("Local", "Private")
        if "export" in s:
            return ("Export", None)
        if "cmo" in s:
            return ("CMO & Others", "CMO")
        if "agreement" in s:
            return ("CMO & Others", "Service Agreement")
        return ("Other", sheet_name or "(Unclassified)")

    tree: dict = {}  # tree[top][sub_or_country][product] = [12 floats]
    for plan in plans:
        sheet_name = (plan.get("content") or {}).get("meta", {}).get("sheet_name")
        top, sub = _classify(sheet_name)
        for row in (plan.get("content") or {}).get("rows") or []:
            if len(row) < 16:
                continue
            product = str(row[3] or "").strip()
            if not product:
                continue
            months = [float(row[4 + m] or 0) for m in range(12)]
            if not any(months):
                continue
            key2 = sub if sub else (str(row[1] or "").strip() or "(Unspecified)")
            arr = tree.setdefault(top, {}).setdefault(key2, {}).setdefault(product, [0.0] * 12)
            for i in range(12):
                arr[i] += months[i]

    if not tree:
        raise HTTPException(404, f"Tidak ada data Sales Plan untuk tahun {plan_year}")

    def sum_months(node):
        if isinstance(node, list):
            return list(node)
        total = [0.0] * 12
        for child in node.values():
            cs = sum_months(child)
            for i in range(12):
                total[i] += cs[i]
        return total

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sales_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_top = PatternFill("solid", fgColor="D9C6F2")
    fill_sub = PatternFill("solid", fgColor="F2F2F2")

    headers = ["Product / Group"] + MONTH_LABELS + ["Total"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Sales Plan — {plan_year} (unit: mil Rp) — from Simulation Data > Sales Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    r = HR + 1

    def emit(label, level, months, fill=None, bold_row=False):
        nonlocal r
        cell = ws.cell(row=r, column=1, value=("  " * level) + label)
        if bold_row:
            cell.font = bold
        for i, v in enumerate(months, 2):
            c = ws.cell(row=r, column=i, value=v or None)
            if bold_row:
                c.font = bold
        tot = ws.cell(row=r, column=ncols, value=sum(months) or None)
        if bold_row:
            tot.font = bold
        if fill:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    grand_total = [0.0] * 12
    for top in ["Local", "CMO & Others", "Export", "Other"]:
        if top not in tree:
            continue
        top_months = sum_months(tree[top])
        emit(top, 0, top_months, fill=fill_top, bold_row=True)
        for i in range(12):
            grand_total[i] += top_months[i]
        for sub in sorted(tree[top]):
            sub_months = sum_months(tree[top][sub])
            emit(sub, 1, sub_months, fill=fill_sub, bold_row=True)
            for product in sorted(tree[top][sub]):
                emit(product, 2, tree[top][sub][product])

    emit("TOTAL SALES PLAN", 0, grand_total, fill=fill_top, bold_row=True)

    ws.freeze_panes = ws.cell(row=HR + 1, column=2)
    ws.column_dimensions["A"].width = 34
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Sales_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_personnel_plan_report_xlsx(plans: list, plan_year: int) -> StreamingResponse:
    """Reporting format reference: sumber/01.V3.2026 Business_Plan_Report_
    Dec20_2025.xlsx, sheet "9.Personnel plan" — its "2) Personnel Plan"
    section only (headcount by Department -> Level, Prev/Curr Year +
    Increase, each split Permanent/Temporary/Total). The sheet's "1)
    Organization Chart" section is out of scope per the user's request —
    it's hand-drawn/free-text in the source file anyway, nothing in
    Personnel Plan Simulation Data corresponds to it.

    One difference from the reference worth knowing: the reference's
    second grouping level under each Department is by TEAM/FUNCTION (e.g.
    Plant -> "Quality Management"/"Production"/"Engineering"). Personnel
    Plan Simulation Data groups headcount by LEVEL instead (e.g. "General
    Manager"/"Manager"/"Staff") — that's the dimension this input actually
    captures, so it's what's reported here; there's no function/team field
    to fall back to."""
    FIELDS = ["prev_permanent", "prev_temporary", "prev_total",
              "curr_permanent", "curr_temporary", "curr_total",
              "inc_permanent", "inc_temporary", "inc_total"]

    def sum_rows(rows):
        return {f: sum(float(r.get(f) or 0) for r in rows) for f in FIELDS}

    dept_rows: dict = {}
    dept_totals: dict = {}
    year_prev = year_curr = None
    for plan in plans:
        dept = str(plan.get("department") or "").strip() or "(Unspecified)"
        hc = (plan.get("content") or {}).get("headcount") or {}
        if hc.get("year_prev") and not year_prev:
            year_prev = hc["year_prev"]
        if hc.get("year_curr") and not year_curr:
            year_curr = hc["year_curr"]
        dept_rows.setdefault(dept, []).extend(hc.get("rows") or [])
        # Trust the uploaded file's own department total when present
        # (matches whatever the source Excel said), fall back to summing
        # the level rows only when it's missing.
        if hc.get("total"):
            dept_totals[dept] = hc["total"]

    if not dept_rows:
        raise HTTPException(404, f"Tidak ada data Personnel Plan untuk tahun {plan_year}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Personnel_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_dept = PatternFill("solid", fgColor="F2F2F2")
    fill_total = PatternFill("solid", fgColor="D9C6F2")

    y_prev = year_prev or (plan_year - 1)
    y_curr = year_curr or plan_year
    headers = ["Department / Level",
               f"{y_prev} Permanent", f"{y_prev} Temporary", f"{y_prev} Total",
               f"{y_curr} Permanent", f"{y_curr} Temporary", f"{y_curr} Total",
               "Increase Permanent", "Increase Temporary", "Increase Total"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Personnel Plan — {plan_year} (headcount) — from Simulation Data > Personnel Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    def emit_row(r, label, values, level, fill=None, bold_row=False):
        cell = ws.cell(row=r, column=1, value=("  " * level) + label)
        if bold_row:
            cell.font = bold
        for ci, f in enumerate(FIELDS, 2):
            c = ws.cell(row=r, column=ci, value=values.get(f) or None)
            if bold_row:
                c.font = bold
        if fill:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = fill

    r = HR + 1
    grand = {f: 0.0 for f in FIELDS}
    for dept in sorted(dept_rows):
        rows = dept_rows[dept]
        dept_total = dept_totals.get(dept) or sum_rows(rows)
        emit_row(r, dept, dept_total, 0, fill=fill_dept, bold_row=True)
        r += 1
        for lvl_row in rows:
            emit_row(r, str(lvl_row.get("level") or ""), lvl_row, 1)
            r += 1
        for f in FIELDS:
            grand[f] += float(dept_total.get(f) or 0)

    emit_row(r, "Total", grand, 0, fill=fill_total, bold_row=True)
    r += 1
    emit_row(r, "Permanent Employees", {
        "prev_total": grand["prev_permanent"], "curr_total": grand["curr_permanent"], "inc_total": grand["inc_permanent"],
    }, 1, fill=fill_total)
    r += 1
    emit_row(r, "Temporary Employees", {
        "prev_total": grand["prev_temporary"], "curr_total": grand["curr_temporary"], "inc_total": grand["inc_temporary"],
    }, 1, fill=fill_total)

    ws.freeze_panes = ws.cell(row=HR + 1, column=2)
    ws.column_dimensions["A"].width = 30
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Personnel_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_opex_plan_report_xlsx(plans: list, plan_year: int) -> StreamingResponse:
    """Format reference: user-provided "Operating Expense" summary
    screenshot (Expense | Estimation PL {YY} (total)-FS), sourced from
    OPEX Plan Simulation Data. Grouped by Managerial Account Name (row's
    own budget category, e.g. "Salary & Allowance") — a category expands
    into its Chart of Account Name sub-lines only when more than one
    distinct COA exists under it, matching the screenshot's "Sales &
    Marketing activity" -> Seminar & Event / Sponsorship / Entertainment
    breakdown; every other category in the screenshot has just one COA
    line, so it stays flat. Category (and sub-line) order follows first-
    appearance order in the uploaded data, not alphabetical — the source
    template's own category ordering is business-meaningful (matches the
    screenshot's ordering), alphabetizing it would scramble that. Sums
    every department/team plan for the year; the Sales & Mkt/Strategy
    Development/Plant/Admin columns in the input are cost-center
    ALLOCATION markers, not additional amounts, so only each row's own
    Total (Jan-Dec already summed) is used."""
    groups: dict = {}  # ma_name -> {coa_name: total}, insertion-ordered
    for plan in plans:
        for row in (plan.get("content") or {}).get("rows") or []:
            if len(row) < 23:
                continue
            ma_name = str(row[2] or "").strip()
            if not ma_name:
                continue
            coa_name = str(row[4] or "").strip() or ma_name
            total = float(row[22] or 0)
            g = groups.setdefault(ma_name, {})
            g[coa_name] = g.get(coa_name, 0.0) + total

    if not groups:
        raise HTTPException(404, f"Tidak ada data OPEX Plan untuk tahun {plan_year}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"OPEX_Plan_{plan_year}"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    fill_total = PatternFill("solid", fgColor="D9C6F2")
    sub_font = Font(italic=True, color="4472C4")

    yy = str(plan_year)[2:].zfill(2)
    headers = ["Expense", f"Estimation PL {yy} (total)-FS"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Operating Expense — {plan_year} — from Simulation Data > OPEX Plan").font = Font(italic=True, size=10)

    HR = 4
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=HR, column=c, value=label)
        cell.font, cell.alignment, cell.fill = bold, center, fill_hdr

    r = HR + 1
    grand_total = sum(sum(coa.values()) for coa in groups.values())
    ws.cell(row=r, column=1, value="Operating Expense").font = bold
    ws.cell(row=r, column=2, value=grand_total or None).font = bold
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = fill_total
    r += 1

    for ma_name, coa_totals in groups.items():
        ma_total = sum(coa_totals.values())
        ws.cell(row=r, column=1, value="  " + ma_name)
        ws.cell(row=r, column=2, value=ma_total or None)
        r += 1
        if len(coa_totals) > 1:
            for coa_name, coa_total in coa_totals.items():
                c1 = ws.cell(row=r, column=1, value="    " + coa_name)
                c1.font = sub_font
                c2 = ws.cell(row=r, column=2, value=coa_total or None)
                c2.font = sub_font
                r += 1

    ws.freeze_panes = ws.cell(row=HR + 1, column=2)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"OPEX_Plan_Report_{plan_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Manufacture Plan ────────────────────────────────────────────────────────────

class ManufacturePlanPayload(BaseModel):
    id:          Optional[int]  = None
    plan_year:   int
    department:  str            = ""
    team_code:   str            = ""
    team_name:   str            = ""
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/manufacture-plans")
async def list_manufacture_plans(
    plan_year:  Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    team_code:  Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List manufacture plans filtered by year/department/team."""
    return await ManufacturePlanService().list_manufacture_plans(db, plan_year, department, team_code)


@router.get("/manufacture-plans/detail-report")
async def export_manufacture_plan_detail_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Export the Manufacturing Plan detail report — same column layout and
    calculations as sumber/'BOM, Manufacturing plan.xlsx' sheet
    'Detail Manufacturing plan_All', built from every Manufacture Plan
    record for the given year (2026 scope only — no 2027 H1 columns, since
    this app has no 2027 planning data source to build them from).
    Registered before /manufacture-plans/{plan_id} for the same reason as
    /sales-plans/gross-sales-report — otherwise Starlette's path matching
    would swallow this route as plan_id="detail-report"."""
    result = await ManufacturePlanService().list_manufacture_plans(db, plan_year=plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Manufacture Plan data")
    rows = [row for plan in result["data"] for row in (plan.get("content") or {}).get("rows", [])]
    if not rows:
        raise HTTPException(404, f"Tidak ada data Manufacture Plan untuk tahun {plan_year}")
    try:
        return _build_manufacture_plan_detail_xlsx(rows, plan_year)
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {e}")


@router.get("/manufacture-plans/report/export")
async def export_manufacture_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download — computed live from Simulation Data > Manufacture
    Plan, no separate upload. See ManufacturePlanService.get_report for
    the Local/CMO/Export and Liquid/Freeze Dry classification this infers
    from that input data. Registered before /manufacture-plans/{plan_id}
    for the same reason as /manufacture-plans/detail-report above."""
    result = await ManufacturePlanService().get_report(db, plan_year)
    if not result.get("success"):
        raise HTTPException(404, result.get("error") or "No report data available")
    try:
        return _build_manufacture_plan_report_xlsx(result, plan_year)
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {e}")


@router.get("/manufacture-plans/{plan_id}")
async def get_manufacture_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single manufacture plan."""
    return await ManufacturePlanService().get_manufacture_plan(db, plan_id)


@router.post("/manufacture-plans")
async def upsert_manufacture_plan(
    body: ManufacturePlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update manufacture plan."""
    return await ManufacturePlanService().upsert_manufacture_plan(db, body.model_dump(), user.username)


@router.delete("/manufacture-plans/{plan_id}")
async def delete_manufacture_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete manufacture plan."""
    return await ManufacturePlanService().delete_manufacture_plan(db, plan_id)


@router.post("/manufacture-plans/upload")
async def upload_manufacture_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import Manufacture Plan data from an Excel file matching the
    "manufacture plan template.xlsx" layout."""
    content = await file.read()
    return await ManufacturePlanService().import_excel(db, content, plan_year, user.username)


# ── Investment Plan ──────────────────────────────────────────────────────────────

class InvestmentPlanPayload(BaseModel):
    id:          Optional[int]  = None
    plan_year:   int
    department:  str            = ""
    team_code:   str            = ""
    team_name:   str            = ""
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/investment-plans")
async def list_investment_plans(
    plan_year:  Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    team_code:  Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List investment plans filtered by year/department/team."""
    return await InvestmentPlanService().list_investment_plans(db, plan_year, department, team_code)


@router.get("/investment-plans/report/export")
async def export_investment_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download of Investment Plan Simulation Data, formatted like
    "5. Investment Plan" in the Business Plan Report workbook — grouped by
    Classification, Q1-Q4 + Total. See _build_investment_plan_report_xlsx
    for what's deliberately left out (Acquisition/Depreciation columns).
    Registered before /investment-plans/{plan_id} for the same reason as
    /manufacture-plans/detail-report."""
    result = await InvestmentPlanService().list_investment_plans(db, plan_year=plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load Investment Plan data")
    return _build_investment_plan_report_xlsx(result["data"], plan_year)


@router.get("/investment-plans/{plan_id}")
async def get_investment_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single investment plan."""
    return await InvestmentPlanService().get_investment_plan(db, plan_id)


@router.post("/investment-plans")
async def upsert_investment_plan(
    body: InvestmentPlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update investment plan."""
    return await InvestmentPlanService().upsert_investment_plan(db, body.model_dump(), user.username)


@router.delete("/investment-plans/{plan_id}")
async def delete_investment_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete investment plan."""
    return await InvestmentPlanService().delete_investment_plan(db, plan_id)


@router.post("/investment-plans/upload")
async def upload_investment_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import Investment Plan data from an Excel file matching the
    "investment plan template.xlsx" layout."""
    content = await file.read()
    return await InvestmentPlanService().import_excel(db, content, plan_year, user.username)


# ── OPEX Plan ─────────────────────────────────────────────────────────────────────

class OpexPlanPayload(BaseModel):
    id:          Optional[int]  = None
    plan_year:   int
    department:  str            = ""
    team_code:   str            = ""
    team_name:   str            = ""
    content:     dict           = {}
    status:      Optional[str]  = "draft"


@router.get("/opex-plans")
async def list_opex_plans(
    plan_year:  Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    team_code:  Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """List OPEX plans filtered by year/department/team."""
    return await OpexPlanService().list_opex_plans(db, plan_year, department, team_code)


@router.get("/opex-plans/report/export")
async def export_opex_plan_report(
    plan_year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Excel download of OPEX Plan Simulation Data, formatted as an
    "Operating Expense" summary (Expense | Estimation PL {YY} (total)-FS).
    See _build_opex_plan_report_xlsx for the Managerial-Account/Chart-of-
    Account grouping rule. Registered before /opex-plans/{plan_id} for the
    same reason as /manufacture-plans/detail-report."""
    result = await OpexPlanService().list_opex_plans(db, plan_year=plan_year)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to load OPEX Plan data")
    return _build_opex_plan_report_xlsx(result["data"], plan_year)


@router.get("/opex-plans/{plan_id}")
async def get_opex_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Get single OPEX plan."""
    return await OpexPlanService().get_opex_plan(db, plan_id)


@router.post("/opex-plans")
async def upsert_opex_plan(
    body: OpexPlanPayload,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Create / update OPEX plan."""
    return await OpexPlanService().upsert_opex_plan(db, body.model_dump(), user.username)


@router.delete("/opex-plans/{plan_id}")
async def delete_opex_plan(
    plan_id: int,
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Delete OPEX plan."""
    return await OpexPlanService().delete_opex_plan(db, plan_id)


@router.post("/opex-plans/upload")
async def upload_opex_plan_excel(
    plan_year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Import OPEX Plan data from an Excel file matching the
    "(O1) OPEX Plan_Summary_Department.xlsx" layout."""
    content = await file.read()
    return await OpexPlanService().import_excel(db, content, plan_year, user.username)


# ── Exchange Rates ─────────────────────────────────────────────────────────────

@router.get("/exchange-rates")
async def get_exchange_rates(
    refresh: bool = Query(False, description="Force re-scrape even if cache is fresh"),
    source: str = Query("auto", description="Data source: auto, bi_html, exchangerate_api, frankfurter"),
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """Kurs — multi-source fallback: Bank Indonesia, ExchangeRate-API, Frankfurter."""
    return await asyncio.to_thread(exchange_rate_service.get_rates, source, refresh)


class PushToEBSRequest(BaseModel):
    rate_date:   str        # "2026-07-03"
    rate_type:   str        = "Corporate"       # Corporate | Spot | user-defined
    rate_source: str        = "tengah"          # jual | beli | tengah
    currencies:  list[str]  = ["USD", "EUR", "SGD", "JPY", "GBP", "AUD", "CNY", "MYR"]


@router.post("/exchange-rates/push-to-ebs")
async def push_exchange_rates_to_ebs(
    body: PushToEBSRequest,
    user: CurrentUser = Depends(require_role(Roles.PAC)),
):
    """
    Push Kurs Transaksi BI ke Oracle EBS GL Daily Rates menggunakan
    GL_DAILY_RATES_API.INSERT_RATE / UPDATE_RATE.
    """
    # Get current cached rates (don't re-scrape unless cache is empty)
    cached = await asyncio.to_thread(exchange_rate_service.get_rates, "auto", False)
    if not cached.get("rates"):
        return {"success": False, "error": "Tidak ada data kurs — ambil data kurs terlebih dahulu", "results": []}

    results = await asyncio.to_thread(
        exchange_rate_service.push_rates_to_ebs,
        cached["rates"],
        body.rate_date,
        body.rate_type,
        body.rate_source,
        body.currencies,
    )

    success_count = sum(1 for r in results if r["status"] == "success")
    error_count   = sum(1 for r in results if r["status"] == "error")
    return {
        "success":       error_count == 0,
        "rate_date":     body.rate_date,
        "rate_type":     body.rate_type,
        "rate_source":   body.rate_source,
        "total":         len(results),
        "success_count": success_count,
        "error_count":   error_count,
        "results":       results,
    }
