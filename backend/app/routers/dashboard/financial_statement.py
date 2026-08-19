"""
Financial Statement Router
Route prefix : /api/v1/dashboard/accounting/financial-statement
Required role: accounting_staff OR admin

Balance Sheet / Profit & Loss reporting sourced live from Oracle EBS 12.2.8
GL_BALANCES — see app/services/financial_statement_service.py for the
account-mapping and balance-formula documentation.

Endpoints:
  GET  /periods                     — GL periods available for the period selectors
  GET  /balance-sheet               — line-item Balance Sheet, one column per period
  GET  /balance-sheet/export        — same, as .xlsx (format matches FS_CKD OTTO 2015-2026_sent.xlsx)
  GET  /balance-sheet-detail        — natural-account-level Balance Sheet drill-down
  GET  /balance-sheet-detail/export — same, as .xlsx
  GET  /profit-loss                 — line-item P&L, one column per fiscal year/YTD range
  GET  /profit-loss/export          — same, as .xlsx
  GET  /profit-loss-monthly         — MTD/YTD this year vs same period last year
  GET  /profit-loss-monthly/export  — same, as .xlsx
  GET  /cash-flow                   — Cash Flow statement, Excel-only (no Oracle equivalent)
"""
import io
import json
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import require_role, CurrentUser, Roles
from app.database import get_db
from app.services.financial_statement_service import FinancialStatementService
from app.services.financial_statement_upload_service import FinancialStatementUploadService

router = APIRouter()


@router.get("/periods")
async def get_periods(user: CurrentUser = Depends(require_role(Roles.ACCOUNTING))):
    """List GL periods for the primary ledger, with a has_activity flag so
    the frontend can default to the latest period with posted balances."""
    return await FinancialStatementService().get_periods()


# ── Excel upload source (transition from manual Excel reporting to Oracle) ──
# One uploaded snapshot per report_type, replacing the live Oracle query
# when the frontend's source toggle is set to "excel". See
# financial_statement_upload_service.py for the parser / storage shape.

@router.get("/upload-status")
async def get_upload_status(
    report_type: str = Query(..., pattern="^(balance_sheet|profit_loss|profit_loss_monthly|cash_flow)$"),
    month: Optional[int] = Query(None, ge=1, le=12, description="profit_loss_monthly only — which stored snapshot; omit for the most recent"),
    year:  Optional[int] = Query(None, description="profit_loss_monthly only — which stored snapshot; omit for the most recent"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Metadata about a stored Excel upload for a report, if any — lets the
    frontend show "last uploaded by X on Y" and restrict its year pickers
    to whatever years the uploaded file actually covers. For
    profit_loss_monthly (which stores one snapshot per month, unlike every
    other report_type's single row), month/year selects which one; omitted
    defaults to the most recently uploaded."""
    upload = await FinancialStatementUploadService().get_upload(db, report_type, month, year)
    if not upload:
        return {"success": True, "data": None}
    content = upload["content"]
    return {
        "success": True,
        "data": {
            "original_filename": upload["original_filename"],
            "uploaded_by": upload["uploaded_by"],
            "uploaded_at": upload["uploaded_at"],
            "years": content.get("years"),
            "as_of_label": content.get("as_of_label"),
            "date_last": content.get("date_last"),
            "date_this": content.get("date_this"),
            "period_month": upload.get("period_month"),
            "period_year": upload.get("period_year"),
        },
    }


@router.get("/profit-loss-monthly/snapshots")
async def get_profit_loss_monthly_snapshots(
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Every stored Profit or Loss Monthly snapshot (month, year), newest
    first — populates the Month+Year picker with only the periods that
    actually have an uploaded file, instead of offering all 12 months."""
    snapshots = await FinancialStatementUploadService().list_snapshots(db, "profit_loss_monthly")
    return {"success": True, "data": snapshots}


@router.post("/upload")
async def upload_financial_statement_excel(
    report_type: str = Query(..., pattern="^(balance_sheet|profit_loss|profit_loss_monthly|cash_flow)$"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Upload the manual Excel report for one of the 4 supported report
    types — parses the matching sheet (Balance sheet / Profit or loss /
    PL_monthly / Cashflow) and stores it. profit_loss_monthly is stored as
    its own (month, year) snapshot, additive to whatever other months are
    already stored; every other report_type still replaces its single
    previous upload wholesale."""
    content = await file.read()
    result = await FinancialStatementUploadService().save_upload(db, report_type, content, file.filename, user.username)
    if not result.get("success"):
        raise HTTPException(400, result.get("error") or "Failed to process the file")
    return {"success": True, "uploaded_at": result.get("uploaded_at")}


async def _balance_sheet_from_excel(db: AsyncSession, periods_csv: str) -> dict:
    upload = await FinancialStatementUploadService().get_upload(db, "balance_sheet")
    if not upload:
        return {"success": False, "error": "No Excel data uploaded yet for Balance Sheet."}
    content = upload["content"]
    all_years = content["years"]
    years = [int(p.strip()) for p in periods_csv.split(",") if p.strip().isdigit() and int(p.strip()) in all_years]
    if not years:
        return {"success": False, "error": "The requested year isn't in the uploaded Excel data."}
    idx = [all_years.index(y) for y in years]
    latest_year = max(all_years)

    def pick(arr):
        return [arr[i] for i in idx]

    def pick_rows(rows):
        return [{"label": r["label"], "values": pick(r["values"])} for r in rows]

    def col_label(y):
        # The uploaded file's newest column is whatever mid-year snapshot
        # date it was saved at (e.g. "June 30, 2026"), not a Dec 31
        # close -- closed prior years genuinely are Dec 31 snapshots.
        if y == latest_year and content.get("as_of_label"):
            return content["as_of_label"]
        return f"Dec {y}"

    return {
        "success": True,
        "periods": [str(y) for y in years],
        "column_labels": [col_label(y) for y in years],
        "current_assets": pick_rows(content["current_assets"]), "total_current_assets": pick(content["total_current_assets"]),
        "noncurrent_assets": pick_rows(content["noncurrent_assets"]), "total_noncurrent_assets": pick(content["total_noncurrent_assets"]),
        "total_assets": pick(content["total_assets"]),
        "current_liabilities": pick_rows(content["current_liabilities"]), "total_current_liabilities": pick(content["total_current_liabilities"]),
        "noncurrent_liabilities": pick_rows(content["noncurrent_liabilities"]), "total_noncurrent_liabilities": pick(content["total_noncurrent_liabilities"]),
        "total_liabilities": pick(content["total_liabilities"]),
        "equity": pick_rows(content["equity"]), "total_equity": pick(content["total_equity"]),
        "total_liabilities_and_equity": pick(content["total_liabilities_and_equity"]),
        "check_diff": pick(content["check_diff"]),
        "unmapped_accounts": content["unmapped_accounts"],
    }


async def _profit_loss_from_excel(db: AsyncSession, years_csv: str) -> dict:
    upload = await FinancialStatementUploadService().get_upload(db, "profit_loss")
    if not upload:
        return {"success": False, "error": "No Excel data uploaded yet for Profit or Loss."}
    content = upload["content"]
    all_years = content["years"]
    years = [int(y.strip()) for y in years_csv.split(",") if y.strip().isdigit() and int(y.strip()) in all_years]
    if not years:
        return {"success": False, "error": "The requested year isn't in the uploaded Excel data."}
    idx = [all_years.index(y) for y in years]

    def pick(arr):
        return [arr[i] for i in idx]

    def pick_rows(rows):
        return [{"label": r["label"], "values": pick(r["values"])} for r in rows]

    return {
        "success": True, "columns": [f"FY {y}" for y in years],
        "sales_lines": pick_rows(content["sales_lines"]), "contra_lines": [], "total_net_sales": pick(content["total_net_sales"]),
        "cogs_lines": pick_rows(content["cogs_lines"]), "total_cogs": pick(content["total_cogs"]),
        "gross_profit": pick(content["gross_profit"]),
        "expense_lines": pick_rows(content["expense_lines"]), "total_expenses": pick(content["total_expenses"]),
        "other_lines": pick_rows(content["other_lines"]), "total_other": pick(content["total_other"]),
        "profit_before_tax": pick(content["profit_before_tax"]),
        "tax_lines": pick_rows(content["tax_lines"]), "total_tax": pick(content["total_tax"]),
        "profit_after_tax": pick(content["profit_after_tax"]),
        "oci": pick(content["oci"]), "total_comprehensive": pick(content["total_comprehensive"]),
        "unmapped_accounts": [],
    }


async def _cash_flow_from_excel(db: AsyncSession, years_csv: str) -> dict:
    upload = await FinancialStatementUploadService().get_upload(db, "cash_flow")
    if not upload:
        return {"success": False, "error": "No Excel data uploaded yet for Cash Flow."}
    content = upload["content"]
    all_years = content["years"]
    years = [int(y.strip()) for y in years_csv.split(",") if y.strip().isdigit() and int(y.strip()) in all_years]
    if not years:
        return {"success": False, "error": "The requested year isn't in the uploaded Excel data."}
    idx = [all_years.index(y) for y in years]

    def col_label(y):
        # The in-progress year's column is a partial-year figure (through
        # whatever month was last posted), not a full Dec 31 close.
        if y == content.get("partial_year") and content.get("as_of_label"):
            return content["as_of_label"]
        return str(y)

    return {
        "success": True,
        "columns": [col_label(y) for y in years],
        "rows": [{**r, "values": [r["values"][i] for i in idx]} for r in content["rows"]],
    }


async def _profit_loss_monthly_from_excel(db: AsyncSession, month: Optional[int] = None, year: Optional[int] = None) -> dict:
    upload = await FinancialStatementUploadService().get_upload(db, "profit_loss_monthly", month, year)
    if not upload:
        msg = "No Excel data uploaded yet for Profit or Loss Monthly."
        if month and year:
            msg = f"No Profit or Loss Monthly snapshot uploaded for {month}/{year}."
        return {"success": False, "error": msg}
    result = dict(upload["content"])
    result["success"] = True
    result.setdefault("unmapped_accounts", [])
    result["period_month"] = upload.get("period_month")
    result["period_year"] = upload.get("period_year")
    return result


@router.get("/balance-sheet")
async def get_balance_sheet(
    periods: str = Query(..., description="Comma-separated GL period names (oracle) or fiscal years (excel)"),
    source: str = Query("oracle", pattern="^(oracle|excel)$"),
    account_group: Optional[str] = Query(None, pattern="^(ASSETS|LIABILITIES|EQUITY)$",
                                          description="Narrow to one section — omit for All (Oracle source only)"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Balance Sheet grouped into line items — one column per period."""
    if source == "excel":
        return await _balance_sheet_from_excel(db, periods)
    period_list = [p.strip() for p in periods.split(",") if p.strip()]
    return await FinancialStatementService().get_balance_sheet(period_list, account_group)


@router.get("/balance-sheet-detail")
async def get_balance_sheet_detail(
    periods: str = Query(..., description="Comma-separated GL period names"),
    account_group: Optional[str] = Query(None, pattern="^(ASSETS|LIABILITIES|EQUITY)$",
                                          description="Narrow to one section — omit for All"),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Balance Sheet at natural-account granularity — drill-down view.
    Oracle-only: the manual Excel report has no natural-account detail to
    drill into, only the same bucketed line items the summary view shows."""
    period_list = [p.strip() for p in periods.split(",") if p.strip()]
    return await FinancialStatementService().get_balance_sheet_detail(period_list, account_group)


@router.get("/profit-loss")
async def get_profit_and_loss(
    columns: Optional[str] = Query(None, description='JSON list of {"label","periods":[...]} — oracle source only'),
    source: str = Query("oracle", pattern="^(oracle|excel)$"),
    years: Optional[str] = Query(None, description="Comma-separated fiscal years — excel source only"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """P&L grouped into line items — one column per fiscal year / YTD range."""
    if source == "excel":
        return await _profit_loss_from_excel(db, years or "")
    col_list = json.loads(columns)
    return await FinancialStatementService().get_profit_and_loss(col_list)


@router.get("/profit-loss-monthly")
async def get_profit_and_loss_monthly(
    period_this: Optional[str] = Query(None, description="This year's MTD period, e.g. JUN-26 — oracle source only"),
    ytd_this: Optional[str] = Query(None, description="Comma-separated periods JAN-26,...,JUN-26 — oracle source only"),
    period_last: Optional[str] = Query(None, description="Same month last year, e.g. JUN-25 — oracle source only"),
    ytd_last: Optional[str] = Query(None, description="Comma-separated periods JAN-25,...,JUN-25 — oracle source only"),
    source: str = Query("oracle", pattern="^(oracle|excel)$"),
    month: Optional[int] = Query(None, ge=1, le=12, description="excel source only — which stored snapshot; omit for the most recent"),
    year:  Optional[int] = Query(None, description="excel source only — which stored snapshot; omit for the most recent"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """MTD/YTD comparison — this year's period vs the same period last year."""
    if source == "excel":
        return await _profit_loss_monthly_from_excel(db, month, year)
    ytd_this_list = [p.strip() for p in ytd_this.split(",") if p.strip()]
    ytd_last_list = [p.strip() for p in ytd_last.split(",") if p.strip()]
    return await FinancialStatementService().get_profit_and_loss_monthly(
        period_this, ytd_this_list, period_last, ytd_last_list
    )


@router.get("/cash-flow")
async def get_cash_flow(
    years: str = Query(..., description="Comma-separated fiscal years"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """Cash Flow statement — Excel-only. There's no live Oracle equivalent:
    a statutory cash flow isn't a direct GL_BALANCES query, it's manually
    prepared/derived from the other statements each period."""
    return await _cash_flow_from_excel(db, years)


@router.get("/cash-flow/export")
async def export_cash_flow(
    years: str = Query(..., description="Comma-separated fiscal years"),
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    data = await _cash_flow_from_excel(db, years)
    if not data.get("success"):
        raise HTTPException(400, data.get("error") or "Failed to export")
    return _build_cash_flow_xlsx(data["rows"], data["columns"])


# ── Excel exports — layout mirrors FS_CKD OTTO 2015-2026_sent.xlsx ────────────
# (title block, dark-navy header row, indented hierarchical line items, bold
# light-blue total rows, accounting number format with parentheses for
# negatives) — that's the format management is used to seeing, per the user.

TITLE_FONT   = Font(bold=True, size=18)
SUB_FONT     = Font(bold=True, size=11)
HEADER_FONT  = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
TOTAL_FONT   = Font(bold=True, size=10)
TOTAL_FILL   = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=10)
LABEL_FONT   = Font(size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
ACC_NUMFMT   = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'


def _title_block(ws, title2: str, date_label: str, last_col: int):
    ws.cell(row=1, column=1, value="PT CKD OTTO Pharmaceuticals").font = TITLE_FONT
    ws.cell(row=2, column=1, value=title2).font = TITLE_FONT
    ws.cell(row=3, column=1, value=date_label).font = SUB_FONT
    ws.cell(row=4, column=1, value="Amount in IDR").font = SUB_FONT


def _header_row(ws, row: int, columns: list, last_col: int):
    c = ws.cell(row=row, column=1, value="ACCOUNT")
    c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for i, label in enumerate(columns, start=2):
        c = ws.cell(row=row, column=i, value=label)
        c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for c_idx in range(1, last_col + 1):
        ws.cell(row=row, column=c_idx).fill = HEADER_FILL


def _write_row(ws, row: int, label: str, values: list, level: int = 0, bold: bool = False, fill=None):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = TOTAL_FONT if bold else LABEL_FONT
    lc.alignment = Alignment(indent=level)
    if fill:
        lc.fill = fill
    for i, v in enumerate(values, start=2):
        vc = ws.cell(row=row, column=i, value=v)
        vc.number_format = ACC_NUMFMT
        vc.font = TOTAL_FONT if bold else LABEL_FONT
        if fill:
            vc.fill = fill


def _autosize(ws, ncols: int, first_width: int = 34, other_width: int = 15):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = first_width
    for i in range(2, ncols + 2):
        ws.column_dimensions[get_column_letter(i)].width = other_width
    ws.freeze_panes = "B7"


def _stream(wb, fname: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _bs_sections(data: dict) -> list:
    """(section_label, line_rows, (total_label, total_values), indent_level)"""
    return [
        ("ASSETS", None, None, 0),
        ("CURRENT ASSETS", data["current_assets"], ("TOTAL CURRENT ASSETS", data["total_current_assets"]), 1),
        ("NON CURRENT ASSET", data["noncurrent_assets"], ("TOTAL NON CURRENT ASSETS", data["total_noncurrent_assets"]), 1),
        (None, None, ("TOTAL ASSETS", data["total_assets"]), 0),
        ("LIABILITIES", None, None, 0),
        ("CURRENT LIABILITIES", data["current_liabilities"], ("TOTAL CURRENT LIABILITIES", data["total_current_liabilities"]), 1),
        ("NONCURRENT LIABILITIES", data["noncurrent_liabilities"], ("TOTAL NONCURRENT LIABILITIES", data["total_noncurrent_liabilities"]), 1),
        (None, None, ("TOTAL  LIABILITIES", data["total_liabilities"]), 0),
        ("EQUITY", data["equity"], ("TOTAL  EQUITY", data["total_equity"]), 0),
        (None, None, ("TOTAL  LIABILITIES AND EQUITY", data["total_liabilities_and_equity"]), 0),
    ]


def _build_balance_sheet_xlsx(data: dict, column_labels: list, as_of_label: str, detail: bool) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet Detail" if detail else "Balance Sheet"
    ncols = len(column_labels)
    _title_block(ws, "Balance Sheet Detail" if detail else "Balance Sheet", as_of_label, ncols + 1)
    _header_row(ws, 6, column_labels, ncols + 1)

    row = 7
    if detail:
        by_type = {"A": [], "L": [], "O": []}
        for a in data["accounts"]:
            by_type.setdefault(a["account_type"], []).append(a)
        for section_label, key in [("ASSETS", "A"), ("LIABILITIES", "L"), ("EQUITY", "O")]:
            _write_row(ws, row, section_label, [], level=0, bold=True)
            row += 1
            for acc in by_type.get(key, []):
                _write_row(ws, row, f"{acc['account_code']} — {acc['account_desc'] or acc['line_item']}", acc["values"], level=1)
                row += 1
            row += 1
    else:
        for section_label, rows_, total_, level in _bs_sections(data):
            if section_label:
                _write_row(ws, row, section_label, [], level=level, bold=True)
                row += 1
            if rows_:
                for line in rows_:
                    _write_row(ws, row, line["label"], line["values"], level=level + 1)
                    row += 1
            if total_:
                _write_row(ws, row, total_[0], total_[1], level=level, bold=True, fill=TOTAL_FILL)
                row += 1

    _autosize(ws, ncols)
    fname = f"Balance_Sheet{'_Detail' if detail else ''}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _stream(wb, fname)


def _build_cash_flow_xlsx(rows: list, column_labels: list) -> StreamingResponse:
    """Cash Flow's rows are already flat (label/level/type/values) — no
    section grouping to unpack, unlike Balance Sheet/P&L, so this just
    writes them straight through in file order."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    ncols = len(column_labels)
    _title_block(ws, "Cash Flow", datetime.now().strftime("%B %d, %Y"), ncols + 1)
    _header_row(ws, 6, column_labels, ncols + 1)

    row = 7
    for r in rows:
        bold = r.get("type") == "total"
        _write_row(ws, row, r["label"], r.get("values") or [], level=r.get("level", 0), bold=bold, fill=TOTAL_FILL if bold else None)
        row += 1

    _autosize(ws, ncols)
    fname = f"Cash_Flow_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _stream(wb, fname)


def _pl_sections(data: dict) -> list:
    return [
        ("NET SALES", data["sales_lines"] + data["contra_lines"], ("TOTAL NET SALES", data["total_net_sales"])),
        ("COGS", data["cogs_lines"], ("TOTAL COGS", data["total_cogs"])),
        (None, None, ("GROSS PROFIT", data["gross_profit"])),
        ("EXPENSES", data["expense_lines"], ("TOTAL EXPENSES", data["total_expenses"])),
        ("OTHER INCOME / EXPENSES", data["other_lines"], ("TOTAL OTHER INCOME (EXPENSES)", data["total_other"])),
        (None, None, ("PROFIT (LOSS) BEFORE TAX", data["profit_before_tax"])),
        ("INCOME TAX", data["tax_lines"], ("TOTAL INCOME TAX BENEFIT (EXPENSE)", data["total_tax"])),
        (None, None, ("PROFIT (LOSS) AFTER TAX", data["profit_after_tax"])),
        (None, [{"label": "OTHER COMPREHENSIVE INCOME", "values": data["oci"]}],
         ("TOTAL COMPREHENSIVE INCOME (LOSS) FOR THE YEAR", data["total_comprehensive"])),
    ]


def _build_pl_xlsx(data: dict, title2: str, date_label: str, sheet_title: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    column_labels = data["columns"]
    ncols = len(column_labels)
    _title_block(ws, title2, date_label, ncols + 1)
    _header_row(ws, 6, column_labels, ncols + 1)

    row = 7
    for section_label, rows_, total_ in _pl_sections(data):
        if section_label:
            _write_row(ws, row, section_label, [], level=0, bold=True)
            row += 1
        if rows_:
            for line in rows_:
                _write_row(ws, row, line["label"], line["values"], level=1)
                row += 1
        if total_:
            _write_row(ws, row, total_[0], total_[1], level=0, bold=True, fill=TOTAL_FILL)
            row += 1
        row += 1

    _autosize(ws, ncols)
    fname = f"{sheet_title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _stream(wb, fname)


@router.get("/balance-sheet/export")
async def export_balance_sheet(
    periods: str = Query(...),
    as_of_label: str = Query("", description="Display label for the report date line, e.g. 'June 30, 2026'"),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    period_list = [p.strip() for p in periods.split(",") if p.strip()]
    data = await FinancialStatementService().get_balance_sheet(period_list)
    labels = [FinancialStatementService.period_display_label(p) for p in period_list]
    return _build_balance_sheet_xlsx(data, labels, as_of_label or labels[-1], detail=False)


@router.get("/balance-sheet-detail/export")
async def export_balance_sheet_detail(
    periods: str = Query(...),
    as_of_label: str = Query(""),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    period_list = [p.strip() for p in periods.split(",") if p.strip()]
    data = await FinancialStatementService().get_balance_sheet_detail(period_list)
    labels = [FinancialStatementService.period_display_label(p) for p in period_list]
    return _build_balance_sheet_xlsx(data, labels, as_of_label or labels[-1], detail=True)


@router.get("/profit-loss/export")
async def export_profit_and_loss(
    columns: str = Query(...),
    date_label: str = Query(""),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    col_list = json.loads(columns)
    data = await FinancialStatementService().get_profit_and_loss(col_list)
    return _build_pl_xlsx(data, "Profit or Loss", date_label or datetime.now().strftime("%B %d, %Y"), "Profit or Loss")


@router.get("/profit-loss-monthly/export")
async def export_profit_and_loss_monthly(
    period_this: str = Query(...),
    ytd_this: str = Query(...),
    period_last: str = Query(...),
    ytd_last: str = Query(...),
    date_label: str = Query(""),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    ytd_this_list = [p.strip() for p in ytd_this.split(",") if p.strip()]
    ytd_last_list = [p.strip() for p in ytd_last.split(",") if p.strip()]
    data = await FinancialStatementService().get_profit_and_loss_monthly(
        period_this, ytd_this_list, period_last, ytd_last_list
    )
    label = date_label or FinancialStatementService.period_display_label(period_this)
    return _build_pl_xlsx(data, "Profit or Loss", label, "PL Monthly")
