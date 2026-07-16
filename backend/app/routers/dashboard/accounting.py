"""
ACCOUNTING Dashboard Router
Route prefix : /api/v1/dashboard/accounting
Required role: accounting_staff OR admin

Endpoints:
  GET  /summary                        — placeholder
  GET  /material-transactions          — MTL_MATERIAL_TRANSACTIONS export
"""
import io
from calendar import monthrange

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.dependencies import require_role, CurrentUser, Roles
from app.services.accounting_service import AccountingService

router = APIRouter()


@router.get("/summary")
async def get_summary(user: CurrentUser = Depends(require_role(Roles.ACCOUNTING))):
    return {"module": "accounting", "status": "ready"}


# ── AP Outstanding ───────────────────────────────────────────────────────────

@router.get("/ap-outstanding")
async def get_ap_outstanding(
    as_of_date:     str  = Query(None, description="As-of date YYYY-MM-DD (default: today)"),
    supplier_name:  str  = Query(None, description="Partial supplier name filter"),
    operating_unit: str  = Query(None, description="Partial operating unit name filter"),
    payment_status: str  = Query(None, description="Not Paid | Partially Paid | ALL"),
    limit:          int  = Query(500, ge=1, le=2000),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    AP Outstanding — AP_INVOICES_ALL + AP_PAYMENT_SCHEDULES_ALL.
    Excludes fully Paid invoices. As-of date defaults to SYSDATE.
    """
    return await AccountingService().get_ap_outstanding(
        as_of_date, supplier_name, operating_unit, payment_status, limit
    )


# ── AR Outstanding ───────────────────────────────────────────────────────────

@router.get("/ar-outstanding")
async def get_ar_outstanding(
    customer_name:  str  = Query(None, description="Partial customer name filter"),
    invoice_number: str  = Query(None, description="Partial invoice number filter"),
    date_from:      str  = Query(None, description="Invoice date from YYYY-MM-DD"),
    date_to:        str  = Query(None, description="Invoice date to YYYY-MM-DD"),
    status:         str  = Query("OP",  description="OP=open, CL=closed, ALL=both"),
    limit:          int  = Query(500, ge=1, le=2000),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    AR Outstanding from Oracle EBS — AR_PAYMENT_SCHEDULES_ALL + RA_CUSTOMER_TRX_ALL.
    Class filter: INV + DM (invoices and debit memos only).
    """
    return await AccountingService().get_ar_outstanding(
        customer_name, invoice_number, date_from, date_to, status, limit
    )


# ── COGS / Inventory RM PM ───────────────────────────────────────────────────

@router.get("/inventory-rm-pm")
async def get_inventory_rm_pm(
    period:        str  = Query(...,  description="OPM period code e.g. JAN-26"),
    include_begin: bool = Query(True, description="Calculate beginning balance (slower)"),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    Inventory RM PM monthly report — movements + price + beginning/ending balance.
    Fixed: org 121. Data: MTL_MATERIAL_TRANSACTIONS + CKDO_GET_ITEM_COST.
    """
    return await AccountingService().get_inventory_rm_pm(period, include_begin)


@router.get("/inventory-rm-pm/export")
async def export_inventory_rm_pm(
    period:        str  = Query(...,  description="OPM period code e.g. JAN-26"),
    include_begin: bool = Query(True, description="Calculate beginning balance (slower)"),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    Export Inventory RM PM to Excel, laid out to match
    sumber/ouput-inventory RMPM.xlsx (title block, Beginning Balance /
    Price-Qty-Amount movement columns, grouped by material type).
    """
    result = await AccountingService().get_inventory_rm_pm(period, include_begin)
    if not result.get("success"):
        return result

    try:
        return _build_inventory_rm_pm_xlsx(result, period)
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {e}")


def _build_inventory_rm_pm_xlsx(result: dict, period: str) -> StreamingResponse:
    QTY_COLS = AccountingService.QTY_COLS
    AMT_COLS = AccountingService.AMT_COLS
    rows = result["data"]

    MONTHS = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
              "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
    mon_str, yr_str = period.upper().split("-")
    month, year = MONTHS[mon_str], 2000 + int(yr_str)
    _, last_day = monthrange(year, month)
    as_of = f"As of {['January','February','March','April','May','June','July','August','September','October','November','December'][month-1]} {last_day}, {year}"

    N_QTY, N_AMT = len(QTY_COLS), len(AMT_COLS)
    COL_NO, COL_TYPE, COL_CODE, COL_NAME, COL_UOM = 1, 2, 3, 4, 5
    COL_BEG_PRICE, COL_BEG_QTY, COL_BEG_AMT = 6, 7, 8
    COL_PRICE = 9
    COL_QTY_START = 10
    COL_QTY_END = COL_QTY_START + N_QTY               # "QTY Ending" column
    COL_AMT_START = COL_QTY_END + 1
    COL_AMT_END = COL_AMT_START + N_AMT               # "Amount Ending" column
    LAST_COL = COL_AMT_END

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory RM PM"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fillHdr = PatternFill("solid", fgColor="D9E1F2")
    fillGroup = PatternFill("solid", fgColor="F2F2F2")

    def merge(r1, c1, r2, c2, value=None, style=None):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        if style: cell.font, cell.alignment = style
        return cell

    merge(1, 1, 1, LAST_COL, "PT CKD OTTO Pharmaceuticals", (Font(bold=True, size=13), center))
    merge(2, 1, 2, LAST_COL, "Inventory Raw Material_Average Cost", (bold, center))
    merge(3, 1, 3, LAST_COL, as_of, (bold, center))
    merge(4, 1, 4, LAST_COL, "Amount in IDR", (Font(italic=True), center))

    HR1, HR2, HR3 = 6, 7, 8  # 3-tier header rows

    def vmerge_header(col, label):
        merge(HR1, col, HR3, col, label, (bold, center))
        ws.cell(row=HR1, column=col).fill = fillHdr

    vmerge_header(COL_NO, "No")
    vmerge_header(COL_TYPE, "Material Type")
    vmerge_header(COL_CODE, "Item Code")
    vmerge_header(COL_NAME, "Material Name")
    vmerge_header(COL_UOM, "UOM")
    merge(HR1, COL_BEG_PRICE, HR1, COL_BEG_AMT, "Beginning Balance", (bold, center))
    for c, lbl in [(COL_BEG_PRICE, "Price/UOM"), (COL_BEG_QTY, "QTY Ending"), (COL_BEG_AMT, "Amount Ending")]:
        merge(HR2, c, HR3, c, lbl, (bold, center))
        ws.cell(row=HR2, column=c).fill = fillHdr
    vmerge_header(COL_PRICE, "Price/UOM")

    # NOTE: the "Qty"/"Amount" group merges stop one column short of
    # COL_QTY_END/COL_AMT_END — those Ending columns get their own vertical
    # merge via vmerge_header() below, and merged ranges in Excel must not
    # overlap (unlike HTML colSpan/rowSpan, which can visually cover the
    # same column from different cells across rows without conflict).
    merge(HR1, COL_QTY_START, HR1, COL_QTY_END - 1, "Qty", (bold, center))
    merge(HR2, COL_QTY_START, HR2, COL_QTY_START, "In", (bold, center))
    merge(HR2, COL_QTY_START + 1, HR2, COL_QTY_END - 1, "Out", (bold, center))
    vmerge_header(COL_QTY_END, "QTY Ending")
    for i, (_, lbl) in enumerate(QTY_COLS):
        c = ws.cell(row=HR3, column=COL_QTY_START + i, value=lbl)
        c.font, c.alignment = bold, center

    merge(HR1, COL_AMT_START, HR1, COL_AMT_END - 1, "Amount", (bold, center))
    merge(HR2, COL_AMT_START, HR2, COL_AMT_START, "In", (bold, center))
    merge(HR2, COL_AMT_START + 1, HR2, COL_AMT_END - 1, "Out", (bold, center))
    vmerge_header(COL_AMT_END, "Amount Ending")
    for i, (_, lbl) in enumerate(AMT_COLS):
        c = ws.cell(row=HR3, column=COL_AMT_START + i, value=lbl)
        c.font, c.alignment = bold, center

    for c in range(1, LAST_COL + 1):
        ws.cell(row=HR1, column=c).fill = fillHdr
        ws.cell(row=HR2, column=c).fill = fillHdr
        ws.cell(row=HR3, column=c).fill = fillHdr

    NUMFMT = "#,##0.####"

    grouped: dict[str, list] = {}
    for r in rows:
        grouped.setdefault(r["material_type"], []).append(r)

    r_idx = HR3 + 1
    row_no = 0
    for mat_type, mrows in grouped.items():
        ws.cell(row=r_idx, column=1, value=f"{mat_type} — {len(mrows)} items").font = bold
        ws.cell(row=r_idx, column=1).fill = fillGroup
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=LAST_COL)
        r_idx += 1

        totals = {c: 0 for c in ["begin_amount", "end_amount"] + [k for k, _ in QTY_COLS] + [k for k, _ in AMT_COLS]}
        for row in mrows:
            row_no += 1
            ws.cell(row=r_idx, column=COL_NO, value=row_no)
            ws.cell(row=r_idx, column=COL_TYPE, value=row["material_type"])
            ws.cell(row=r_idx, column=COL_CODE, value=row["item_code"])
            ws.cell(row=r_idx, column=COL_NAME, value=row["item_name"])
            ws.cell(row=r_idx, column=COL_UOM, value=row["uom"])
            ws.cell(row=r_idx, column=COL_BEG_PRICE, value=row["unit_price"])
            ws.cell(row=r_idx, column=COL_BEG_QTY, value=row["begin_qty"])
            ws.cell(row=r_idx, column=COL_BEG_AMT, value=row["begin_amount"])
            ws.cell(row=r_idx, column=COL_PRICE, value=row["unit_price"])
            for i, (k, _) in enumerate(QTY_COLS):
                ws.cell(row=r_idx, column=COL_QTY_START + i, value=row[k])
                totals[k] += row[k]
            ws.cell(row=r_idx, column=COL_QTY_END, value=row["end_qty"])
            for i, (k, _) in enumerate(AMT_COLS):
                ws.cell(row=r_idx, column=COL_AMT_START + i, value=row[k])
                totals[k] += row[k]
            ws.cell(row=r_idx, column=COL_AMT_END, value=row["end_amount"])
            totals["begin_amount"] += row["begin_amount"]
            totals["end_amount"]   += row["end_amount"]
            for c in range(COL_BEG_PRICE, LAST_COL + 1):
                ws.cell(row=r_idx, column=c).number_format = NUMFMT
            r_idx += 1

        ws.cell(row=r_idx, column=COL_NAME, value=f"TOTAL {mat_type}").font = bold
        ws.cell(row=r_idx, column=COL_BEG_AMT, value=totals["begin_amount"]).font = bold
        for i, (k, _) in enumerate(QTY_COLS):
            ws.cell(row=r_idx, column=COL_QTY_START + i, value=totals[k]).font = bold
        for i, (k, _) in enumerate(AMT_COLS):
            ws.cell(row=r_idx, column=COL_AMT_START + i, value=totals[k]).font = bold
        ws.cell(row=r_idx, column=COL_AMT_END, value=totals["end_amount"]).font = bold
        for c in range(COL_BEG_PRICE, LAST_COL + 1):
            ws.cell(row=r_idx, column=c).number_format = NUMFMT
        ws.cell(row=r_idx, column=1).fill = fillGroup
        r_idx += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 8
    for c in range(6, LAST_COL + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 13
    ws.freeze_panes = f"{openpyxl.utils.get_column_letter(COL_CODE + 1)}{HR3 + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"inventory_rm_pm_{period.upper()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── COGS / Item Cost Components ───────────────────────────────────────────────

@router.get("/item-cost-components")
async def get_item_cost_components(
    period: str = Query(..., description="OPM period code e.g. JAN-2025"),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    Item cost breakdown per cost component — Oracle OPM CM_CMPT_DTL.
    Fixed org=121 (CKDO), cost_type=1000 (Actual).
    """
    return await AccountingService().get_item_cost_components(period)


# ── COGS / Material Transactions ─────────────────────────────────────────────

@router.get("/material-transactions")
async def get_material_transactions(
    date_from:   str = Query(..., description="Start date YYYY-MM-DD"),
    date_to:     str = Query(..., description="End date   YYYY-MM-DD"),
    org_code:    str = Query(None, description="Filter by organization code"),
    item_number: str = Query(None, description="Filter by item number (partial)"),
    trx_type:    str = Query(None, description="Filter by transaction type (partial)"),
    limit:       int = Query(1000, ge=1, le=5000),
    user: CurrentUser = Depends(require_role(Roles.ACCOUNTING)),
):
    """
    Material Transactions from MTL_MATERIAL_TRANSACTIONS — mirrors
    Oracle EBS Inventory > Material Transactions > Export to Excel.
    """
    return await AccountingService().get_material_transactions(
        date_from, date_to, org_code, item_number, trx_type, limit
    )
