"""
Seed script for Sales Plan from Excel files in sumber/.
Reads (S1) Sales plan_Value_Local_FY 2026_Scenario 2.xlsx and inserts into DB.
"""
import asyncio
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import load_workbook
from app.database import async_engine, Base, get_db
from app.models.sales_plan import SalesPlan

XLSX_PATH = "../../sumber/(S1) Sales plan_Value_Local_FY 2026_Scenario 2.xlsx"


def read_sheet(ws):
    rows = []
    for r in range(16, ws.max_row + 1):
        no = ws.cell(row=r, column=1).value
        product = ws.cell(row=r, column=2).value
        if no is None or product is None:
            continue
        months = [ws.cell(row=r, column=c).value for c in range(4, 16)]  # D-O = Jan-Dec
        total_value = ws.cell(row=r, column=16).value or 0
        total_unit = ws.cell(row=r, column=17).value or 0
        price = ws.cell(row=r, column=18).value
        rows.append({
            "no": int(no),
            "product": str(product),
            "months": [int(v or 0) for v in months],
            "total_value": int(total_value or 0),
            "total_unit": int(total_unit or 0),
            "price": price,
        })
    return rows


async def seed():
    async with async_engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    xlsx_path = os.path.join(os.path.dirname(__file__), XLSX_PATH)
    if not os.path.exists(xlsx_path):
        print(f"Excel not found: {xlsx_path}")
        return

    wb = load_workbook(xlsx_path, data_only=True)
    sheets_to_read = [
        ("Sales plan_V_Total", "Total", ""),
        ("Sales plan_V_National_Public", "National", "Public"),
        ("Sales plan_V_National_Private", "National", "Private"),
        ("Sales plan_V_West_Public", "West", "Public"),
        ("Sales plan_V_West_Private", "West", "Private"),
        ("Sales plan_V_East_Public", "East", "Public"),
        ("Sales plan_V_East_Private", "East", "Private"),
    ]

    async for session in get_db():
        try:
            count = 0
            for sheet_name, area, type_ in sheets_to_read:
                ws = wb[sheet_name]
                rows = read_sheet(ws)
                if not rows:
                    continue

                content = {
                    "headers": ["No", "Product", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total Value", "Total Unit", "Price (Rp)"],
                    "rows": [],
                    "meta": {
                        "area": area,
                        "type": type_,
                        "department": "Sales & Marketing",
                        "team_code": "21",
                        "team_name": "Sales",
                    }
                }

                for row in rows:
                    content["rows"].append([
                        row["no"],
                        row["product"],
                        *row["months"],
                        row["total_value"],
                        row["total_unit"],
                        row["price"] or "",
                    ])

                plan = SalesPlan(
                    plan_year=2026,
                    department="Sales & Marketing",
                    team_code="21",
                    team_name="Sales",
                    plan_type="value",
                    content=content,
                    status="final",
                    created_by="seed",
                )
                session.add(plan)
                count += 1

            await session.commit()
            print(f"Seeded {count} sales plans from Excel.")
        except Exception as e:
            print(f"Seed error: {e}")
            await session.rollback()
            raise
        finally:
            await session.close()
        break


if __name__ == "__main__":
    asyncio.run(seed())
